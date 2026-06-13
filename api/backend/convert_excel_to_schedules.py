"""
Convert Excel schedules to JSON format for blob storage
"""
from openpyxl import load_workbook
import json
from datetime import datetime
import uuid

def parse_shift_time(shift_str):
    """Parse shift time string like '9a-5p' to start/end times"""
    if not shift_str:
        return None, None
    
    shift_str = str(shift_str).upper()
    
    # Handle OFF
    if 'OFF' in shift_str:
        return None, None
    
    # Parse time range
    if '-' not in shift_str:
        return None, None
    
    parts = shift_str.split('-')
    if len(parts) != 2:
        return None, None
    
    start_str, end_str = parts
    
    def parse_time(t):
        t = t.strip()
        if not t:
            return None
        
        # Extract time part (ignore text like "event", "GR", etc.)
        import re
        time_match = re.search(r'(\d{1,2}):?(\d{2})?\s*([AP])M?', t)
        if time_match:
            hour = int(time_match.group(1))
            minute = int(time_match.group(2)) if time_match.group(2) else 0
            ampm = time_match.group(3) if time_match.group(3) else ''
            
            if ampm == 'P' and hour != 12:
                hour += 12
            elif ampm == 'A' and hour == 12:
                hour = 0
            
            return f"{hour:02d}:{minute:02d}"
        
        # Try simple format like "9a" or "5p"
        simple_match = re.search(r'(\d{1,2})([AP])', t)
        if simple_match:
            hour = int(simple_match.group(1))
            ampm = simple_match.group(2)
            
            if ampm == 'P' and hour != 12:
                hour += 12
            elif ampm == 'A' and hour == 12:
                hour = 0
            
            return f"{hour:02d}:00"
        
        return None
    
    start_time = parse_time(start_str)
    end_time = parse_time(end_str)
    
    return start_time, end_time

def parse_location(shift_str):
    """Extract location from shift string"""
    if not shift_str:
        return None
    
    shift_str = str(shift_str).upper()
    
    # Common location patterns
    if 'F6' in shift_str or '6TH' in shift_str:
        return '6th Floor'
    elif 'F2' in shift_str or '2ND' in shift_str:
        return '2nd Floor'
    elif 'GR' in shift_str or 'GROUND' in shift_str:
        return 'Ground Floor'
    elif 'CC' in shift_str or 'CALL CENTER' in shift_str:
        return 'Call Center'
    elif 'WFH' in shift_str:
        return 'Working From Home'
    elif 'EVENT' in shift_str:
        return 'Event'
    
    return None

def convert_excel_to_schedules(excel_path, output_path, employees_path):
    """Convert Excel schedules to JSON format"""
    wb = load_workbook(excel_path)
    
    # Load employees to map names to IDs
    with open(employees_path, 'r') as f:
        employees = json.load(f)
    
    # Create name -> ID mapping
    name_to_id = {}
    for emp in employees:
        name_to_id[emp['name'].lower()] = emp['id']
    
    schedules = []
    
    # Process each sheet (each sheet is a week)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Extract week start date from row 1 (Monday date)
        week_start = None
        for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
            # Find the Monday date (second column)
            for cell in row:
                if cell and isinstance(cell, datetime):
                    week_start = cell.strftime('%Y-%m-%d')
                    break
            if week_start:
                break
        
        if not week_start:
            print(f"Skipping {sheet_name}: no week start date found")
            continue
        
        # Create schedule object
        schedule_id = f"schedule_{week_start}"
        shifts = []
        total_hours = {}
        
        # Parse employee schedules starting from row 3
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            employee_name = row[0]
            if not employee_name or employee_name == 'EVENTS':
                continue
            
            # Clean employee name and map to ID
            employee_name = str(employee_name).strip()
            employee_id = name_to_id.get(employee_name.lower())
            
            if not employee_id:
                print(f"Warning: No employee ID found for '{employee_name}', skipping")
                continue
            
            # Map columns to days (Monday=col1, Tuesday=col3, etc.)
            days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
            day_columns = [1, 3, 5, 7, 9, 11, 13]  # Column indices for each day
            hour_columns = [2, 4, 6, 8, 10, 12, 14]  # Hour columns
            
            employee_hours = 0
            
            for day_idx, (col_idx, hour_col_idx) in enumerate(zip(day_columns, hour_columns)):
                shift_text = row[col_idx]
                hours = row[hour_col_idx]
                
                if shift_text and shift_text != 'OFF' and 'OFF' not in str(shift_text):
                    shift_str = str(shift_text)
                    
                    # Parse shift
                    start_time, end_time = parse_shift_time(shift_str)
                    location = parse_location(shift_str)
                    
                    if start_time and end_time:
                        shift = {
                            "id": f"shift_{uuid.uuid4().hex[:8]}",
                            "employee_id": employee_id,
                            "day_of_week": days[day_idx],
                            "start_time": start_time,
                            "end_time": end_time,
                            "job_type": "employee",
                            "floor": None,
                            "location": location,
                            "hours": float(str(hours).rstrip('.')) if hours else 0,
                            "is_event": "EVENT" in shift_str.upper(),
                            "event_name": None,
                            "color": None,
                            "comment": shift_str,
                            "requires_break": False,
                            "break_provided": False
                        }
                        shifts.append(shift)
                        employee_hours += float(str(hours).rstrip('.')) if hours else 0
            
            total_hours[employee_id] = employee_hours
        
        schedule = {
            "id": schedule_id,
            "week_start_date": week_start,
            "shifts": shifts,
            "total_hours": total_hours,
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat(),
            "created_by": "excel_import",
            "status": "published",
            "metadata": {
                "source_sheet": sheet_name,
                "imported_at": datetime.now().isoformat()
            }
        }
        
        schedules.append(schedule)
        print(f"Imported {sheet_name}: {len(shifts)} shifts")
    
    # Save to JSON
    with open(output_path, 'w') as f:
        json.dump(schedules, f, indent=2)
    
    print(f"\nSaved {len(schedules)} schedules to {output_path}")
    return schedules

if __name__ == "__main__":
    excel_path = "/Users/levisantosaraujo/Downloads/Python/Schedule Sheet IBU/Last OPS Schedule June-Dec 2026.xlsx"
    output_path = "/Users/levisantosaraujo/Downloads/Python/Schedule Sheet IBU/api/backend/data/schedules.json"
    employees_path = "/Users/levisantosaraujo/Downloads/Python/Schedule Sheet IBU/api/backend/data/employees.json"
    
    convert_excel_to_schedules(excel_path, output_path, employees_path)
