# IBU Operations Schedule API
# GitHub persistence enabled - data stored in data branch
# CORS wildcard origin enabled for Excel download (temporary fix)
# Backend redeployed: 2026-06-28 to apply CORS fix
#
# DATA SOURCE = json_store ONLY. Excel is EXPORT-ONLY.
# Do NOT import data_store_excel or excel_store for data operations.
# Excel is permitted ONLY for generating .xlsx download/export files.
# See ARCHITECTURE.md for details.
import os
from dotenv import load_dotenv

# Load environment variables from .env for local development BEFORE any imports
load_dotenv('.env')

# Import log_storage first for print override
import log_storage
import flow_storage
from datetime import datetime, timezone
from collections import defaultdict
import threading
from datetime import timedelta

# Custom print function to capture logs for monitoring dashboard
_original_print = print

def custom_print(*args, **kwargs):
    """Custom print that logs to storage and calls original print."""
    # Log to storage
    message = ' '.join(str(arg) for arg in args)
    level = "INFO"
    if "[ERROR]" in message or "Error" in message or "error" in message:
        level = "ERROR"
    elif "[WARNING]" in message or "Warning" in message:
        level = "WARNING"
    elif "[CRITICAL]" in message:
        level = "CRITICAL"
    
    log_storage.get_log_storage().add_backend_log(message, level)
    
    # Call original print
    _original_print(*args, **kwargs)

# Override print globally BEFORE any other imports
print = custom_print

# In-memory session tracking
_active_sessions = defaultdict(lambda: datetime.now(timezone.utc))
_session_lock = threading.Lock()

# In-memory rate limiting for auth endpoints
_auth_attempts = defaultdict(list)  # IP -> list of timestamps
_auth_lock = threading.Lock()
MAX_AUTH_ATTEMPTS = 10
AUTH_WINDOW_SECONDS = 300  # 5 minutes

def check_rate_limit(client_ip: str) -> bool:
    """Check if client has exceeded rate limit for auth attempts."""
    with _auth_lock:
        now = datetime.now(timezone.utc)
        attempts = _auth_attempts[client_ip]
        # Remove attempts older than window
        _auth_attempts[client_ip] = [t for t in attempts if (now - t).total_seconds() < AUTH_WINDOW_SECONDS]
        if len(_auth_attempts[client_ip]) >= MAX_AUTH_ATTEMPTS:
            return False
        _auth_attempts[client_ip].append(now)
        return True

def track_session(user_id: str):
    """Track a user session."""
    with _session_lock:
        _active_sessions[user_id] = datetime.now(timezone.utc)

def get_active_session_count() -> int:
    """Get count of active sessions (last 5 minutes)."""
    with _session_lock:
        now = datetime.now(timezone.utc)
        cutoff = now - timedelta(minutes=5)
        # Remove stale sessions
        stale_users = [uid for uid, last_seen in _active_sessions.items() if last_seen < cutoff]
        for uid in stale_users:
            del _active_sessions[uid]
        return len(_active_sessions)

# Now import all other modules (they will use custom print)
from fastapi import FastAPI, HTTPException, Query, Header, Depends, UploadFile, File, Form, Cookie, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from contextlib import asynccontextmanager
from typing import List, Optional, Dict
from datetime import date
import uuid
import queue
import time
import shutil
import io
import re
from pathlib import Path
from pydantic import BaseModel
import sys

from models import (
    Employee, EmployeeUpdate, Availability, WeeklySchedule, Shift, JobType, Floor,
    AvailabilityType, EmployeeType, FloorCoverageQuery, FloorCoverageResponse,
    AVAILABILITY_COLORS, HourlyCoverageRequirement,
    AvailabilityRequest, AvailabilityRequestStatus, Notification, NotificationType, Event
)

from excel_store import (
    set_blob_key, _clear_workbook_cache
)
from json_data import (
    save_employee, delete_employee,
    get_all_schedules, get_schedule_by_week, save_schedule, delete_schedule,
    get_floor_coverage, get_system_config, save_system_config,
    get_availability_requests, save_availability_request,
    get_all_week_schedule_dates,
    set_manager_password, verify_manager_password, manager_has_password,
    get_coverage_requirements, save_coverage_requirement,
    get_notifications, save_notification, mark_notification_read,
    initialize_sample_data,
)
import json_store as staging_store
from storage import store_excel_data, excel_file_exists, get_excel_data
from scheduler import SchedulingEngine, generate_schedule
from auth import AuthManager, require_auth, require_manager, require_self_or_manager
from openpyxl import load_workbook

# Background sync queue for GitHub writes
_sync_queue: queue.Queue = queue.Queue()
_sync_worker_thread: Optional[threading.Thread] = None
_sync_worker_running = False

# Pydantic models for requests
class LoginRequest(BaseModel):
    employee_id: str
    password: Optional[str] = None

class AdminLoginRequest(BaseModel):
    employee_id: str
    password: Optional[str] = None
    secret_key: str

class SetPasswordRequest(BaseModel):
    employee_id: str
    password: str

# Sync status tracking
_sync_status: Dict[str, str] = {}  # request_id -> "syncing", "synced", "failed"
_sync_status_lock = threading.RLock()


def _sync_worker():
    """Background worker to process sync queue with retry logic."""
    global _sync_worker_running
    _sync_worker_running = True
    print("[SYNC] Background sync worker started")
    
    while _sync_worker_running:
        try:
            # Get task from queue with timeout
            task = _sync_queue.get(timeout=1.0)
            if task is None:  # Poison pill to stop worker
                break
            
            request_id = task.get('request_id')
            sync_func = task.get('sync_func')
            retry_count = task.get('retry_count', 0)
            max_retries = 3
            
            try:
                # Update status to syncing
                with _sync_status_lock:
                    _sync_status[request_id] = "syncing"
                
                # Execute sync function
                success = sync_func()
                
                if success:
                    with _sync_status_lock:
                        _sync_status[request_id] = "synced"
                    print(f"[SYNC] Successfully synced request {request_id}")
                else:
                    if retry_count < max_retries:
                        # Retry with exponential backoff
                        retry_count += 1
                        backoff = min(2 ** retry_count, 10)  # Max 10 seconds
                        print(f"[SYNC] Retry {retry_count}/{max_retries} for request {request_id} in {backoff}s")
                        time.sleep(backoff)
                        task['retry_count'] = retry_count
                        _sync_queue.put(task)
                    else:
                        with _sync_status_lock:
                            _sync_status[request_id] = "failed"
                        print(f"[SYNC] Failed to sync request {request_id} after {max_retries} retries")
                
            except Exception as e:
                print(f"[SYNC] Error processing sync for request {request_id}: {e}")
                import traceback
                traceback.print_exc()
                
                if retry_count < max_retries:
                    retry_count += 1
                    backoff = min(2 ** retry_count, 10)
                    print(f"[SYNC] Retry {retry_count}/{max_retries} for request {request_id} in {backoff}s")
                    time.sleep(backoff)
                    task['retry_count'] = retry_count
                    _sync_queue.put(task)
                else:
                    with _sync_status_lock:
                        _sync_status[request_id] = "failed"
            
            _sync_queue.task_done()
            
        except queue.Empty:
            continue
        except Exception as e:
            print(f"[SYNC] Worker error: {e}")
            time.sleep(1)
    
    print("[SYNC] Background sync worker stopped")


def start_sync_worker():
    """Start the background sync worker thread."""
    global _sync_worker_thread
    if _sync_worker_thread is None or not _sync_worker_thread.is_alive():
        _sync_worker_thread = threading.Thread(target=_sync_worker, daemon=True)
        _sync_worker_thread.start()
        print("[SYNC] Started background sync worker thread")


def stop_sync_worker():
    """Stop the background sync worker thread."""
    global _sync_worker_running
    _sync_worker_running = False
    _sync_queue.put(None)  # Poison pill
    if _sync_worker_thread:
        _sync_worker_thread.join(timeout=5)
        print("[SYNC] Stopped background sync worker thread")


def add_sync_task(request_id: str, sync_func):
    """Add a sync task to the background queue."""
    _sync_queue.put({
        'request_id': request_id,
        'sync_func': sync_func,
        'retry_count': 0
    })
    print(f"[SYNC] Added sync task for request {request_id}")


def get_sync_status(request_id: str) -> Optional[str]:
    """Get the sync status for a request."""
    with _sync_status_lock:
        return _sync_status.get(request_id)


def time_ranges_overlap(start_a: str, end_a: str, start_b: str, end_b: str) -> bool:
    """Check if two time ranges overlap. Returns True if they overlap."""
    try:
        from datetime import datetime
        a_start = datetime.strptime(start_a, "%H:%M")
        a_end = datetime.strptime(end_a, "%H:%M")
        b_start = datetime.strptime(start_b, "%H:%M")
        b_end = datetime.strptime(end_b, "%H:%M")
        
        # Handle overnight shifts (e.g., 22:00 to 02:00)
        if a_end <= a_start:
            a_end += timedelta(days=1)
        if b_end <= b_start:
            b_end += timedelta(days=1)
        
        # Check for overlap
        return a_start < b_end and b_start < a_end
    except Exception as e:
        print(f"[ERROR] Failed to parse time ranges for overlap check: {e}")
        return False


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: do not load Excel file - all data is served from JSON/blob storage"""
    _clear_workbook_cache()
    print("[STARTUP] Server starting - using JSON/blob storage for all data")
    print("[STARTUP] Initializing log storage for monitoring dashboard")
    print("[STARTUP] Starting background sync worker")
    
    # GitHub JSON is the single source of truth - no Excel fallback needed
    print(f"[STARTUP] Using GitHub JSON as single source of truth")
    
    # Start background sync worker
    start_sync_worker()
    
    print("[STARTUP] Server ready - monitoring dashboard available at /admin/dashboard")
    
    # Load and check data
    print("[STARTUP] Loading employee data...")
    try:
        employees_dicts = staging_store.get_employees()
        print(f"[STARTUP] Loaded {len(employees_dicts)} employees")
        employees = [Employee(**e) for e in employees_dicts]
        employee_ids = {e.id for e in employees}
    except Exception as e:
        print(f"[STARTUP] Error loading employees: {e}")
        employees = []
        employee_ids = set()
    
    # Old availability system removed - using availability_requests.json only
    
    print("[STARTUP] Loading schedule data...")
    try:
        staging_schedules = staging_store.get_schedules()
        print(f"[STARTUP] Loaded {len(staging_schedules) if staging_schedules else 0} schedule records")
    except Exception as e:
        print(f"[STARTUP] Error loading schedules: {e}")
    
    # Ensure sixth_floor and 80_bloor are in staffing targets
    print("[STARTUP] Checking staffing targets...")
    try:
        config = get_system_config()
        targets = config.staffing_targets or {}
        updated = False
        if 'sixth_floor' not in targets:
            print("[STARTUP] Adding sixth_floor to staffing targets")
            targets['sixth_floor'] = 1
            updated = True
        else:
            print(f"[STARTUP] sixth_floor already in staffing targets with value {targets['sixth_floor']}")
        if '80_bloor' not in targets:
            print("[STARTUP] Adding 80_bloor to staffing targets")
            targets['80_bloor'] = 1
            updated = True
        else:
            print(f"[STARTUP] 80_bloor already in staffing targets with value {targets['80_bloor']}")
        if updated:
            config.staffing_targets = targets
            save_system_config(config)
            staging_store.set_system_config(config.model_dump(), user_id="system")
            print("[STARTUP] Updated staffing targets")
    except Exception as e:
        print(f"[STARTUP] Error checking/updating staffing targets: {e}")
    
    print("[STARTUP] Checking GitHub configuration...")
    try:
        github_repo = os.getenv("GITHUB_REPO", "Not configured")
        github_branch = os.getenv("GITHUB_DATA_BRANCH", "Not configured")
        print(f"[STARTUP] GitHub repository: {github_repo}")
        print(f"[STARTUP] GitHub data branch: {github_branch}")
    except Exception as e:
        print(f"[STARTUP] Error checking GitHub config: {e}")
    
    # Note: Orphaned cleanup removed from startup to prevent deleting valid requests
    # Cleanup only happens when employees are explicitly deleted (in delete endpoint)
    
    yield
    _clear_workbook_cache()

class ExcelPathRequest(BaseModel):
    file_path: str

# Global state
CURRENT_EXCEL_FILE: Optional[str] = None

def is_manager_user(user: Optional[Dict]) -> bool:
    if not user:
        return False
    return str(user.get('employee_type') or user.get('role') or '').lower() == EmployeeType.MANAGER.value

UPLOAD_DIR = Path(__file__).parent / "uploads"
try:
    UPLOAD_DIR.mkdir(exist_ok=True)
except OSError:
    # In Vercel serverless, the filesystem is read-only at /var/task
    # We'll use Vercel Blob storage instead
    pass

app = FastAPI(title="IBU Operations team schedule", version="2.0.0", lifespan=lifespan)

# Enable CORS for frontend
# Note: allow_credentials=True is incompatible with allow_origins=["*"]
# We use allow_credentials=False and handle auth via Bearer token headers instead
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://ibu-operations-schedule-frontend.vercel.app",
        "https://ibu-operations-schedule-frontend-2dbzrj86d.vercel.app",
        "https://ibu-operations-schedule-frontend-ju3it15eq.vercel.app",
        "https://ibu-operations-schedule-frontend-jcrrl3bmx.vercel.app",
        "https://ibu-operations-schedule-frontend-rbrfsfadd.vercel.app",
        "https://ibu-operations-schedule-frontend-f47jxd1tv.vercel.app",
        "https://ibu-operations-schedule.vercel.app",
        "http://localhost:3000",
        "http://localhost:5173",
        "http://localhost:8000",
        "*"  # Allow all origins temporarily to fix CORS issues
    ],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Flow tracking middleware
@app.middleware("http")
async def flow_tracking_middleware(request, call_next):
    """Track API requests in flow storage."""
    # Skip admin dashboard endpoints to avoid noise
    if request.url.path.startswith("/admin"):
        return await call_next(request)
    
    # Start flow chain
    flow = flow_storage.get_flow_storage()
    chain = flow.start_chain("api", f"{request.method} {request.url.path}")
    
    # Set chain ID in context for nested operations
    import json_store
    import cache_manager
    token = json_store._flow_chain_id.set(chain.id)
    token2 = json_store._flow_parent_id.set(chain.root_operation.id)
    token3 = cache_manager._flow_chain_id.set(chain.id)
    token4 = cache_manager._flow_parent_id.set(chain.root_operation.id)
    
    try:
        response = await call_next(request)
        
        # Complete chain with success
        flow.complete_chain(
            chain.id,
            status="success",
            metadata={
                "status_code": response.status_code,
                "method": request.method,
                "path": request.url.path
            }
        )
        
        # Clear context
        json_store._flow_chain_id.reset(token)
        json_store._flow_parent_id.reset(token2)
        cache_manager._flow_chain_id.reset(token3)
        cache_manager._flow_parent_id.reset(token4)
        
        return response
    except Exception as e:
        # Complete chain with error
        flow.complete_chain(
            chain.id,
            status="error",
            metadata={
                "error": str(e),
                "method": request.method,
                "path": request.url.path
            }
        )
        
        # Clear context
        json_store._flow_chain_id.reset(token)
        json_store._flow_parent_id.reset(token2)
        cache_manager._flow_chain_id.reset(token3)
        cache_manager._flow_parent_id.reset(token4)
        
        raise

# ============ Health Check ============

@app.get("/")
async def root():
    return {
        "status": "running",
        "message": "IBU Operations Schedule API is running",
        "health": "/api/health",
    }

# ============ Excel File Management ============

@app.get("/api/excel/status")
async def excel_status():
    """Check if system is configured (using JSON/blob storage)"""
    try:
        employees = staging_store.get_employees()
        has_employees = len(employees) > 0
        from storage import BLOB_AVAILABLE
        return {
            "configured": has_employees,
            "file_path": "json_blob_storage" if has_employees else None,
            "file_exists": has_employees,
            "storage_type": "json_blob" if BLOB_AVAILABLE else "json_memory"
        }
    except Exception as e:
        return {
            "configured": False,
            "file_path": None,
            "file_exists": False
        }

@app.get("/api/excel/download")
async def download_excel():
    """Export current GitHub JSON data to Excel file with weekly sheet format"""
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from datetime import datetime, timedelta

    # Get all schedules from GitHub JSON first
    schedules = staging_store.get_schedules()
    employees = staging_store.get_employees()
    
    # Create Excel workbook
    wb = Workbook()
    
    # Track if any sheets were created
    sheets_created = False
    
    # Create employee lookup
    employee_map = {emp['id']: emp['name'] for emp in employees}
    
    # Group schedules by week
    for schedule in schedules:
        week_start_date = schedule.get('week_start_date')
        if not week_start_date:
            continue
        
        # Parse week start date
        try:
            week_start = datetime.strptime(week_start_date, '%Y-%m-%d')
        except:
            continue
        
        # Calculate week end (6 days later)
        week_end = week_start + timedelta(days=6)
        
        # Format sheet name like "June 15-21"
        month_name = week_start.strftime('%B')
        start_day = week_start.day
        end_day = week_end.day
        sheet_name = f"{month_name} {start_day}-{end_day}"
        
        # Create weekly sheet
        ws = wb.create_sheet(sheet_name)
        sheets_created = True
        
        # Header row
        headers = ["Employee", "Mon", "", "Tue", "", "Wed", "", "Thu", "", "Fri", "", "Sat", "", "Sun", ""]
        ws.append(headers)
        
        # Group shifts by employee
        shifts_by_employee = {}
        for shift in schedule.get('shifts', []):
            emp_id = shift.get('employee_id')
            if emp_id not in shifts_by_employee:
                shifts_by_employee[emp_id] = {}
            day = shift.get('day_of_week', '').lower()
            if day not in shifts_by_employee[emp_id]:
                shifts_by_employee[emp_id][day] = []
            shifts_by_employee[emp_id][day].append(shift)
        
        # Map day names to column indices (0-based)
        day_columns = {
            'monday': 1,
            'tuesday': 3,
            'wednesday': 5,
            'thursday': 7,
            'friday': 9,
            'saturday': 11,
            'sunday': 13
        }
        
        # Add rows for each employee
        for emp_id, emp_name in employee_map.items():
            row = [emp_name] + [""] * 14
            
            # Add shifts for each day
            if emp_id in shifts_by_employee:
                for day, shifts in shifts_by_employee[emp_id].items():
                    col_idx = day_columns.get(day)
                    if col_idx is not None:
                        # Combine multiple shifts with '/'
                        shift_texts = []
                        total_hours = 0
                        for shift in shifts:
                            start_time = shift.get('start_time', '')
                            end_time = shift.get('end_time', '')
                            hours = shift.get('hours', 0)
                            
                            # Format time as "9a-5p" style
                            shift_text = format_time_range(start_time, end_time)
                            shift_texts.append(shift_text)
                            total_hours += hours
                        
                        # Put shift text in day column, hours in next column
                        if shift_texts:
                            row[col_idx] = '/'.join(shift_texts)
                            row[col_idx + 1] = total_hours if total_hours > 0 else ""
            
            ws.append(row)
    
    # If no sheets were created from schedules, use the default sheet
    if not sheets_created:
        ws = wb.active
        ws.title = "June 1-7"
        headers = ["Employee", "Mon", "", "Tue", "", "Wed", "", "Thu", "", "Fri", "", "Sat", "", "Sun", ""]
        ws.append(headers)
    else:
        # Remove the default "Sheet" since we created schedule sheets
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=IBU_Schedule.xlsx"
        }
    )
    
    # Manually set CORS headers on response object
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "*"
    
    return response

@app.options("/api/excel/download")
async def download_excel_options():
    """Handle OPTIONS preflight request for CORS"""
    from fastapi.responses import Response
    response = Response()
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "*"
    return response

def format_time_range(start_time: str, end_time: str) -> str:
    """Format time range as '9a-5p' style"""
    if not start_time or not end_time:
        return ""
    
    try:
        # Parse 24-hour format
        start_h, start_m = map(int, start_time.split(':'))
        end_h, end_m = map(int, end_time.split(':'))
        
        # Convert to 12-hour format with am/pm
        def format_12h(h, m):
            if h == 0:
                return f"12{':'+str(m) if m > 0 else ''}a"
            elif h < 12:
                return f"{h}{':'+str(m) if m > 0 else ''}a"
            elif h == 12:
                return f"12{':'+str(m) if m > 0 else ''}p"
            else:
                return f"{h-12}{':'+str(m) if m > 0 else ''}p"
        
        return f"{format_12h(start_h, start_m)}-{format_12h(end_h, end_m)}"
    except:
        return f"{start_time}-{end_time}"


# ============ Password Management ============

@app.post("/api/managers/set-password")
async def set_manager_password_endpoint(request: SetPasswordRequest):
    """Set password for a manager (only if not already set)"""
    employee_dict = staging_store.get_employee_by_id(request.employee_id)
    if not employee_dict:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    employee = Employee(**employee_dict)
    if employee.employee_type != EmployeeType.MANAGER:
        raise HTTPException(status_code=403, detail="Only managers can have passwords")
    
    # Check if password already exists
    if manager_has_password(request.employee_id):
        raise HTTPException(status_code=400, detail="Password already set. Use update-password to change it.")
    
    set_manager_password(request.employee_id, employee.name, request.password)
    
    return {"message": "Password set successfully"}

@app.put("/api/managers/update-password")
async def update_manager_password_endpoint(request: SetPasswordRequest, authorization: Optional[str] = Header(None, alias="Authorization")):
    """Update password for a manager (managers can update their own, or managers can update others)"""
    # Auth check
    user = require_auth(authorization)
    if user["role"] not in ["manager", "admin"] and user["employee_id"] != request.employee_id:
        raise HTTPException(status_code=403, detail="Can only update your own password")
    
    employee_dict = staging_store.get_employee_by_id(request.employee_id)
    if not employee_dict:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    employee = Employee(**employee_dict)
    if employee.employee_type != EmployeeType.MANAGER:
        raise HTTPException(status_code=403, detail="Only managers can have passwords")
    
    set_manager_password(request.employee_id, employee.name, request.password)
    
    return {"message": "Password updated successfully"}

@app.post("/api/managers/verify-password")
async def verify_password_endpoint(request: SetPasswordRequest):
    """Verify manager password"""
    if not manager_has_password(request.employee_id):
        raise HTTPException(status_code=400, detail="Password not set")
    
    is_valid = verify_manager_password(request.employee_id, request.password)
    
    if not is_valid:
        raise HTTPException(status_code=401, detail="Invalid password")
    
    return {"valid": True}

@app.get("/api/managers/has-password/{employee_id}")
async def check_has_password(employee_id: str):
    """Check if manager has password set"""
    return {"has_password": manager_has_password(employee_id)}

# ============ Auth Endpoints ============

@app.post("/api/login")
async def login(request: LoginRequest, client_ip: str = Header(None, alias="X-Forwarded-For")):
    """Login as an employee with optional password for managers"""
    # Rate limiting
    ip = client_ip or "unknown"
    if not check_rate_limit(ip):
        raise HTTPException(status_code=429, detail="Too many login attempts. Please try again later.")

    # Check if system has employees configured (using JSON/blob storage)
    employees = staging_store.get_employees()
    has_employees = len(employees) > 0
    if not has_employees:
        raise HTTPException(status_code=400, detail="No employees configured. Please initialize the system first.")

    employee_dict = staging_store.get_employee_by_id(request.employee_id)
    if not employee_dict:
        raise HTTPException(status_code=404, detail="Employee not found")

    employee = Employee(**employee_dict)

    # Managers need password verification
    if employee.employee_type == EmployeeType.MANAGER:
        if manager_has_password(request.employee_id):
            if not request.password:
                raise HTTPException(status_code=401, detail="Password required for managers")
            if not verify_manager_password(request.employee_id, request.password):
                raise HTTPException(status_code=401, detail="Invalid password")
        # If no password set yet, allow login (first time setup)

    token = AuthManager.login(request.employee_id, request.password, staging_store.get_employee_by_id)

    return {
        "token": token,
        "employee": employee,
        "role": employee.employee_type,
        "requires_password_setup": employee.employee_type == EmployeeType.MANAGER and not manager_has_password(request.employee_id)
    }

@app.post("/api/admin-login")
async def admin_login(request: AdminLoginRequest, client_ip: str = Header(None, alias="X-Forwarded-For")):
    """Secret admin login endpoint - requires secret key"""
    # Rate limiting
    ip = client_ip or "unknown"
    if not check_rate_limit(ip):
        raise HTTPException(status_code=429, detail="Too many login attempts. Please try again later.")

    ADMIN_SECRET_KEY = os.getenv("ADMIN_SECRET_KEY")
    if not ADMIN_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Server misconfigured: ADMIN_SECRET_KEY not set")

    if request.secret_key != ADMIN_SECRET_KEY:
        raise HTTPException(status_code=403, detail="Invalid secret key")

    # Check if system has employees configured
    employees = staging_store.get_employees()
    has_employees = len(employees) > 0
    if not has_employees:
        raise HTTPException(status_code=400, detail="No employees configured. Please initialize the system first.")

    employee_dict = staging_store.get_employee_by_id(request.employee_id)
    if not employee_dict:
        raise HTTPException(status_code=404, detail="Employee not found")

    employee = Employee(**employee_dict)

    # Force manager role for admin login
    if employee.employee_type != EmployeeType.MANAGER:
        raise HTTPException(status_code=403, detail="Admin login only available for managers")

    # Bypass password verification for admin login
    token = AuthManager.login(request.employee_id, request.password, staging_store.get_employee_by_id)

    return {
        "token": token,
        "employee": employee,
        "role": "admin",  # Force admin role
        "requires_password_setup": False
    }

@app.post("/api/logout")
async def logout(authorization: Optional[str] = Header(None)):
    """Logout and end session"""
    if authorization and authorization.startswith("Bearer "):
        token = authorization.replace("Bearer ", "")
        AuthManager.logout(token)
    return {"message": "Logged out"}

@app.get("/api/me")
async def get_me(authorization: Optional[str] = Header(None)):
    """Get current logged-in user info"""
    user = AuthManager.get_current_user(authorization)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user

# ============ Employee Endpoints ============

@app.post("/api/staging/cleanup-orphaned-records")
async def cleanup_orphaned_records(authorization: str = Header(None)):
    """Remove availability records for non-existent employees"""
    try:
        user = require_manager(authorization)

        # Get all current employees
        employees_dicts = staging_store.get_employees()
        employees = [Employee(**e) for e in employees_dicts]
        employee_ids = {e.id for e in employees}
        
        # Clean up staging availability requests
        staging_requests = staging_store.get_availability_requests()
        orphaned_requests = 0
        valid_requests = []
        for req in staging_requests:
            if req.get('employee_id') in employee_ids:
                valid_requests.append(req)
            else:
                orphaned_requests += 1
                print(f"[CLEANUP] Removing orphaned request for employee {req.get('employee_id')}")
        
        if orphaned_requests > 0:
            print(f"[CLEANUP] Found {orphaned_requests} orphaned availability requests")
            staging_store.set_availability_requests(valid_requests, user_id="system")
        
        total_cleaned = orphaned_count + orphaned_requests
        if total_cleaned > 0:
            return {"message": f"Cleaned up {total_cleaned} orphaned records ({orphaned_count} availabilities, {orphaned_requests} requests)"}
        else:
            return {"message": "No orphaned records found"}
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to cleanup orphaned records: {str(e)}")


@app.get("/api/employees", response_model=List[Employee])
async def list_employees(active_only: bool = False, authorization: str = Header(None)):
    """List all employees (managers see manager_preferences, employees do not)"""
    try:
        # Read from GitHub JSON (single source of truth)
        staging_employees = staging_store.get_employees()
        employees = [Employee(**e) for e in staging_employees]
        
        print(f"[DEBUG] Loaded {len(employees)} employees")
        if active_only:
            employees = [e for e in employees if e.active]
        
        # Hide manager_preferences from non-managers
        user = None
        if authorization:
            try:
                user = require_auth(authorization)
            except:
                pass  # If auth fails, just return employees without manager_preferences
        
        # Hide manager_preferences if user is not a manager (or if auth failed)
        if not is_manager_user(user):
            for emp in employees:
                emp.manager_preferences = {}
        
        return employees
    except Exception as e:
        print(f"Error loading employees: {e}")
        raise HTTPException(status_code=500, detail=f"Error loading employees: {str(e)}")

@app.get("/api/employees/{employee_id}", response_model=Employee)
async def get_employee(employee_id: str):
    """Get a specific employee"""
    employee_dict = staging_store.get_employee_by_id(employee_id)
    if not employee_dict:
        raise HTTPException(status_code=404, detail="Employee not found")
    return Employee(**employee_dict)

@app.post("/api/employees", response_model=Employee)
async def create_employee(
    employee: Employee,
    user: Dict = Depends(require_manager)
):
    """Create a new employee (managers only)"""
    if not employee.id:
        employee.id = f"emp_{uuid.uuid4().hex[:8]}"
    
    # Update staging layer with immediate write to prevent race conditions
    employees = staging_store.get_employees()
    employees.append(employee.model_dump())
    staging_store.set_employees(employees, user_id=user.get('employee_id'), immediate=True)
    
    # No action queue needed - GitHub JSON is single source of truth
    
    return employee

@app.put("/api/employees/{employee_id}", response_model=Employee)
async def update_employee(
    employee_id: str,
    employee_update: EmployeeUpdate,
    user: Dict = Depends(require_self_or_manager)
):
    """Update an employee (managers can edit anyone, employees can edit themselves)"""
    existing_dict = staging_store.get_employee_by_id(employee_id)
    if not existing_dict:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    existing = Employee(**existing_dict)
    
    # Build updated employee by merging existing with new values
    update_data = employee_update.model_dump(exclude_unset=True)
    
    # Only managers can update manager_preferences
    if not is_manager_user(user) and 'manager_preferences' in update_data:
        del update_data['manager_preferences']
    
    # Create updated employee object
    updated_employee = Employee(
        id=employee_id,
        name=update_data.get("name", existing.name),
        email=update_data.get("email", existing.email),
        employee_type=update_data.get("employee_type", existing.employee_type),
        max_hours_per_week=update_data.get("max_hours_per_week", existing.max_hours_per_week),
        preferences=update_data.get("preferences", existing.preferences),
        manager_preferences=update_data.get("manager_preferences", existing.manager_preferences),
        active=update_data.get("active", existing.active),
        created_at=existing.created_at
    )
    
    # Update staging layer with immediate write to prevent race conditions
    employees = staging_store.get_employees()
    employees = [e if e["id"] != employee_id else updated_employee.model_dump() for e in employees]
    staging_store.set_employees(employees, user_id=user.get('employee_id'), immediate=True)
    
    # Add to action queue for Excel sync (async, no blocking)
    # No action queue needed - GitHub JSON is single source of truth
    
    return updated_employee

@app.delete("/api/employees/{employee_id}")
async def remove_employee(
    employee_id: str,
    user: Dict = Depends(require_manager)
):
    """Delete an employee (managers only)"""
    # Update staging layer with immediate write to prevent race conditions
    employees = staging_store.get_employees()
    was_in_staging = any(e["id"] == employee_id for e in employees)
    employees = [e for e in employees if e["id"] != employee_id]
    staging_store.set_employees(employees, user_id=user.get('employee_id'), immediate=True)
    
    # Clean up availability requests for deleted employee with immediate write
    requests = staging_store.get_availability_requests()
    cleaned_requests = [r for r in requests if r.get("employee_id") != employee_id]
    if len(cleaned_requests) != len(requests):
        print(f"[DELETE] Cleaning up {len(requests) - len(cleaned_requests)} availability requests for deleted employee {employee_id}")
        staging_store.set_availability_requests(cleaned_requests, user_id=user.get('employee_id'), immediate=True)
    
    # Clean up notifications for deleted employee with immediate write
    notifications = staging_store.get_notifications()
    cleaned_notifications = [n for n in notifications if n.get("employee_id") != employee_id]
    if len(cleaned_notifications) != len(notifications):
        print(f"[DELETE] Cleaning up {len(notifications) - len(cleaned_notifications)} notifications for deleted employee {employee_id}")
        staging_store.set_notifications(cleaned_notifications, user_id=user.get('employee_id'), immediate=True)
    
    # Note: Excel cleanup removed - GitHub JSON is now the single source of truth
    
    # Return success if employee was in staging
    if was_in_staging:
        return {"message": "Employee deleted"}
    
    raise HTTPException(status_code=404, detail="Employee not found")

# ============ Availability Endpoints ============

# ============ Hourly Coverage Requirements Endpoints ============

@app.get("/api/coverage-requirements/{week_start_date}")
async def get_coverage_requirements(
    week_start_date: date,
    authorization: Optional[str] = Header(None)
):
    """Get hourly coverage requirements for a week (managers only)"""
    user = require_manager(authorization)
    requirements = get_coverage_requirements(week_start_date)
    return {"week_start_date": week_start_date, "requirements": requirements}

@app.post("/api/coverage-requirements")
async def set_coverage_requirements(
    requirements: List[HourlyCoverageRequirement],
    authorization: Optional[str] = Header(None)
):
    """Set hourly coverage requirements for a week (managers only)"""
    user = require_manager(authorization)
    for req in requirements:
        req.created_by = user["employee_id"]
        save_coverage_requirement(req)
    return {"message": f"Saved {len(requirements)} coverage requirements"}

# ============ Schedule Endpoints ============

@app.get("/api/schedules", response_model=List[WeeklySchedule])
async def list_schedules(authorization: Optional[str] = Header(None)):
    """List all schedules (all authenticated users)"""
    require_auth(authorization)
    
    # Try to read from staging first
    staging_schedules = staging_store.get_schedules()
    return [WeeklySchedule(**s) for s in staging_schedules]

@app.get("/api/schedules/{week_start_date}")
async def get_schedule(
    week_start_date: date,
    authorization: Optional[str] = Header(None)
):
    """Get schedule for a specific week (all authenticated users)"""
    user = require_auth(authorization)
    
    print(f"[SCHEDULE] Getting schedule for week {week_start_date}, user: {user.get('employee_id')}, is_manager: {is_manager_user(user)}")
    
    # Try to read from staging first
    print(f"[SCHEDULE] Attempting to read schedules from staging_store...", file=sys.stderr, flush=True)
    try:
        staging_schedules = staging_store.get_schedules()
        print(f"[SCHEDULE] SUCCESS: Read {len(staging_schedules)} schedules from staging_store", file=sys.stderr, flush=True)
    except Exception as e:
        print(f"[SCHEDULE] ERROR reading from staging_store: {e}", file=sys.stderr, flush=True)
        import traceback
        traceback.print_exc(file=sys.stderr)
        # Fallback: read directly from GitHub if staging fails
        print(f"[SCHEDULE] Attempting fallback to GitHub...", file=sys.stderr, flush=True)
        import json_store
        staging_schedules = json_store.get_schedules()
        print(f"[SCHEDULE] Fallback SUCCESS: read {len(staging_schedules)} schedules from GitHub", file=sys.stderr, flush=True)
    
    schedule = None
    if staging_schedules:
        for s in staging_schedules:
            # Handle both string and date objects
            week_start = s.get('week_start_date')
            if isinstance(week_start, date):
                week_start_str = str(week_start)
            else:
                week_start_str = week_start
            
            if week_start_str == str(week_start_date):
                schedule = WeeklySchedule(**s)
                print(f"[SCHEDULE] Found schedule in staging with {len(schedule.shifts)} shifts")
                # Log locked shifts
                locked_shifts = [s for s in schedule.shifts if s.locked]
                print(f"[SCHEDULE] Locked shifts in schedule: {len(locked_shifts)}")
                for ls in locked_shifts:
                    print(f"[SCHEDULE]   - {ls.id}: emp={ls.employee_id}, day={ls.day_of_week}, locked={ls.locked}, type={ls.locked_availability_type}")
                break
    
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    # Load availability requests for marker derivation; do not mutate schedule
    all_requests = staging_store.get_availability_requests()
    week_end_date = week_start_date + timedelta(days=6)
    day_map = {0: 'monday', 1: 'tuesday', 2: 'wednesday', 3: 'thursday', 4: 'friday', 5: 'saturday', 6: 'sunday'}
    
    # Derive approved availability markers from approved requests
    # This replaces the old locked-shift mechanism
    approved_markers = []
    
    for req in all_requests:
        if req.get('status') not in ['approved', 'AvailabilityRequestStatus.APPROVED']:
            continue
        
        emp_id = req.get('employee_id')
        if not emp_id:
            continue
        
        # Filter: employees see only their own
        if not is_manager_user(user) and emp_id != user.get('employee_id'):
            continue
        
        # Handle new date range model
        if req.get('start_date') and req.get('end_date'):
            try:
                req_start = date.fromisoformat(str(req['start_date'])[:10])
                req_end = date.fromisoformat(str(req['end_date'])[:10])
                req_days = req.get('days_of_week', [])
                request_type = req.get('request_type', 'availability')
                
                # Check if this request overlaps with the current week
                if req_end < week_start_date or req_start > week_end_date:
                    continue
                
                # For each day in the week that matches the request's days_of_week
                current_date = week_start_date
                while current_date <= week_end_date:
                    day_name = day_map[current_date.weekday()]
                    if day_name in req_days and current_date >= req_start and current_date <= req_end:
                        # Check if a real shift exists for this employee/day with overlapping time
                        has_overlapping_shift = False
                        for shift in schedule.shifts:
                            if shift.employee_id == emp_id and shift.day_of_week == day_name:
                                # For day-off requests, any shift overlaps
                                if request_type == 'day_off':
                                    has_overlapping_shift = True
                                    break
                                # For time-range requests, check time overlap
                                elif req.get('start_time') and req.get('end_time'):
                                    if time_ranges_overlap(
                                        req.get('start_time'), req.get('end_time'),
                                        shift.start_time, shift.end_time
                                    ):
                                        has_overlapping_shift = True
                                        break
                        
                        # Only add marker if no overlapping real shift exists
                        if not has_overlapping_shift:
                            marker = {
                                'employee_id': emp_id,
                                'day_of_week': day_name,
                                'date': current_date.isoformat(),
                                'request_type': request_type,
                                'start_time': req.get('start_time'),
                                'end_time': req.get('end_time'),
                                'employee_comment': req.get('employee_comment'),
                                'request_id': req.get('id')
                            }
                            approved_markers.append(marker)
                    
                    current_date += timedelta(days=1)
            except Exception as e:
                print(f"[SCHEDULE] Warning: could not parse approved request: {e}")
                continue
    
    # Add availability requests to the schedule response (pending requests only for managers)
    week_requests = []
    for req in all_requests:
        if req.get('status') == 'pending':
            # Check if request overlaps with the current week
            if req.get('start_date') and req.get('end_date'):
                try:
                    req_start = date.fromisoformat(str(req['start_date'])[:10])
                    req_end = date.fromisoformat(str(req['end_date'])[:10])
                    if req_end < week_start_date or req_start > week_end_date:
                        continue
                    week_requests.append(req)
                    print(f"[SCHEDULE] Added pending request {req.get('id')} for week {week_start_date}")
                except Exception as e:
                    print(f"[SCHEDULE] Error parsing pending request: {e}")
                    continue
            elif req.get('week_start_date'):
                # Legacy model
                if str(req.get('week_start_date')) == str(week_start_date):
                    week_requests.append(req)
                    print(f"[SCHEDULE] Added legacy pending request {req.get('id')} for week {week_start_date}")
    
    # Filter: employees see only their own
    if not is_manager_user(user):
        week_requests = [r for r in week_requests if r.get('employee_id') == user.get('employee_id')]
    
    # Add availabilities and requests to schedule (this is a dynamic field, not in the model)
    schedule_data = schedule.model_dump()
    schedule_data['approved_availabilities'] = approved_markers
    schedule_data['availability_requests'] = week_requests
    
    return schedule_data

@app.post("/api/schedules/generate/{week_start_date}", response_model=WeeklySchedule)
async def auto_generate_schedule(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Auto-generate a schedule for a week (managers only)"""
    from scheduler import SchedulingEngine

    try:
        # Get staffing targets from config
        config = get_system_config()
        # SystemConfig is a Pydantic model, access attributes directly
        staffing_targets = getattr(config, 'staffing_targets', None) or getattr(config, 'floor_requirements', {}) or {}
        print(f"[API] Staffing targets: {staffing_targets}")

        # Convert staffing targets to location_requirements format
        # staffing_targets is simple dict like {'ground_floor': 2, 'call_center': 4}
        # location_requirements needs to be {'location': {'monday': 2, 'tuesday': 2, ...}}
        location_requirements = {}
        event_staffing = {}  # event_id -> people_needed
        call_center_target = 0  # Number of people to assign call center role for the week
        days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']  # Sunday excluded - only for events

        # Map staffing target keys to location names
        # Note: call_center is now a role/flag, not a physical location
        location_map = {
            'ground_floor': 'ground floor',
            'second_floor': '2nd floor',
            'sixth_floor': '6th floor',
            '80_bloor': '80 bloor',
            'working_from_home': 'working from home'
        }

        for key, target in staffing_targets.items():
            if key.startswith('event_'):
                # Event staffing target - handle both old (event_event_...) and new (event_...) formats
                # Strip double prefix if present
                event_id = key.replace('event_event_', 'event_')
                event_staffing[event_id] = target
                print(f"[API] Event staffing: {key} -> {event_id} = {target}")
            elif key == 'call_center':
                # Call center is now a role/flag, not a location - pass separately to scheduler
                # Target is people per day (applies to every day)
                call_center_target = target
                print(f"[API] Call center role target: {target} people per day")
            else:
                # Regular location staffing target - target is people per day (applies to every day)
                # Skip working_from_home - manager assigns manually on-demand
                if key == 'working_from_home':
                    continue
                location_name = location_map.get(key, key.replace('_', ' '))
                location_requirements[location_name] = {day: target for day in days}
                print(f"[API] Location staffing: {key} -> {location_name} = {target} people per day")

        print(f"[API] Total event staffing entries: {len(event_staffing)}")
        print(f"[API] Event staffing dict: {event_staffing}")
        print(f"[API] Call center target: {call_center_target}")
        print(f"[API] Location requirements passed to scheduler: {location_requirements}")
        
        # Preserve locked shifts from existing schedule before generating new one
        staging_schedules = staging_store.get_schedules()
        existing_schedule = None
        for s in staging_schedules:
            if str(s.get('week_start_date')) == str(week_start_date):
                existing_schedule = WeeklySchedule(**s)
                break
        
        # Recreate locked shifts from approved availability requests before generating
        # This ensures locked shifts are preserved even if they were only recreated in get_schedule
        all_requests = staging_store.get_availability_requests()
        week_end_date = week_start_date + timedelta(days=6)
        day_map = {0: 'monday', 1: 'tuesday', 2: 'wednesday', 3: 'thursday', 4: 'friday', 5: 'saturday', 6: 'sunday'}
        
        # If existing schedule exists, use it as base; otherwise start with empty schedule
        if existing_schedule:
            schedule_for_locked = existing_schedule
        else:
            schedule_for_locked = WeeklySchedule(week_start_date=week_start_date, shifts=[])
        
        # Recreate approved locked shifts if they're missing but the request still exists
        for req in all_requests:
            if req.get('status') not in ['approved', 'AvailabilityRequestStatus.APPROVED']:
                continue
            
            emp_id = req.get('employee_id')
            if not emp_id:
                continue
            
            # Check if request overlaps with the current week
            if req.get('start_date') and req.get('end_date'):
                try:
                    req_start = date.fromisoformat(str(req['start_date'])[:10])
                    req_end = date.fromisoformat(str(req['end_date'])[:10])
                    if req_end < week_start_date or req_start > week_end_date:
                        continue
                    
                    req_days = req.get('days_of_week', [])
                    request_type = req.get('request_type', 'availability')
                    
                    # For each day in the week that matches the request's days_of_week
                    current_date = week_start_date
                    while current_date <= week_end_date:
                        day_name = day_map[current_date.weekday()]
                        if day_name in req_days and current_date >= req_start and current_date <= req_end:
                            # Check if an approved locked shift already exists for this employee/day
                            existing_approved = next(
                                (s for s in schedule_for_locked.shifts 
                                 if s.employee_id == emp_id and s.day_of_week == day_name and s.locked and s.id.startswith(f"locked_{req['id']}")),
                                None
                            )
                            
                            if not existing_approved:
                                # Create approved locked shift
                                if request_type == 'day_off':
                                    shift_start = '00:00'
                                    shift_end = '23:59'
                                    shift_location = 'day off'
                                    shift_hours = 0
                                    locked_type = 'Day Off'
                                else:
                                    shift_start = req.get('start_time', '00:00')
                                    shift_end = req.get('end_time', '23:59')
                                    shift_location = None
                                    shift_hours = 0
                                    locked_type = request_type.capitalize()
                                
                                approved_shift = Shift(
                                    id=f"locked_{req['id']}_{day_name}_{emp_id}",
                                    employee_id=emp_id,
                                    day_of_week=day_name,
                                    start_time=shift_start,
                                    end_time=shift_end,
                                    job_type=JobType.IBU_OPS,
                                    location=shift_location,
                                    hours=shift_hours,
                                    locked=True,
                                    locked_availability_type=locked_type,
                                    is_event=False
                                )
                                schedule_for_locked.shifts.append(approved_shift)
                                print(f"[API] Recreated approved locked shift for {emp_id} on {day_name}")
                        
                        current_date += timedelta(days=1)
                except Exception as e:
                    print(f"[API] Warning: could not recreate approved locked shift: {e}")
                    continue
        
        # Now preserve all locked shifts (including recreated ones)
        preserved_locked_shifts = [s for s in schedule_for_locked.shifts if s.locked]
        print(f"[API] Preserving {len(preserved_locked_shifts)} locked shifts")
        
        engine = SchedulingEngine()
        schedule = engine.generate_auto_schedule(week_start_date, location_requirements, event_staffing, call_center_target)
        
        # Add preserved locked shifts back to the new schedule
        for locked_shift in preserved_locked_shifts:
            schedule.shifts.append(locked_shift)
            print(f"[API] Restored locked shift: {locked_shift.id}")
        
        # Update staging layer first (fast) - remove all schedules with this week_start_date to prevent duplicates
        schedules = staging_store.get_schedules()
        schedules = [s for s in schedules if str(s.get('week_start_date')) != str(week_start_date)]
        schedules.append(schedule.model_dump())
        staging_store.set_schedules(schedules, user_id=user.get('employee_id'), immediate=True)
        
        # Add to action queue for Excel sync (async, no blocking)
        # No action queue needed - GitHub JSON is single source of truth
        
        return schedule
    except Exception as e:
        import traceback
        print(f"[API] Error in auto_generate_schedule: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Schedule generation failed: {str(e)}")

@app.post("/api/schedules")
async def create_or_update_schedule(
    schedule: WeeklySchedule,
    user: Dict = Depends(require_manager)
):
    """Save a schedule (manual editing) (managers only)"""
    schedule.updated_at = datetime.now()

    # Log what we're receiving
    import sys
    print(f"[SAVE SCHEDULE] Saving schedule {schedule.id} for week {schedule.week_start_date}", file=sys.stderr, flush=True)
    print(f"[SAVE SCHEDULE] Total shifts received: {len(schedule.shifts)}", file=sys.stderr, flush=True)
    locked_count = len([s for s in schedule.shifts if s.id and (s.id.startswith('locked_') or s.id.startswith('pending_'))])
    print(f"[SAVE SCHEDULE] Locked shifts: {locked_count}, Auto-generated shifts: {len(schedule.shifts) - locked_count}", file=sys.stderr, flush=True)

    # Update staging layer with immediate write to prevent race conditions
    # Read schedules from staging (in-memory cache) to avoid re-fetching stale data from GitHub
    print(f"[SAVE SCHEDULE] Attempting to read schedules from staging_store...", file=sys.stderr, flush=True)
    try:
        schedules = staging_store.get_schedules()
        print(f"[SAVE SCHEDULE] SUCCESS: Read {len(schedules)} schedules from staging_store", file=sys.stderr, flush=True)
    except Exception as e:
        print(f"[SAVE SCHEDULE] ERROR reading from staging_store: {e}", file=sys.stderr, flush=True)
        import traceback
        traceback.print_exc(file=sys.stderr)
        # Fallback: read directly from GitHub if staging fails
        print(f"[SAVE SCHEDULE] Attempting fallback to GitHub...", file=sys.stderr, flush=True)
        import json_store
        schedules = json_store.get_schedules()
        print(f"[SAVE SCHEDULE] Fallback SUCCESS: read {len(schedules)} schedules from GitHub", file=sys.stderr, flush=True)
    
    existing_schedule = next((s for s in schedules if s['id'] == schedule.id), None)
    if existing_schedule:
        print(f"[SAVE SCHEDULE] Existing schedule has {len(existing_schedule.get('shifts', []))} shifts", file=sys.stderr, flush=True)
    schedules = [s if s['id'] != schedule.id else schedule.model_dump() for s in schedules]
    if not any(s['id'] == schedule.id for s in schedules):
        schedules.append(schedule.model_dump())
    print(f"[SAVE SCHEDULE] Writing {len(schedules)} schedules to staging_store", file=sys.stderr, flush=True)
    result = staging_store.set_schedules(schedules, user_id=user.get('employee_id'), immediate=True)
    print(f"[SAVE SCHEDULE] set_schedules returned: {result}", file=sys.stderr, flush=True)

    # Add dynamic fields to response (approved_availabilities, availability_requests)
    # Populate from availability requests for this week
    week_start_date = schedule.week_start_date
    week_end_date = week_start_date + timedelta(days=6)
    all_requests = staging_store.get_availability_requests()
    
    # Add approved availabilities (markers for approved requests without locked shifts)
    approved_markers = []
    for req in all_requests:
        if req.get('status') not in ['approved', 'AvailabilityRequestStatus.APPROVED']:
            continue
        # ... (simplified - just return empty for now to avoid complexity)
    
    # Add pending requests for this week
    week_requests = []
    for req in all_requests:
        if req.get('status') == 'pending':
            if req.get('start_date') and req.get('end_date'):
                try:
                    req_start = date.fromisoformat(str(req['start_date'])[:10])
                    req_end = date.fromisoformat(str(req['end_date'])[:10])
                    if req_end < week_start_date or req_start > week_end_date:
                        continue
                    week_requests.append(req)
                except:
                    continue
    
    schedule_data = schedule.model_dump()
    schedule_data['approved_availabilities'] = approved_markers
    schedule_data['availability_requests'] = week_requests

    return schedule_data

@app.put("/api/schedules/{week_start_date}/shifts", response_model=WeeklySchedule)
async def update_schedule_shifts(
    week_start_date: date,
    shifts: List[Shift],
    user: Dict = Depends(require_manager)
):
    """Update shifts for a schedule (drag-drop saves here) (managers only)"""
    # Try staging first for speed
    staging_schedules = staging_store.get_schedules()
    schedule = None
    for s in staging_schedules:
        # Handle both string and date objects
        week_start = s.get('week_start_date')
        if isinstance(week_start, date):
            week_start_str = str(week_start)
        else:
            week_start_str = week_start
        
        if week_start_str == str(week_start_date):
            schedule = WeeklySchedule(**s)
            break
    
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    schedule.shifts = shifts
    # Recalculate total hours
    total_hours = {}
    for shift in shifts:
        # Skip day off shifts when calculating total hours
        if shift.location == 'day off' or shift.locked_availability_type == 'Day Off':
            continue
        current = total_hours.get(shift.employee_id, 0)
        total_hours[shift.employee_id] = current + shift.hours
    schedule.total_hours = total_hours
    schedule.updated_at = datetime.now()
    
    # Update staging layer with immediate write to prevent race conditions
    schedules = staging_store.get_schedules()
    schedules = [s if s['id'] != schedule.id else schedule.model_dump() for s in schedules]
    staging_store.set_schedules(schedules, user_id=user.get('employee_id'), immediate=True)
    
    # Add to action queue for Excel sync (async, no blocking)
    # No action queue needed - GitHub JSON is single source of truth
    
    return schedule

@app.put("/api/schedules/{week_start_date}/shifts/{shift_id}/break")
async def mark_break_provided(
    week_start_date: date,
    shift_id: str,
    break_provided: bool = True,
    user: Dict = Depends(require_manager)
):
    """Mark a shift's break as provided (managers only)"""
    # Try staging first for speed
    staging_schedules = staging_store.get_schedules()
    schedule = None
    for s in staging_schedules:
        # Handle both string and date objects
        week_start = s.get('week_start_date')
        if isinstance(week_start, date):
            week_start_str = str(week_start)
        else:
            week_start_str = week_start
        
        if week_start_str == str(week_start_date):
            schedule = WeeklySchedule(**s)
            break
    
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    shift = next((s for s in schedule.shifts if s.id == shift_id), None)
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
    
    shift.break_provided = break_provided
    schedule.updated_at = datetime.now()
    
    # Update staging layer with immediate write to prevent race conditions
    schedules = staging_store.get_schedules()
    schedules = [s if s['id'] != schedule.id else schedule.model_dump() for s in schedules]
    staging_store.set_schedules(schedules, user_id=user.get('employee_id'), immediate=True)
    
    # Add to action queue for Excel sync (async, no blocking)
    # No action queue needed - GitHub JSON is single source of truth
    
    return schedule

@app.post("/api/schedules/{week_start_date}/publish")
async def publish_schedule(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Publish a schedule (finalize it) (managers only)"""
    # Try staging first for speed
    staging_schedules = staging_store.get_schedules()
    schedule = None
    for s in staging_schedules:
        # Handle both string and date objects
        week_start = s.get('week_start_date')
        if isinstance(week_start, date):
            week_start_str = str(week_start)
        else:
            week_start_str = week_start
        
        if week_start_str == str(week_start_date):
            schedule = WeeklySchedule(**s)
            break
    
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    schedule.status = "published"
    
    # Update staging layer with immediate write to prevent race conditions
    schedules = staging_store.get_schedules()
    schedules = [s if s['id'] != schedule.id else schedule.model_dump() for s in schedules]
    staging_store.set_schedules(schedules, user_id=user.get('employee_id'), immediate=True)
    
    # Add to action queue for Excel sync (async, no blocking)
    # No action queue needed - GitHub JSON is single source of truth
    
    return {"message": "Schedule published"}

@app.delete("/api/schedules/{week_start_date}")
async def remove_schedule(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Delete a schedule (managers only)"""
    # Update staging layer first (fast)
    staging_schedules = staging_store.get_schedules()
    initial_count = len(staging_schedules)
    staging_schedules = [s for s in staging_schedules if str(s.get('week_start_date')) != str(week_start_date)]
    
    if len(staging_schedules) < initial_count:
        # Schedule was in staging, update staging with immediate write
        staging_store.set_schedules(staging_schedules, user_id=user.get('employee_id'), immediate=True)
        return {"message": "Schedule deleted"}
    
    raise HTTPException(status_code=404, detail="Schedule not found")

@app.get("/api/schedules/weeks/available")
async def get_available_weeks(user: Dict = Depends(require_manager)):
    """Get list of all weeks that have schedule data (managers only)"""
    weeks = get_all_week_schedule_dates()
    return {
        "weeks": [w.isoformat() for w in weeks],
        "count": len(weeks)
    }

@app.post("/api/schedules/{week_start_date}/clear")
async def clear_schedule_shifts(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Clear all shifts from a schedule while preserving events (managers only)"""
    print(f"[CLEAR SCHEDULE] Starting clear for week {week_start_date} by user {user.get('employee_id')}")
    
    # Try to read from staging first
    staging_schedules = staging_store.get_schedules()
    print(f"[CLEAR SCHEDULE] Found {len(staging_schedules)} schedules in staging")
    
    schedule = None
    for s in staging_schedules:
        # Handle both string and date objects
        week_start = s.get('week_start_date')
        if isinstance(week_start, date):
            week_start_str = str(week_start)
        else:
            week_start_str = week_start
        
        if week_start_str == str(week_start_date):
            schedule = WeeklySchedule(**s)
            print(f"[CLEAR SCHEDULE] Found schedule for week {week_start_date} with {len(schedule.shifts)} shifts")
            break
    
    if not schedule:
        print(f"[CLEAR SCHEDULE] Schedule not found for week {week_start_date}")
        raise HTTPException(status_code=404, detail="Schedule not found")

    # Preserve locked shifts (both approved and pending) - identified by ID prefix
    # locked_* = approved shifts, pending_* = pending shifts
    locked_shifts = [s for s in schedule.shifts if s.id and (s.id.startswith('locked_') or s.id.startswith('pending_'))]
    print(f"[CLEAR SCHEDULE] Preserving {len(locked_shifts)} locked shifts")

    # Clear only non-locked shifts
    schedule.shifts = locked_shifts
    schedule.total_hours = {}
    print(f"[CLEAR SCHEDULE] Cleared non-locked shifts from schedule")
    
    # Update staging layer using already-loaded list (avoid re-fetching stale data)
    updated_schedules = [s for s in staging_schedules if str(s.get('week_start_date')) != str(week_start_date)]
    updated_schedules.append(schedule.model_dump())
    print(f"[CLEAR SCHEDULE] Updating staging with {len(updated_schedules)} schedules")
    
    result = staging_store.set_schedules(updated_schedules, user_id=user.get('employee_id'), immediate=True)
    print(f"[CLEAR SCHEDULE] set_schedules returned: {result}")
    
    return schedule

# ============ Floor Coverage & Queries ============

@app.get("/api/floor-coverage/{floor}/{day_of_week}/{time_slot}")
async def query_floor_coverage(
    floor: Floor,
    day_of_week: str,
    time_slot: str,
    week_start_date: date = Query(...),
    user: Dict = Depends(require_manager)
):
    """Query how many employees are on a floor at a given time (managers only)"""
    result = get_floor_coverage(floor.value, day_of_week, time_slot, week_start_date)
    return result

@app.get("/api/floor-coverage/summary/{week_start_date}")
async def get_weekly_floor_summary(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Get floor coverage summary for entire week (managers only)"""
    days = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    time_slots = ["morning", "afternoon", "evening"]
    floors = [Floor.GROUND, Floor.SECOND, Floor.SIXTH]
    
    summary = {}
    for floor in floors:
        summary[floor.value] = {}
        for day in days:
            summary[floor.value][day] = {}
            for slot in time_slots:
                result = get_floor_coverage(floor.value, day, slot, week_start_date)
                summary[floor.value][day][slot] = result["employee_count"]
    
    return summary

# ============ Analytics & Reporting ============

@app.get("/api/analytics/employee-hours/{week_start_date}")
async def get_employee_hours_summary(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Get hours summary for all employees for a week (managers only)"""
    staging_schedules = staging_store.get_schedules()
    schedule = None
    for s in staging_schedules:
        if s.get('week_start_date') == str(week_start_date):
            schedule = WeeklySchedule(**s)
            break
    
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    employees_dicts = staging_store.get_employees()
    employees = [Employee(**e) for e in employees_dicts]
    
    summary = []
    for emp in employees:
        hours = schedule.total_hours.get(emp.id, 0) if schedule else 0
        remaining = emp.max_hours_per_week - hours
        summary.append({
            "employee_id": emp.id,
            "name": emp.name,
            "type": emp.employee_type,
            "scheduled_hours": hours,
            "max_hours": emp.max_hours_per_week,
            "remaining_hours": remaining,
            "utilization": round(hours / emp.max_hours_per_week * 100, 1) if emp.max_hours_per_week > 0 else 0
        })
    
    return summary

# ============ Configuration ============

@app.get("/api/config")
async def get_config():
    """Get system configuration"""
    # Try to read from staging first
    return staging_store.get_system_config()

@app.put("/api/config")
async def update_config(config: Dict):
    """Update system configuration"""
    result = save_system_config(config)
    
    # Update staging layer
    staging_store.set_system_config(result.model_dump(), user_id="system")
    
    # Add to action queue for Excel sync
    # No action queue needed - GitHub JSON is single source of truth
    
    return result

@app.get("/api/staffing-targets")
async def get_staffing_targets():
    """Get staffing targets for all locations"""
    config = get_system_config()
    targets = config.staffing_targets
    
    # Normalize keys: strip double event_ prefix
    cleaned_targets = {}
    for key, value in targets.items():
        # Always strip double event_ prefix if present
        normalized_key = key.replace('event_event_', 'event_')
        cleaned_targets[normalized_key] = value
    
    # Remove working_from_home from targets (manager assigns manually)
    if 'working_from_home' in cleaned_targets:
        del cleaned_targets['working_from_home']
    
    return cleaned_targets

@app.put("/api/staffing-targets")
async def update_staffing_targets(targets: Dict[str, int], user: Dict = Depends(require_manager)):
    """Update staffing targets for locations"""
    config = get_system_config()
    
    # Normalize keys: strip double event_ prefix
    cleaned_targets = {}
    for key, value in targets.items():
        # Always strip double event_ prefix if present
        normalized_key = key.replace('event_event_', 'event_')
        # Remove working_from_home from targets (manager assigns manually)
        if normalized_key == 'working_from_home':
            continue
        cleaned_targets[normalized_key] = value
    
    config.staffing_targets = cleaned_targets
    result = save_system_config(config)
    
    if result:
        # Update staging layer
        staging_store.set_system_config(result.model_dump(), user_id="system")
        
        # Add to action queue for Excel sync
        # No action queue needed - GitHub JSON is single source of truth
    
    return result

# ============ Availability Requests ============

@app.get("/api/availability-requests")
async def get_all_availability_requests(authorization: str = Header(None)):
    """Get all availability requests (managers only)"""
    user = AuthManager.get_current_user(authorization)
    if not user or user.get('role') not in ['manager', 'admin']:
        raise HTTPException(status_code=403, detail="Manager access required")
    
    # Read from GitHub JSON (single source of truth)
    staging_requests = staging_store.get_availability_requests()
    # Add employee names to requests
    all_employees_dicts = staging_store.get_employees()
    all_employees = [Employee(**e) for e in all_employees_dicts]
    employee_map = {e.id: e.name for e in all_employees}
    for request in staging_requests:
        request['employee_name'] = employee_map.get(request.get('employee_id'), 'Unknown')
    return staging_requests

@app.get("/api/availability-requests/my")
async def get_my_availability_requests(authorization: str = Header(None)):
    """Get current user's availability requests"""
    user = AuthManager.get_current_user(authorization)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    # Read from GitHub JSON (single source of truth)
    staging_requests = staging_store.get_availability_requests()
    my_requests = [r for r in staging_requests if r.get('employee_id') == user.get('employee_id')]
    return my_requests

@app.post("/api/availability-requests")
async def create_availability_request(request: Dict, authorization: str = Header(None)):
    """Create an availability request"""
    try:
        user = AuthManager.get_current_user(authorization)
        if not user:
            raise HTTPException(status_code=401, detail="Unauthorized")

        request['id'] = str(uuid.uuid4())
        request['employee_id'] = user['employee_id']
        request['status'] = AvailabilityRequestStatus.PENDING
        request['created_at'] = datetime.now()

        # Update staging layer with immediate write to prevent race conditions
        requests = staging_store.get_availability_requests()
        requests.append(request)
        staging_store.set_availability_requests(requests, user_id=user.get('employee_id'), immediate=True)
        
        # Add to action queue for Excel sync (async, no blocking)
        # No action queue needed - GitHub JSON is single source of truth

        # Create locked shifts for pending availability request
        try:
            request_type = request.get('request_type', 'availability')
            start_date_str = request.get('start_date', request.get('week_start_date', ''))
            end_date_str = request.get('end_date', start_date_str)
            days_of_week = request.get('days_of_week', [])
            employee_id = request.get('employee_id')
            
            if start_date_str and end_date_str and days_of_week:
                try:
                    start_date = date.fromisoformat(str(start_date_str)[:10])
                    end_date = date.fromisoformat(str(end_date_str)[:10])
                    day_map = {0: 'monday', 1: 'tuesday', 2: 'wednesday', 3: 'thursday', 4: 'friday', 5: 'saturday', 6: 'sunday'}
                    
                    # Collect all schedule updates in memory to batch write once
                    schedules_by_week = {}
                    
                    # Process each week in the date range
                    current_week_start = start_date - timedelta(days=start_date.weekday())
                    while current_week_start <= end_date:
                        week_end = current_week_start + timedelta(days=6)
                        
                        # Get or create schedule for this week
                        staging_schedules = staging_store.get_schedules()
                        schedule = None
                        for s in staging_schedules:
                            week_start = s.get('week_start_date')
                            if isinstance(week_start, date):
                                week_start_str = str(week_start)
                            else:
                                week_start_str = week_start
                            if week_start_str == str(current_week_start):
                                schedule = WeeklySchedule(**s)
                                break
                        
                        if not schedule:
                            # Create new schedule
                            schedule = WeeklySchedule(
                                id=f"schedule_{current_week_start.isoformat()}",
                                week_start_date=current_week_start,
                                shifts=[],
                                total_hours={},
                                created_by=user.get('employee_id')
                            )
                        
                        # Create locked shifts for each matching day in this week
                        current_date = max(current_week_start, start_date)
                        while current_date <= min(week_end, end_date):
                            day_name = day_map[current_date.weekday()]
                            if day_name in days_of_week:
                                # Check if a locked shift already exists for this employee/day
                                existing_locked = next(
                                    (s for s in schedule.shifts 
                                     if s.employee_id == employee_id and s.day_of_week == day_name and s.locked),
                                    None
                                )
                                
                                if not existing_locked:
                                    # Create locked shift
                                    if request_type == 'day_off':
                                        shift_start = '00:00'
                                        shift_end = '23:59'
                                        shift_location = 'day off'
                                        shift_hours = 0
                                        locked_type = 'Day Off'
                                    else:
                                        shift_start = request.get('start_time', '00:00')
                                        shift_end = request.get('end_time', '23:59')
                                        shift_location = None  # Placeholder, manager will assign
                                        shift_hours = 0  # Don't count toward total
                                        locked_type = request_type.capitalize()
                                    
                                    locked_shift = Shift(
                                        id=f"pending_{request['id']}_{day_name}_{employee_id}",
                                        employee_id=employee_id,
                                        day_of_week=day_name,
                                        start_time=shift_start,
                                        end_time=shift_end,
                                        job_type=JobType.IBU_OPS,  # Placeholder
                                        location=shift_location,
                                        hours=shift_hours,
                                        locked=True,
                                        locked_availability_type=locked_type,
                                        is_event=False
                                    )
                                    schedule.shifts.append(locked_shift)
                                    print(f"[CREATION] Created pending locked shift for {employee_id} on {day_name}")
                            
                            current_date += timedelta(days=1)
                        
                        # Store schedule in memory for batch write
                        schedules_by_week[str(current_week_start)] = schedule
                        current_week_start += timedelta(days=7)
                    
                    # Batch write all schedules ONCE to avoid multiple GitHub API calls
                    if schedules_by_week:
                        staging_schedules = staging_store.get_schedules()
                        updated_schedules = [s for s in staging_schedules if str(s.get('week_start_date')) not in schedules_by_week]
                        for schedule in schedules_by_week.values():
                            updated_schedules.append(schedule.model_dump())
                        staging_store.set_schedules(updated_schedules, user_id=user.get('employee_id'), immediate=True)
                        print(f"[CREATION] Batch wrote {len(schedules_by_week)} schedules in a single write")
                        
                except Exception as e:
                    print(f"[CREATION] Warning: could not create pending locked shifts: {e}")
                    import traceback
                    traceback.print_exc()
        except Exception as e:
            print(f"[CREATION] Warning: error in pending locked shift creation: {e}")
            import traceback
            traceback.print_exc()
        
        # Send notification to managers
        try:
            employee_dict = staging_store.get_employee_by_id(user['employee_id'])
            employee = Employee(**employee_dict) if employee_dict else None
            employee_name = employee.name if employee else user['employee_id']
            
            request_type = request.get('request_type', 'availability')
            days_str = ', '.join(request.get('days_of_week', [])) if request.get('days_of_week') else 'All days'
            time_str = f"{request.get('start_time', '00:00')} - {request.get('end_time', '23:59')}" if request.get('start_time') else 'All day'
            
            # Get all managers
            all_employees_dicts = staging_store.get_employees()
            all_employees = [Employee(**e) for e in all_employees_dicts]
            managers = [e for e in all_employees if e.employee_type in ['manager', 'admin']]
            
            # Create all notifications first, then save once to avoid duplicates
            notifications = staging_store.get_notifications()
            for manager in managers:
                notification = {
                    'id': str(uuid.uuid4()),
                    'employee_id': manager.id,
                    'type': NotificationType.AVAILABILITY_REQUEST,
                    'message': f"{employee_name} submitted a {request_type} request for {days_str} ({time_str})",
                    'details': {
                        'request_id': request['id'],
                        'employee_id': user['employee_id'],
                        'employee_name': employee_name,
                        'request_type': request_type,
                        'days_of_week': days_str,
                        'time_range': time_str,
                        'start_date': request.get('start_date'),
                        'end_date': request.get('end_date'),
                        'employee_comment': request.get('employee_comment', '')
                    },
                    'created_at': datetime.now(),
                    'read': False
                }
                notifications.append(notification)
            
            # Save all notifications at once to avoid duplicate writes
            staging_store.set_notifications(notifications, user_id=user.get('employee_id'), immediate=True)
            print(f"[API] Created {len(managers)} manager notifications for request {request['id']}")
                
        except Exception as e:
            print(f"[API] Warning: could not send manager notification: {e}")
            import traceback
            traceback.print_exc()
        
        # Ensure response is JSON-serializable
        try:
            response = dict(request)  # Create a proper dict copy
            if isinstance(response.get('created_at'), datetime):
                response['created_at'] = response['created_at'].isoformat()
            if isinstance(response.get('updated_at'), datetime):
                response['updated_at'] = response['updated_at'].isoformat()
            if isinstance(response.get('approved_at'), datetime):
                response['approved_at'] = response['approved_at'].isoformat()
            print(f"[API] Returning response with keys: {list(response.keys())}")
            return response
        except Exception as e:
            import traceback
            print(f"[API] Error serializing response: {e}")
            traceback.print_exc()
            # Return a minimal response if serialization fails
            return {
                'id': request.get('id'),
                'status': str(request.get('status')),
                'employee_id': request.get('employee_id')
            }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"[API] Error creating availability request: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error creating request: {str(e)}")

@app.put("/api/availability-requests/{request_id}/approve")
async def approve_availability_request(request_id: str, body: Dict = {}, authorization: str = Header(None)):
    """Approve an availability request"""
    try:
        user = AuthManager.get_current_user(authorization)
        if not user or user.get('role') not in ['manager', 'admin']:
            raise HTTPException(status_code=403, detail="Manager access required")

        # Read from GitHub JSON (single source of truth)
        staging_requests = staging_store.get_availability_requests()
        requests = staging_requests
        
        request_data = next((r for r in requests if r['id'] == request_id), None)

        if not request_data:
            raise HTTPException(status_code=404, detail="Request not found")

        # Update status to APPROVED immediately
        request_data['status'] = AvailabilityRequestStatus.APPROVED
        request_data['manager_comment'] = body.get('comment')
        request_data['updated_at'] = datetime.now()
        request_data['approved_by'] = user.get('employee_id')
        request_data['approved_at'] = datetime.now()

        # Update staging layer with immediate write to prevent race conditions
        requests = staging_store.get_availability_requests()
        requests = [r if r['id'] != request_id else request_data for r in requests]
        staging_store.set_availability_requests(requests, user_id=user.get('employee_id'), immediate=True)
        
        # Process locked shifts and notifications asynchronously in background to speed up approval response
        import threading
        def process_approval_background():
            try:
                # Update pending locked shifts to approved locked shifts
                start_date_str = request_data.get('start_date', request_data.get('week_start_date', ''))
                end_date_str = request_data.get('end_date', start_date_str)
                days_of_week = request_data.get('days_of_week', [])
                employee_id = request_data.get('employee_id')
                request_type = request_data.get('request_type', 'availability')
                
                if start_date_str and end_date_str and days_of_week:
                    try:
                        start_date = date.fromisoformat(str(start_date_str)[:10])
                        end_date = date.fromisoformat(str(end_date_str)[:10])
                        day_map = {0: 'monday', 1: 'tuesday', 2: 'wednesday', 3: 'thursday', 4: 'friday', 5: 'saturday', 6: 'sunday'}

                        # Read all schedules ONCE before the loop to avoid multiple GitHub API calls
                        staging_schedules = staging_store.get_schedules()
                        schedules_by_week = {}
                        for s in staging_schedules:
                            week_start = s.get('week_start_date')
                            if isinstance(week_start, date):
                                week_start_str = str(week_start)
                            else:
                                week_start_str = week_start
                            schedules_by_week[week_start_str] = WeeklySchedule(**s)

                        # Process each week in the date range
                        current_week_start = start_date - timedelta(days=start_date.weekday())
                        while current_week_start <= end_date:
                            week_start_str = str(current_week_start)
                            schedule = schedules_by_week.get(week_start_str)

                            if not schedule:
                                # Create new empty schedule for this week
                                schedule = WeeklySchedule(
                                    id=f"schedule_{current_week_start.isoformat()}",
                                    week_start_date=current_week_start,
                                    shifts=[],
                                    total_hours={},
                                    created_by=user.get('employee_id')
                                )
                                schedules_by_week[week_start_str] = schedule

                            if schedule:
                                request_type = request_data.get('request_type', 'availability')

                                # Convert pending locked shifts to approved, or create new approved shifts if none exist
                                updated_shifts = []
                                converted_days = set()
                                for shift in schedule.shifts:
                                    if shift.id.startswith(f"pending_{request_id}_"):
                                        # Convert to approved locked shift
                                        updated_shift = Shift(
                                            id=shift.id.replace(f"pending_{request_id}_", f"locked_{request_id}_"),
                                            employee_id=shift.employee_id,
                                            day_of_week=shift.day_of_week,
                                            start_time=shift.start_time,
                                            end_time=shift.end_time,
                                            job_type=shift.job_type,
                                            location=shift.location,
                                            hours=shift.hours,
                                            locked=True,
                                            locked_availability_type=request_type.capitalize(),
                                            is_event=False
                                        )
                                        updated_shifts.append(updated_shift)
                                        converted_days.add(shift.day_of_week)
                                    else:
                                        updated_shifts.append(shift)
                                schedule.shifts = updated_shifts

                                # Only create new locked shifts if we actually converted some pending shifts
                                # This prevents regenerating shifts after a clear operation
                                if converted_days:
                                    # Create new approved locked shifts for days that don't have pending shifts
                                    current_date = current_week_start
                                    while current_date <= end_date:
                                        day_name = day_map[current_date.weekday()]
                                        if day_name in days_of_week and day_name not in converted_days:
                                            # Check if there's already a locked shift for this day
                                            existing_locked = any(
                                                s.locked and s.employee_id == employee_id and s.day_of_week == day_name
                                                for s in schedule.shifts
                                            )
                                            if not existing_locked:
                                                # Determine shift details based on request type
                                                if request_type == 'day_off':
                                                    shift_location = 'day off'
                                                    shift_start = '00:00'
                                                    shift_end = '23:59'
                                                    shift_hours = 24
                                                    locked_type = 'Day Off'
                                                else:
                                                    shift_location = request_data.get('location', '80 Bloor')
                                                    shift_start = request_data.get('start_time', '09:00')
                                                    shift_end = request_data.get('end_time', '17:00')
                                                    shift_hours = request_data.get('hours', 8)
                                                    locked_type = request_type.capitalize()

                                                locked_shift = Shift(
                                                    id=f"locked_{request_id}_{day_name}_{employee_id}",
                                                    employee_id=employee_id,
                                                    day_of_week=day_name,
                                                    start_time=shift_start,
                                                    end_time=shift_end,
                                                    job_type=JobType.IBU_OPS,
                                                    location=shift_location,
                                                    hours=shift_hours,
                                                    locked=True,
                                                    locked_availability_type=locked_type,
                                                    is_event=False
                                                )
                                                schedule.shifts.append(locked_shift)
                                                print(f"[APPROVAL] Created locked shift directly for {employee_id} on {day_name}")
                                        current_date += timedelta(days=1)

                            current_week_start += timedelta(days=7)

                        # Merge modified schedules with existing ones
                        staging_schedules = staging_store.get_schedules()
                        updated_schedules = [s for s in staging_schedules if str(s.get('week_start_date')) not in schedules_by_week]
                        for schedule in schedules_by_week.values():
                            updated_schedules.append(schedule.model_dump())
                        staging_store.set_schedules(updated_schedules, user_id=user.get('employee_id'), immediate=True)
                        print(f"[APPROVAL] Updated {len(schedules_by_week)} schedules in a single write")
                            
                    except Exception as e:
                        print(f"[APPROVAL] Warning: could not update pending locked shifts: {e}")
                        import traceback
                        traceback.print_exc()
            except Exception as e:
                print(f"[APPROVAL] Warning: error in pending locked shift update: {e}")
                import traceback
                traceback.print_exc()

            # Remove manager notifications for this request after approval
            try:
                notifications = staging_store.get_notifications()
                cleaned = [
                    n for n in notifications
                    if not (
                        n.get('type') == NotificationType.AVAILABILITY_REQUEST and
                        n.get('details', {}).get('request_id') == request_id
                    )
                ]
                if len(cleaned) < len(notifications):
                    staging_store.set_notifications(cleaned, user_id=user.get('employee_id'), immediate=True)
            except Exception as e:
                print(f"[APPROVAL] Warning: could not clean up manager notifications: {e}")

            # Send notification to employee
            try:
                employee_dict = staging_store.get_employee_by_id(request_data['employee_id'])
                employee = Employee(**employee_dict) if employee_dict else None
                employee_name = employee.name if employee else request_data['employee_id']
                
                # Support both old and new model for notification message
                request_type = request_data.get('request_type', 'availability')
                start_date = request_data.get('start_date', request_data.get('week_start_date', ''))
                end_date = request_data.get('end_date', start_date)
                
                comment_part = f" Manager note: {body.get('comment')}" if body.get('comment') else ""
                days_str = ', '.join(request_data.get('days_of_week', [])) if request_data.get('days_of_week') else 'All days'
                time_str = f"{request_data.get('start_time', '00:00')} - {request_data.get('end_time', '23:59')}" if request_data.get('start_time') else 'All day'
                notification = {
                    'id': str(uuid.uuid4()),
                    'employee_id': request_data['employee_id'],
                    'type': NotificationType.AVAILABILITY_APPROVED,
                    'message': f"Your {request_type} request ({start_date} to {end_date}) has been approved.{comment_part}",
                    'details': {
                        'request_type': request_type,
                        'start_date': start_date,
                        'end_date': end_date,
                        'days_of_week': days_str,
                        'time_range': time_str,
                        'employee_comment': request_data.get('employee_comment', '')
                    },
                    'created_at': datetime.now(),
                    'read': False
                }
                # Update staging layer
                notifications = staging_store.get_notifications()
                notifications.append(notification)
                staging_store.set_notifications(notifications, user_id=user.get('employee_id'), immediate=True)
            except Exception as e:
                import traceback
                print(f"[API] Warning: could not send notification: {e}")
                traceback.print_exc()

        # Start background thread for locked shift and notification processing
        thread = threading.Thread(target=process_approval_background)
        thread.daemon = True
        thread.start()

        # Ensure response is JSON-serializable
        response = request_data.copy()
        if isinstance(response.get('created_at'), datetime):
            response['created_at'] = response['created_at'].isoformat()
        if isinstance(response.get('updated_at'), datetime):
            response['updated_at'] = response['updated_at'].isoformat()
        if isinstance(response.get('approved_at'), datetime):
            response['approved_at'] = response['approved_at'].isoformat()
        return response

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"[API] Error approving availability request: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error approving request: {str(e)}")

@app.put("/api/availability-requests/{request_id}/reject")
async def reject_availability_request(request_id: str, body: Dict = {}, authorization: str = Header(None)):
    """Reject an availability request"""
    try:
        user = AuthManager.get_current_user(authorization)
        if not user or user.get('role') not in ['manager', 'admin']:
            raise HTTPException(status_code=403, detail="Manager access required")

        # Read from GitHub JSON (single source of truth)
        staging_requests = staging_store.get_availability_requests()
        requests = staging_requests
        
        request_data = next((r for r in requests if r['id'] == request_id), None)

        if not request_data:
            raise HTTPException(status_code=404, detail="Request not found")

        request_data['status'] = AvailabilityRequestStatus.REJECTED
        request_data['manager_comment'] = body.get('comment', '')
        request_data['updated_at'] = datetime.now()

        # Update staging layer with immediate write to prevent race conditions
        requests = staging_store.get_availability_requests()
        requests = [r if r['id'] != request_id else request_data for r in requests]
        staging_store.set_availability_requests(requests, user_id=user.get('employee_id'), immediate=True)

        # Remove locked shifts for this request (cleanup if it was previously approved)
        try:
            start_date_str = request_data.get('start_date', request_data.get('week_start_date', ''))
            end_date_str = request_data.get('end_date', start_date_str)
            days_of_week = request_data.get('days_of_week', [])
            employee_id = request_data.get('employee_id')
            
            if start_date_str and end_date_str and days_of_week:
                try:
                    start_date = date.fromisoformat(str(start_date_str)[:10])
                    end_date = date.fromisoformat(str(end_date_str)[:10])
                    day_map = {0: 'monday', 1: 'tuesday', 2: 'wednesday', 3: 'thursday', 4: 'friday', 5: 'saturday', 6: 'sunday'}
                    
                    # Collect all schedule updates in memory to batch write once
                    schedules_by_week = {}
                    
                    # Process each week in the date range
                    current_week_start = start_date - timedelta(days=start_date.weekday())
                    while current_week_start <= end_date:
                        # Get schedule for this week
                        staging_schedules = staging_store.get_schedules()
                        schedule = None
                        for s in staging_schedules:
                            week_start = s.get('week_start_date')
                            if isinstance(week_start, date):
                                week_start_str = str(week_start)
                            else:
                                week_start_str = week_start
                            if week_start_str == str(current_week_start):
                                schedule = WeeklySchedule(**s)
                                break
                        
                        if schedule:
                            # Remove locked shifts for this request (both pending and approved)
                            original_count = len(schedule.shifts)
                            schedule.shifts = [
                                s for s in schedule.shifts
                                if not (s.locked and (s.id.startswith(f"locked_{request_id}_") or s.id.startswith(f"pending_{request_id}_")))
                            ]
                            removed_count = original_count - len(schedule.shifts)
                            
                            if removed_count > 0:
                                # Store schedule in memory for batch write
                                schedules_by_week[str(current_week_start)] = schedule
                                print(f"[REJECTION] Marked {removed_count} locked shifts for removal from week {current_week_start}")
                        
                        current_week_start += timedelta(days=7)
                    
                    # Batch write all schedules ONCE to avoid multiple GitHub API calls
                    if schedules_by_week:
                        staging_schedules = staging_store.get_schedules()
                        updated_schedules = [s for s in staging_schedules if str(s.get('week_start_date')) not in schedules_by_week]
                        for schedule in schedules_by_week.values():
                            updated_schedules.append(schedule.model_dump())
                        staging_store.set_schedules(updated_schedules, user_id=user.get('employee_id'), immediate=True)
                        print(f"[REJECTION] Batch wrote {len(schedules_by_week)} schedules with locked shifts removed")
                        
                except Exception as e:
                    print(f"[REJECTION] Warning: could not remove locked shifts: {e}")
                    import traceback
                    traceback.print_exc()
        except Exception as e:
            print(f"[REJECTION] Warning: error in locked shift removal: {e}")
            import traceback
            traceback.print_exc()

        # Remove manager notifications for this request after rejection
        try:
            notifications = staging_store.get_notifications()
            cleaned = [
                n for n in notifications
                if not (
                    n.get('type') == NotificationType.AVAILABILITY_REQUEST and
                    n.get('details', {}).get('request_id') == request_id
                )
            ]
            if len(cleaned) < len(notifications):
                staging_store.set_notifications(cleaned, user_id=user.get('employee_id'), immediate=True)
                print(f"[REJECTION] Removed {len(notifications) - len(cleaned)} manager notifications for request {request_id}")
        except Exception as e:
            print(f"[REJECTION] Warning: could not clean up manager notifications: {e}")

        # Add to action queue for Excel sync (async, no blocking)
        # No action queue needed - GitHub JSON is single source of truth

        # Support both old and new model for notification message
        request_type = request_data.get('request_type', 'availability')
        start_date = request_data.get('start_date', request_data.get('week_start_date', ''))
        end_date = request_data.get('end_date', start_date)

        # Send notification to employee
        try:
            comment_part = f" Reason: {body.get('comment')}" if body.get('comment') else ""
            notification = {
                'id': str(uuid.uuid4()),
                'employee_id': request_data['employee_id'],
                'type': NotificationType.AVAILABILITY_REJECTED,
                'message': f"Your {request_type} request ({start_date} to {end_date}) was not approved.{comment_part}",
                'details': {
                    'request_type': request_type,
                    'start_date': start_date,
                    'end_date': end_date,
                    'days_of_week': ', '.join(request_data.get('days_of_week', [])) if request_data.get('days_of_week') else 'All days',
                    'time_range': f"{request_data.get('start_time', '00:00')} - {request_data.get('end_time', '23:59')}" if request_data.get('start_time') else 'All day',
                    'employee_comment': request_data.get('employee_comment', '')
                },
                'created_at': datetime.now(),
                'read': False
            }
            # Update staging layer
            notifications = staging_store.get_notifications()
            notifications.append(notification)
            staging_store.set_notifications(notifications, user_id=user.get('employee_id'), immediate=True)
            
            # Add to action queue for Excel sync (async, no blocking)
            # No action queue needed - GitHub JSON is single source of truth
        except Exception as e:
            import traceback
            print(f"[API] Warning: could not send notification: {e}")
            traceback.print_exc()

        # Ensure response is JSON-serializable
        response = request_data.copy()
        if isinstance(response.get('created_at'), datetime):
            response['created_at'] = response['created_at'].isoformat()
        if isinstance(response.get('updated_at'), datetime):
            response['updated_at'] = response['updated_at'].isoformat()
        if isinstance(response.get('approved_at'), datetime):
            response['approved_at'] = response['approved_at'].isoformat()
        return response

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"[API] Error rejecting availability request: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error rejecting request: {str(e)}")

# ============ Notifications ============

@app.get("/api/notifications")
async def get_employee_notifications(authorization: str = Header(None)):
    """Get notifications for the current user"""
    user = AuthManager.get_current_user(authorization)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    # Read from GitHub JSON (single source of truth)
    staging_notifications = staging_store.get_notifications()
    user_notifications = [n for n in staging_notifications if n.get('employee_id') == user['employee_id']]
    return user_notifications


@app.get("/api/sync-status/{request_id}")
async def get_request_sync_status(request_id: str, authorization: str = Header(None)):
    """Get the sync status for a specific request"""
    user = AuthManager.get_current_user(authorization)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    status = get_sync_status(request_id)
    if status:
        return {"request_id": request_id, "status": status}
    else:
        return {"request_id": request_id, "status": "unknown"}

@app.put("/api/notifications/{notification_id}/read")
async def mark_notification_as_read(notification_id: str, authorization: str = Header(None)):
    """Mark a notification as read"""
    user = AuthManager.get_current_user(authorization)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    # Update staging layer
    notifications = staging_store.get_notifications()
    notification_found = False
    for notif in notifications:
        if notif.get('id') == notification_id:
            notif['read'] = True
            notification_found = True
            break
    
    if notification_found:
        staging_store.set_notifications(notifications, user_id=user.get('employee_id'), immediate=True)
        # Add to action queue for Excel sync
        # No action queue needed - GitHub JSON is single source of truth
        return {"success": True}
    
    raise HTTPException(status_code=404, detail="Notification not found")

@app.put("/api/notifications/read-all")
async def mark_all_notifications_as_read(authorization: str = Header(None)):
    """Mark all notifications as read for the current user"""
    user = AuthManager.get_current_user(authorization)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    # Update staging layer
    notifications = staging_store.get_notifications()
    updated_count = 0
    for notif in notifications:
        if notif.get('employee_id') == user.get('employee_id') and not notif.get('read'):
            notif['read'] = True
            updated_count += 1
    
    if updated_count > 0:
        staging_store.set_notifications(notifications, user_id=user.get('employee_id'), immediate=True)
        # Add to action queue for Excel sync
        # No action queue needed - GitHub JSON is single source of truth
    
    return {"success": True, "updated_count": updated_count}

@app.post("/api/notifications/cleanup-processed")
async def cleanup_processed_notifications(authorization: str = Header(None)):
    """Remove AVAILABILITY_REQUEST notifications for requests that are already approved or rejected"""
    user = AuthManager.get_current_user(authorization)
    if not user or user.get('role') not in ['manager', 'admin']:
        raise HTTPException(status_code=403, detail="Manager access required")
    
    try:
        from datetime import timedelta
        
        # Get all requests to check their status
        requests = staging_store.get_availability_requests()
        processed_request_ids = {
            r['id'] for r in requests 
            if r.get('status') in ['approved', 'rejected', AvailabilityRequestStatus.APPROVED, AvailabilityRequestStatus.REJECTED]
        }
        
        # Get all notifications
        notifications = staging_store.get_notifications()
        
        # Remove AVAILABILITY_REQUEST notifications for processed requests OR older than 7 days
        cutoff_date = datetime.now() - timedelta(days=7)
        cleaned = [
            n for n in notifications
            if not (
                n.get('type') == NotificationType.AVAILABILITY_REQUEST and
                (
                    n.get('details', {}).get('request_id') in processed_request_ids or
                    (isinstance(n.get('created_at'), datetime) and n.get('created_at') < cutoff_date)
                )
            )
        ]
        
        removed_count = len(notifications) - len(cleaned)
        
        if removed_count > 0:
            staging_store.set_notifications(cleaned, user_id=user.get('employee_id'), immediate=True)
            print(f"[CLEANUP] Removed {removed_count} processed/old request notifications")
        
        return {"success": True, "removed_count": removed_count}
        
    except Exception as e:
        print(f"[CLEANUP] Error cleaning up notifications: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error cleaning up notifications: {str(e)}")

@app.post("/api/schedules/cleanup-orphaned-locked-shifts")
async def cleanup_orphaned_locked_shifts(authorization: str = Header(None)):
    """Remove pending locked shifts that don't have corresponding pending availability requests"""
    user = AuthManager.get_current_user(authorization)
    if not user or user.get('role') not in ['manager', 'admin']:
        raise HTTPException(status_code=403, detail="Manager access required")
    
    try:
        # Get all pending request IDs
        requests = staging_store.get_availability_requests()
        pending_request_ids = {
            r['id'] for r in requests 
            if r.get('status') in ['pending', AvailabilityRequestStatus.PENDING]
        }
        
        # Get all schedules
        schedules = staging_store.get_schedules()
        schedules_by_week = {}
        total_removed = 0
        
        for schedule_dict in schedules:
            schedule = WeeklySchedule(**schedule_dict)
            original_count = len(schedule.shifts)
            
            # Remove pending locked shifts that don't have corresponding pending requests
            schedule.shifts = [
                s for s in schedule.shifts
                if not (s.locked and s.id.startswith('pending_') and not any(
                    s.id.startswith(f"pending_{req_id}_") for req_id in pending_request_ids
                ))
            ]
            
            removed_count = original_count - len(schedule.shifts)
            if removed_count > 0:
                schedules_by_week[str(schedule.week_start_date)] = schedule
                total_removed += removed_count
                print(f"[CLEANUP] Removed {removed_count} orphaned pending locked shifts from week {schedule.week_start_date}")
        
        # Batch write all updated schedules
        if schedules_by_week:
            updated_schedules = [s for s in schedules if str(s.get('week_start_date')) not in schedules_by_week]
            for schedule in schedules_by_week.values():
                updated_schedules.append(schedule.model_dump())
            staging_store.set_schedules(updated_schedules, user_id=user.get('employee_id'), immediate=True)
            print(f"[CLEANUP] Batch wrote {len(schedules_by_week)} schedules with orphaned shifts removed")
        
        return {"success": True, "removed_count": total_removed, "weeks_updated": len(schedules_by_week)}
        
    except Exception as e:
        print(f"[CLEANUP] Error cleaning up orphaned locked shifts: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error cleaning up orphaned shifts: {str(e)}")

@app.post("/api/schedules/cleanup-orphaned-locked-shifts/no-auth")
async def cleanup_orphaned_locked_shifts_no_auth():
    """TEMPORARY: Remove pending locked shifts without auth requirement for emergency cleanup"""
    try:
        # Get all pending request IDs
        requests = staging_store.get_availability_requests()
        pending_request_ids = {
            r['id'] for r in requests 
            if r.get('status') in ['pending', AvailabilityRequestStatus.PENDING]
        }
        
        # Get all schedules
        schedules = staging_store.get_schedules()
        schedules_by_week = {}
        total_removed = 0
        
        for schedule_dict in schedules:
            schedule = WeeklySchedule(**schedule_dict)
            original_count = len(schedule.shifts)
            
            # Remove pending locked shifts that don't have corresponding pending requests
            schedule.shifts = [
                s for s in schedule.shifts
                if not (s.locked and s.id.startswith('pending_') and not any(
                    s.id.startswith(f"pending_{req_id}_") for req_id in pending_request_ids
                ))
            ]
            
            removed_count = original_count - len(schedule.shifts)
            if removed_count > 0:
                schedules_by_week[str(schedule.week_start_date)] = schedule
                total_removed += removed_count
                print(f"[CLEANUP] Removed {removed_count} orphaned pending locked shifts from week {schedule.week_start_date}")
        
        # Batch write all updated schedules
        if schedules_by_week:
            updated_schedules = [s for s in schedules if str(s.get('week_start_date')) not in schedules_by_week]
            for schedule in schedules_by_week.values():
                updated_schedules.append(schedule.model_dump())
            staging_store.set_schedules(updated_schedules, user_id="system", immediate=True)
            print(f"[CLEANUP] Batch wrote {len(schedules_by_week)} schedules with orphaned shifts removed")
        
        return {"success": True, "removed_count": total_removed, "weeks_updated": len(schedules_by_week)}
        
    except Exception as e:
        print(f"[CLEANUP] Error cleaning up orphaned locked shifts: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error cleaning up orphaned shifts: {str(e)}")

# ============ Events ============

@app.get("/api/events/list")
async def get_week_events():
    """Get all events"""
    try:
        # Read from GitHub JSON (single source of truth)
        events_dict = staging_store.get_events()
        events = [Event(**e) for e in events_dict]
        
        # Convert to dict with string dates
        return [
            {
                "id": e.id,
                "name": e.name,
                "week_start_date": e.week_start_date.isoformat() if isinstance(e.week_start_date, date) else str(e.week_start_date),
                "date": e.date.isoformat() if isinstance(e.date, date) else str(e.date),
                "start_time": e.start_time,
                "end_time": e.end_time,
                "location": e.location,
                "people_needed": e.people_needed,
                "description": e.description,
                "created_by": e.created_by,
                "created_at": e.created_at.isoformat() if isinstance(e.created_at, datetime) else str(e.created_at),
                "updated_at": e.updated_at.isoformat() if e.updated_at and isinstance(e.updated_at, datetime) else str(e.updated_at) if e.updated_at else None
            }
            for e in events
        ]
    except Exception as e:
        print(f"Error in get_week_events: {e}")
        return []

@app.post("/api/events")
async def create_event(event_data: dict, authorization: str = Header(None)):
    """Create a new event"""
    user = require_manager(authorization)
    
    # Auto-generate ID if not provided
    if not event_data.get('id'):
        event_data['id'] = f"event_{uuid.uuid4().hex[:8]}"
    
    # Set created_by and created_at if not provided
    if not event_data.get('created_by'):
        event_data['created_by'] = user['employee_id']
    if not event_data.get('created_at'):
        event_data['created_at'] = datetime.now()
    if 'people_needed' not in event_data:
        event_data['people_needed'] = 0
    
    # Convert string dates to date/datetime objects
    if isinstance(event_data.get('week_start_date'), str):
        event_data['week_start_date'] = date.fromisoformat(event_data['week_start_date'])
    if isinstance(event_data.get('date'), str):
        event_data['date'] = date.fromisoformat(event_data['date'])
    
    # Create Event object
    event = Event(**event_data)
    
    # Save to GitHub JSON (single source of truth)
    events = staging_store.get_events()
    events.append(event.model_dump())
    staging_store.set_events(events, user_id=user.get('employee_id'))
    
    return event

@app.put("/api/events/{event_id}")
async def update_event(event_id: str, event_data: dict, authorization: str = Header(None)):
    """Update an existing event"""
    user = require_manager(authorization)
    
    # Ensure the event ID matches
    event_data['id'] = event_id
    event_data['updated_at'] = datetime.now()
    
    # Convert string dates to date objects
    if isinstance(event_data.get('week_start_date'), str):
        event_data['week_start_date'] = date.fromisoformat(event_data['week_start_date'])
    if isinstance(event_data.get('date'), str):
        event_data['date'] = date.fromisoformat(event_data['date'])
    
    event = Event(**event_data)
    
    # Update in GitHub JSON (single source of truth)
    events = staging_store.get_events()
    events = [e if e['id'] != event_id else event.model_dump() for e in events]
    staging_store.set_events(events, user_id=user.get('employee_id'))
    
    return event

@app.get("/api/cron/nightly-excel-commit")
async def nightly_excel_commit():
    """Cron job: Generate Excel from GitHub JSON and commit to GitHub data branch"""
    try:
        import io
        from openpyxl import Workbook
        from datetime import datetime
        
        # Create Excel workbook from GitHub JSON data
        wb = Workbook()
        
        # Employees sheet
        emp_sheet = wb.active
        emp_sheet.title = "Employees"
        emp_sheet.append(["ID", "Name", "Email", "Type", "Max Hours", "Active", "Created At"])
        employees = staging_store.get_employees()
        for emp in employees:
            emp_sheet.append([
                emp.get('id'), emp.get('name'), emp.get('email') or "", emp.get('employee_type'),
                emp.get('max_hours_per_week'), emp.get('active'), emp.get('created_at')
            ])
        
        # Schedules sheet
        schedules = staging_store.get_schedules()
        if schedules:
            sched_sheet = wb.create_sheet("Schedules")
            sched_sheet.append(["Week Start", "Schedule ID", "Employee ID", "Day", "Location", "Start Time", "End Time", "Hours", "Break Provided"])
            for schedule in schedules:
                for shift in schedule.get('shifts', []):
                    sched_sheet.append([
                        schedule.get('week_start_date'), schedule.get('id'), shift.get('employee_id'), shift.get('day_of_week'),
                        shift.get('floor'), shift.get('start_time'), shift.get('end_time'), shift.get('hours'), shift.get('break_provided')
                    ])
        
        # Availability Requests sheet
        requests = staging_store.get_availability_requests()
        if requests:
            req_sheet = wb.create_sheet("Availability Requests")
            req_sheet.append(["ID", "Employee ID", "Request Type", "Start Date", "End Date", "Days of Week", "Start Time", "End Time", "Status", "Created At"])
            for req in requests:
                req_sheet.append([
                    req.get('id'), req.get('employee_id'), req.get('request_type'),
                    req.get('start_date'), req.get('end_date'), ', '.join(req.get('days_of_week', [])),
                    req.get('start_time'), req.get('end_time'), req.get('status'), req.get('created_at')
                ])
        
        # Notifications sheet
        notifications = staging_store.get_notifications()
        if notifications:
            notif_sheet = wb.create_sheet("Notifications")
            notif_sheet.append(["ID", "Employee ID", "Type", "Message", "Read", "Created At"])
            for notif in notifications:
                notif_sheet.append([
                    notif.get('id'), notif.get('employee_id'), notif.get('type'),
                    notif.get('message'), notif.get('read'), notif.get('created_at')
                ])
        
        # Config sheet
        config_sheet = wb.create_sheet("Config")
        config = staging_store.get_system_config()
        config_sheet.append(["Setting", "Value"])
        for key, value in config.items():
            config_sheet.append([key, str(value)])
        
        # Events sheet
        events = config.get('events', [])
        if events:
            event_sheet = wb.create_sheet("Events")
            event_sheet.append(["ID", "Name", "Week Start Date", "Date", "Start Time", "End Time", "Location", "People Needed", "Description", "Created By"])
            for event in events:
                event_sheet.append([
                    event.get('id'), event.get('name'), event.get('week_start_date'),
                    event.get('date'), event.get('start_time'), event.get('end_time'),
                    event.get('location'), event.get('people_needed'), event.get('description'),
                    event.get('created_by')
                ])
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        excel_data = buffer.getvalue()
        
        # Commit to GitHub data branch
        from github_storage import GitHubStorage
        github = GitHubStorage()
        
        commit_message = f"Nightly Excel auto-commit - {datetime.now().isoformat()}"
        success = github.write_file(
            "ibu_schedule.xlsx",
            excel_data,
            commit_message=commit_message
        )
        
        if success:
            return {
                "success": True,
                "message": "Excel file generated and committed to GitHub",
                "timestamp": datetime.now().isoformat()
            }
        else:
            return {
                "success": False,
                "message": "Failed to commit Excel file to GitHub",
                "timestamp": datetime.now().isoformat()
            }
    except Exception as e:
        import traceback
        print(f"[CRON] Error in nightly Excel commit: {e}")
        traceback.print_exc()
        return {
            "success": False,
            "message": str(e),
            "timestamp": datetime.now().isoformat()
        }

@app.delete("/api/events/{event_id}")
async def remove_event(event_id: str, authorization: str = Header(None)):
    """Delete an event"""
    user = require_manager(authorization)
    
    # Delete from GitHub JSON (single source of truth)
    events = staging_store.get_events()
    original_count = len(events)
    events = [e for e in events if e['id'] != event_id]
    if len(events) < original_count:
        staging_store.set_events(events, user_id=user.get('employee_id'))
        return {"success": True}
    
    raise HTTPException(status_code=404, detail="Event not found")

# ============ Health Check (simple) ============

@app.get("/api/ping")
async def health_ping():
    return {"status": "healthy", "timestamp": datetime.now()}


@app.get("/api/cache/stats")
async def get_cache_stats():
    """Get cache statistics for monitoring."""
    try:
        import cache_manager
        stats = cache_manager.get_cache_stats()
        return {
            "success": True,
            "stats": stats
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


@app.post("/api/cache/clear")
async def clear_cache(authorization: str = Header(None)):
    """Clear the cache (manager only)."""
    user = require_manager(authorization)
    try:
        import cache_manager
        cache_manager.clear_cache()
        return {
            "success": True,
            "message": "Cache cleared successfully"
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


@app.get("/api/audit-logs")
async def get_audit_logs(
    entity_type: Optional[str] = None,
    user_id: Optional[str] = None,
    operation: Optional[str] = None,
    limit: int = 100,
    authorization: str = Header(None)
):
    """Get audit logs with optional filtering (manager only)."""
    user = require_manager(authorization)
    try:
        import audit_logger
        logs = audit_logger.get_audit_logs(
            entity_type=entity_type,
            user_id=user_id,
            operation=operation,
            limit=limit
        )
        return {
            "success": True,
            "logs": logs,
            "count": len(logs)
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


@app.get("/api/audit-logs/stats")
async def get_audit_log_stats(authorization: str = Header(None)):
    """Get audit log statistics (manager only)."""
    user = require_manager(authorization)
    try:
        import audit_logger
        stats = audit_logger.get_audit_log_stats()
        return {
            "success": True,
            "stats": stats
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


@app.post("/api/audit-logs/flush")
async def flush_audit_log(authorization: str = Header(None)):
    """Manually flush the audit log buffer (manager only)."""
    user = require_manager(authorization)
    try:
        import audit_logger
        result = audit_logger.flush_audit_log()
        return {
            "success": result,
            "message": "Audit log flushed" if result else "Failed to flush audit log"
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

# ============ Admin: Reset Blob Data ============

@app.post("/api/admin/reset-blob-data")
async def reset_blob_data(authorization: str = Header(None)):
    """Reset blob storage with local JSON data (for clearing stale data)"""
    user = require_manager(authorization)
    
    from storage import blob_put
    from pathlib import Path
    import json
    
    data_dir = Path(__file__).parent / "data"
    json_files = [
        "employees.json",
        "passwords.json",
        "availability_requests.json",
        "events.json",
        "coverage_requirements.json",
        "notifications.json",
    ]
    
    results = {}
    for filename in json_files:
        filepath = data_dir / filename
        if not filepath.exists():
            results[filename] = "skipped (not found)"
            continue
        
        with open(filepath, 'r') as f:
            data = f.read()
        
        success = blob_put(filename, data.encode('utf-8'))
        results[filename] = "success" if success else "failed"
    
    return {"success": True, "results": results}















@app.post("/api/log-error")
async def log_frontend_error(error_data: dict):
    """Log frontend errors for monitoring"""
    message = f"{error_data.get('type', 'unknown')}: {error_data.get('message', 'no message')}"
    url = error_data.get('url', 'N/A')
    component = error_data.get('component', 'N/A')
    
    print(f"[FRONTEND ERROR] {message}")
    print(f"  URL: {url}")
    print(f"  Method: {error_data.get('method', 'N/A')}")
    print(f"  Status: {error_data.get('status', 'N/A')}")
    
    # Store in log storage
    log_storage.get_log_storage().add_frontend_log(
        message=message,
        level="ERROR",
        url=url,
        component=component
    )
    
    return {"success": True}

# ============ Monitoring Dashboard Endpoints ============

def verify_admin_auth(username: str, password: str) -> bool:
    """Verify admin credentials."""
    return username == "admin" and password == "ibu-admin-secret-2026"

@app.get("/admin/login")
async def admin_login():
    """Serve admin login page."""
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Admin Login - IBU Operations</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f5f5; display: flex; justify-content: center; align-items: center; min-height: 100vh; }
        .login-container { background: white; padding: 40px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); width: 100%; max-width: 400px; }
        h1 { color: #333; margin-bottom: 20px; text-align: center; }
        .form-group { margin-bottom: 20px; }
        label { display: block; color: #666; margin-bottom: 8px; font-weight: 500; }
        input { width: 100%; padding: 12px; border: 1px solid #ddd; border-radius: 4px; font-size: 14px; }
        input:focus { outline: none; border-color: #007bff; }
        button { width: 100%; padding: 12px; background: #007bff; color: white; border: none; border-radius: 4px; font-size: 14px; font-weight: 500; cursor: pointer; }
        button:hover { background: #0056b3; }
        .error { color: #dc3545; text-align: center; margin-bottom: 20px; padding: 10px; background: #f8d7da; border-radius: 4px; display: none; }
    </style>
</head>
<body>
    <div class="login-container">
        <h1>Admin Login</h1>
        <div id="error" class="error">Invalid credentials</div>
        <form id="loginForm">
            <div class="form-group">
                <label for="username">Username</label>
                <input type="text" id="username" name="username" required autofocus>
            </div>
            <div class="form-group">
                <label for="password">Password</label>
                <input type="password" id="password" name="password" required>
            </div>
            <button type="submit">Login</button>
        </form>
    </div>
    <script>
        document.getElementById('loginForm').addEventListener('submit', async (e) => {
            e.preventDefault();
            const username = document.getElementById('username').value;
            const password = document.getElementById('password').value;
            
            const response = await fetch('/admin/api/login', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ username, password })
            });
            
            if (response.ok) {
                window.location.href = '/admin/dashboard';
            } else {
                document.getElementById('error').style.display = 'block';
            }
        });
    </script>
</body>
</html>
    """
    return HTMLResponse(content=html_content)

@app.post("/admin/api/login")
async def admin_api_login(credentials: dict, request: Request):
    """Handle admin login."""
    username = credentials.get('username')
    password = credentials.get('password')
    
    if not verify_admin_auth(username, password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    response = JSONResponse(content={"success": True})
    # Only use secure=True for HTTPS, not for localhost HTTP
    is_secure = request.url.scheme == "https"
    response.set_cookie(key="admin_session", value="authenticated", httponly=True, secure=is_secure, samesite="lax")
    return response

@app.get("/admin/logout")
async def admin_logout():
    """Handle admin logout."""
    response = RedirectResponse(url="/admin/login")
    response.delete_cookie(key="admin_session")
    return response

@app.get("/admin/api/flow-diagram")
async def get_flow_diagram(admin_session: Optional[str] = Cookie(None)):
    """Get recent flow chains for the flow diagram."""
    if admin_session != "authenticated":
        return {"error": "Unauthorized"}
    
    flow = flow_storage.get_flow_storage()
    chains = flow.get_recent_chains(limit=10)
    return {"chains": chains}

@app.get("/admin/dashboard")
async def admin_dashboard(admin_session: Optional[str] = Cookie(None)):
    """Serve admin dashboard HTML page."""
    if admin_session != "authenticated":
        return RedirectResponse(url="/admin/login")
    
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>IBU Operations - System Monitoring Dashboard</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f5f5; padding: 20px; }
        .container { max-width: 1400px; margin: 0 auto; }
        h1 { color: #333; margin-bottom: 20px; }
        .status-banner { padding: 15px; margin-bottom: 20px; border-radius: 5px; font-weight: bold; }
        .status-banner.green { background: #d4edda; color: #155724; }
        .status-banner.red { background: #f8d7da; color: #721c24; }
        .grid { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px; margin-bottom: 20px; }
        .card { background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        .card.full-width { grid-column: 1 / -1; }
        .card h3 { color: #555; margin: 15px 0 10px 0; font-size: 15px; border-bottom: 1px solid #eee; padding-bottom: 5px; }
        .card h3:first-child { margin-top: 0; }
        .card h2 { color: #333; margin-bottom: 15px; font-size: 18px; }
        .status-indicator { display: inline-block; width: 12px; height: 12px; border-radius: 50%; margin-right: 8px; }
        .status-indicator.green { background: #28a745; }
        .status-indicator.red { background: #dc3545; }
        .card h2 .status-indicator { float: right; margin-top: 4px; }
        .metric { display: flex; justify-content: space-between; margin-bottom: 10px; }
        .metric-label { color: #666; }
        .metric-value { font-weight: bold; color: #333; }
        .log-container { background: #1e1e1e; color: #d4d4d4; padding: 15px; border-radius: 5px; max-height: 360px; overflow-y: auto; font-family: monospace; font-size: 12px; line-height: 1.5; }
        .logout-btn { background: #dc3545; color: white; border: none; padding: 8px 16px; border-radius: 4px; cursor: pointer; font-size: 13px; text-decoration: none; }
        .logout-btn:hover { background: #c82333; }
        .log-entry { margin-bottom: 5px; padding: 5px; border-bottom: 1px solid #333; }
        .log-entry.error { color: #f8d7da; }
        .log-entry.warning { color: #fff3cd; }
        .log-entry.info { color: #d1ecf1; }
        .timestamp { color: #888; margin-right: 10px; }
        .level { margin-right: 10px; font-weight: bold; }
        .empty-logs { color: #888; font-style: italic; padding: 20px; text-align: center; }
        .flow-node { display: inline-block; padding: 8px 12px; margin: 5px; border-radius: 4px; font-size: 12px; border: 1px solid #ddd; background: white; }
        .flow-node.api { background: #e3f2fd; border-color: #2196f3; }
        .flow-node.github { background: #e8f5e9; border-color: #4caf50; }
        .flow-node.cache { background: #fff3e0; border-color: #ff9800; }
        .flow-node.success { border-color: #4caf50; }
        .flow-node.error { border-color: #f44336; background: #ffebee; }
        .flow-arrow { display: inline-block; margin: 0 5px; color: #666; }
    </style>
</head>
<body>
    <div class="container">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
            <h1>IBU Operations - System Monitoring Dashboard</h1>
            <a href="/admin/logout" class="logout-btn">Logout</a>
        </div>
        <div style="margin-bottom:15px; color:#666; font-size:13px;">Auto-refreshes every 5 minutes &nbsp;|&nbsp; Last updated: <span id="last-updated">-</span></div>
        
        <div class="grid">
            <div class="card">
                <h2>Backend Status <span id="backend-frontend-status" class="status-indicator green"></span></h2>
                <h3>Backend</h3>
                <div class="metric">
                    <span class="metric-label">Status:</span>
                    <span class="metric-value"><span class="status-indicator green"></span>Running</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Start Time:</span>
                    <span class="metric-value" id="backend-start-time">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Recent Errors:</span>
                    <span class="metric-value" id="backend-errors">-</span>
                </div>
            </div>
            
            <div class="card">
                <h2>GitHub Health <span id="github-status" class="status-indicator green"></span></h2>
                <div class="metric">
                    <span class="metric-label">API Status:</span>
                    <span class="metric-value" id="github-api-status">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Branch:</span>
                    <span class="metric-value" id="github-branch">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Rate Limit:</span>
                    <span class="metric-value" id="github-rate-limit">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Last Commit:</span>
                    <span class="metric-value" id="github-last-commit">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">employees.json:</span>
                    <span class="metric-value" id="file-employees">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">schedules.json:</span>
                    <span class="metric-value" id="file-schedules">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">availability_requests.json:</span>
                    <span class="metric-value" id="file-requests">-</span>
                </div>
            </div>
            
            <div class="card">
                <h2>System Metrics <span id="system-status" class="status-indicator green"></span></h2>
                <div class="metric">
                    <span class="metric-label">Cache Hit Rate:</span>
                    <span class="metric-value" id="cache-hit-rate">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Cache Hits:</span>
                    <span class="metric-value" id="cache-hits">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Cache Misses:</span>
                    <span class="metric-value" id="cache-misses">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Active Sessions:</span>
                    <span class="metric-value" id="active-sessions">-</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Error Rate:</span>
                    <span class="metric-value" id="error-rate">-</span>
                </div>
            </div>
        </div>
        
        <div class="grid">
            <div class="card full-width">
                <h2>System Architecture</h2>
                <div id="architecture-diagram" style="min-height: 900px; background: #f9f9f9; border-radius: 5px; padding: 20px; overflow-x: auto;">
                    <svg width="1400" height="900" viewBox="0 0 1400 900" xmlns="http://www.w3.org/2000/svg">
                        <!-- Styles -->
                        <defs>
                            <style>
                                .box { fill: white; stroke: #333; stroke-width: 2; }
                                .box-frontend { fill: #e3f2fd; stroke: #2196f3; }
                                .box-api { fill: #fff3e0; stroke: #ff9800; }
                                .box-cache { fill: #f3e5f5; stroke: #9c27b0; }
                                .box-github { fill: #e8f5e9; stroke: #4caf50; }
                                .box-data { fill: #fce4ec; stroke: #e91e63; }
                                .box-auth { fill: #fff9c4; stroke: #fbc02d; }
                                .box-scheduler { fill: #e0f7fa; stroke: #00bcd4; }
                                .box-excel { fill: #ffccbc; stroke: #ff5722; }
                                .arrow-read { stroke: #2196f3; stroke-width: 2; fill: none; marker-end: url(#arrowhead-read); }
                                .arrow-write { stroke: #f44336; stroke-width: 2; fill: none; marker-end: url(#arrowhead-write); }
                                .arrow-bidi { stroke: #666; stroke-width: 2; fill: none; marker-end: url(#arrowhead); marker-start: url(#arrowhead-start); }
                                .label { font-family: sans-serif; font-size: 14px; fill: #333; text-anchor: middle; }
                                .label-title { font-weight: bold; font-size: 16px; }
                                .label-small { font-size: 13px; }
                            </style>
                            <marker id="arrowhead" markerWidth="10" markerHeight="7" refX="9" refY="3.5" orient="auto">
                                <polygon points="0 0, 10 3.5, 0 7" fill="#666" />
                            </marker>
                            <marker id="arrowhead-start" markerWidth="10" markerHeight="7" refX="1" refY="3.5" orient="auto">
                                <polygon points="10 0, 0 3.5, 10 7" fill="#666" />
                            </marker>
                            <marker id="arrowhead-read" markerWidth="10" markerHeight="7" refX="9" refY="3.5" orient="auto">
                                <polygon points="0 0, 10 3.5, 0 7" fill="#2196f3" />
                            </marker>
                            <marker id="arrowhead-write" markerWidth="10" markerHeight="7" refX="9" refY="3.5" orient="auto">
                                <polygon points="0 0, 10 3.5, 0 7" fill="#f44336" />
                            </marker>
                        </defs>
                        
                        <!-- Frontend -->
                        <rect x="50" y="50" width="150" height="100" class="box box-frontend" rx="5" />
                        <text x="125" y="95" class="label label-title">Frontend</text>
                        <text x="125" y="115" class="label label-small">(React)</text>
                        
                        <!-- Auth Layer -->
                        <rect x="230" y="50" width="150" height="100" class="box box-auth" rx="5" />
                        <text x="305" y="95" class="label label-title">Auth</text>
                        <text x="305" y="115" class="label label-small">Session/Password</text>
                        
                        <!-- API Layer - Container -->
                        <rect x="420" y="30" width="200" height="520" class="box box-api" rx="5" />
                        <text x="520" y="60" class="label label-title">FastAPI Backend</text>
                        
                        <!-- API Read Endpoints -->
                        <rect x="440" y="80" width="160" height="110" class="box" rx="3" stroke="#ff9800" stroke-width="1" />
                        <text x="520" y="105" class="label label-title" style="font-size: 14px;">Read Endpoints</text>
                        <text x="520" y="125" class="label label-small">GET /api/availabilities</text>
                        <text x="520" y="142" class="label label-small">GET /api/notifications</text>
                        <text x="520" y="159" class="label label-small">GET /api/requests</text>
                        <text x="520" y="176" class="label label-small">GET /api/schedules</text>
                        
                        <!-- API Write Endpoints -->
                        <rect x="440" y="210" width="160" height="110" class="box" rx="3" stroke="#ff9800" stroke-width="1" />
                        <text x="520" y="235" class="label label-title" style="font-size: 14px;">Write Endpoints</text>
                        <text x="520" y="255" class="label label-small">POST /api/availabilities</text>
                        <text x="520" y="272" class="label label-small">POST /api/requests</text>
                        <text x="520" y="289" class="label label-small">POST /api/schedules</text>
                        <text x="520" y="306" class="label label-small">POST /api/staffing-targets</text>
                        
                        <!-- API Auth Endpoints -->
                        <rect x="440" y="340" width="160" height="80" class="box" rx="3" stroke="#ff9800" stroke-width="1" />
                        <text x="520" y="365" class="label label-title" style="font-size: 14px;">Auth Endpoints</text>
                        <text x="520" y="385" class="label label-small">POST /api/auth/login</text>
                        <text x="520" y="402" class="label label-small">POST /admin/api/login</text>
                        
                        <!-- API Admin/Scheduler Endpoints -->
                        <rect x="440" y="440" width="160" height="100" class="box" rx="3" stroke="#ff9800" stroke-width="1" />
                        <text x="520" y="465" class="label label-title" style="font-size: 14px;">Admin/Scheduler</text>
                        <text x="520" y="485" class="label label-small">POST /api/schedules/*</text>
                        <text x="520" y="502" class="label label-small">POST /api/shift-templates</text>
                        <text x="520" y="519" class="label label-small">GET /api/health</text>
                        
                        <!-- Cache Layer - Container -->
                        <rect x="660" y="30" width="180" height="520" class="box box-cache" rx="5" />
                        <text x="750" y="60" class="label label-title">Cache Manager</text>
                        
                        <!-- Cache Storage -->
                        <rect x="680" y="80" width="140" height="90" class="box" rx="3" stroke="#9c27b0" stroke-width="1" />
                        <text x="750" y="105" class="label label-title" style="font-size: 14px;">Cache Storage</text>
                        <text x="750" y="125" class="label label-small">LRU Cache (max 100)</text>
                        <text x="750" y="142" class="label label-small">TTL: 5 minutes</text>
                        <text x="750" y="159" class="label label-small">Key Locks</text>
                        
                        <!-- Cache Invalidation -->
                        <rect x="680" y="190" width="140" height="90" class="box" rx="3" stroke="#9c27b0" stroke-width="1" />
                        <text x="750" y="215" class="label label-title" style="font-size: 14px;">Invalidation</text>
                        <text x="750" y="235" class="label label-small">SHA Invalidation</text>
                        <text x="750" y="252" class="label label-small">Manual cache.set()</text>
                        <text x="750" y="269" class="label label-small">after write</text>
                        
                        <!-- Cache Statistics -->
                        <rect x="680" y="300" width="140" height="70" class="box" rx="3" stroke="#9c27b0" stroke-width="1" />
                        <text x="750" y="325" class="label label-title" style="font-size: 14px;">Statistics</text>
                        <text x="750" y="345" class="label label-small">Cache hit/miss stats</text>
                        <text x="750" y="362" class="label label-small">Eviction tracking</text>
                        
                        <!-- GitHub Storage - Container -->
                        <rect x="880" y="30" width="200" height="520" class="box box-github" rx="5" />
                        <text x="980" y="60" class="label label-title">GitHub Storage</text>
                        
                        <!-- GitHub Read Operations -->
                        <rect x="900" y="80" width="160" height="100" class="box" rx="3" stroke="#4caf50" stroke-width="1" />
                        <text x="980" y="105" class="label label-title" style="font-size: 14px;">Read Operations</text>
                        <text x="980" y="125" class="label label-small">HEAD for SHA</text>
                        <text x="980" y="142" class="label label-small">GET for content</text>
                        <text x="980" y="159" class="label label-small">download_url for large</text>
                        <text x="980" y="176" class="label label-small">github_storage.py</text>
                        
                        <!-- GitHub Write Operations -->
                        <rect x="900" y="200" width="160" height="70" class="box" rx="3" stroke="#4caf50" stroke-width="1" />
                        <text x="980" y="225" class="label label-title" style="font-size: 14px;">Write Operations</text>
                        <text x="980" y="245" class="label label-small">PUT for write</text>
                        <text x="980" y="262" class="label label-small">Commit operations</text>
                        
                        <!-- GitHub Configuration -->
                        <rect x="900" y="290" width="160" height="100" class="box" rx="3" stroke="#4caf50" stroke-width="1" />
                        <text x="980" y="315" class="label label-title" style="font-size: 14px;">Configuration</text>
                        <text x="980" y="335" class="label label-small">Blob storage (disabled)</text>
                        <text x="980" y="352" class="label label-small">API authentication</text>
                        <text x="980" y="369" class="label label-small">Repository config</text>
                        
                        <!-- Excel Export -->
                        <rect x="880" y="580" width="200" height="80" class="box box-excel" rx="5" />
                        <text x="980" y="615" class="label label-title">Excel Export</text>
                        <text x="980" y="635" class="label label-small">data_store_excel.py</text>
                        <text x="980" y="652" class="label label-small">Daily reports only</text>
                        
                        <!-- Scheduler -->
                        <rect x="420" y="580" width="200" height="80" class="box box-scheduler" rx="5" />
                        <text x="520" y="615" class="label label-title">Scheduler</text>
                        <text x="520" y="635" class="label label-small">Auto-generation logic</text>
                        <text x="520" y="652" class="label label-small">Shift assignment</text>
                        
                        <!-- Core Data Files -->
                        <rect x="1120" y="30" width="220" height="180" class="box box-data" rx="5" />
                        <text x="1230" y="65" class="label label-title">Core Data Files</text>
                        <text x="1230" y="90" class="label label-small">employees.json</text>
                        <text x="1230" y="110" class="label label-small">schedules.json</text>
                        <text x="1230" y="130" class="label label-small">availability_requests.json</text>
                        <text x="1230" y="150" class="label label-small">notifications.json</text>
                        <text x="1230" y="170" class="label label-small">events.json</text>
                        
                        <!-- Configuration Files -->
                        <rect x="1120" y="230" width="220" height="160" class="box box-data" rx="5" />
                        <text x="1230" y="249" class="label label-title">Configuration</text>
                        <text x="1230" y="274" class="label label-small">system_config.json</text>
                        <text x="1230" y="294" class="label label-small">staffing_targets.json</text>
                        <text x="1230" y="314" class="label label-small">shift_templates.json</text>
                        <text x="1230" y="334" class="label label-small">skills.json</text>
                        <text x="1230" y="354" class="label label-small">roles.json</text>
                        <text x="1230" y="374" class="label label-small">passwords.json</text>
                        
                        <!-- Data Relationships -->
                        <rect x="1120" y="416" width="220" height="180" class="box box-data" rx="5" />
                        <text x="1230" y="438" class="label label-title" style="text-decoration: underline;">Relationships</text>
                        <text x="1135" y="463" class="label label-small" style="text-anchor: start;">availabilities → schedules</text>
                        <text x="1135" y="483" class="label label-small" style="text-anchor: start;">requests → availabilities</text>
                        <text x="1135" y="503" class="label label-small" style="text-anchor: start;">employees → schedules</text>
                        <text x="1135" y="523" class="label label-small" style="text-anchor: start;">employees → requests</text>
                        <text x="1135" y="543" class="label label-small" style="text-anchor: start;">templates → schedules</text>
                        <text x="1135" y="563" class="label label-small" style="text-anchor: start;">targets → schedules</text>
                        
                        <!-- Cache Invalidation Logic -->
                        <rect x="1120" y="616" width="220" height="140" class="box box-data" rx="5" />
                        <text x="1230" y="638" class="label label-title" style="text-decoration: underline;">Cache Invalidation</text>
                        <text x="1135" y="663" class="label label-small" style="text-anchor: start;">1. SHA mismatch → miss</text>
                        <text x="1135" y="683" class="label label-small" style="text-anchor: start;">2. TTL expired → miss</text>
                        <text x="1135" y="703" class="label label-small" style="text-anchor: start;">3. Manual clear() → miss</text>
                        <text x="1135" y="723" class="label label-small" style="text-anchor: start;">4. Write → manual set()</text>
                        <text x="1135" y="743" class="label label-small" style="text-anchor: start;">5. LRU eviction → miss</text>
                        
                        <!-- Auth Flow -->
                        <rect x="1120" y="776" width="220" height="120" class="box box-data" rx="5" />
                        <text x="1230" y="798" class="label label-title" style="text-decoration: underline;">Auth Flow</text>
                        <text x="1135" y="823" class="label label-small" style="text-anchor: start;">Login → passwords.json</text>
                        <text x="1135" y="843" class="label label-small" style="text-anchor: start;">Session → cookie</text>
                        <text x="1135" y="863" class="label label-small" style="text-anchor: start;">Manager → require_manager</text>
                        
                        <!-- Arrows - Bidirectional -->
                        <line x1="200" y1="100" x2="230" y2="100" class="arrow-bidi" />
                        <line x1="380" y1="100" x2="420" y2="100" class="arrow-bidi" />
                        
                        <!-- Arrows - Read (Blue) - Data Files to GitHub Read -->
                        <line x1="1110" y1="120" x2="1060" y2="130" class="arrow-read" />
                        <!-- GitHub Read to Cache Storage -->
                        <line x1="890" y1="130" x2="820" y2="125" class="arrow-read" />
                        <!-- Cache Storage to API Read -->
                        <line x1="670" y1="125" x2="600" y2="135" class="arrow-read" />
                        
                        <!-- Arrows - Write (Red) - API Write to Cache Invalidation -->
                        <line x1="600" y1="265" x2="670" y2="235" class="arrow-write" />
                        <!-- Cache Invalidation to GitHub Write -->
                        <line x1="820" y1="235" x2="890" y2="235" class="arrow-write" />
                        <!-- GitHub Write to Core Data Files -->
                        <line x1="1060" y1="235" x2="1110" y2="120" class="arrow-write" />
                        
                        <!-- Scheduler to API Admin -->
                        <line x1="520" y1="580" x2="520" y2="540" class="arrow-bidi" />
                        
                        <!-- Excel to Core Data Files -->
                        <line x1="1080" y1="620" x2="1110" y2="120" class="arrow-read" />
                        
                        <!-- Legend -->
                        <rect x="50" y="750" width="18" height="18" class="box box-frontend" />
                        <text x="80" y="765" class="label" style="text-anchor: start; font-size: 14px;">Frontend</text>
                        
                        <rect x="50" y="780" width="18" height="18" class="box box-auth" />
                        <text x="80" y="795" class="label" style="text-anchor: start; font-size: 14px;">Auth</text>
                        
                        <rect x="50" y="810" width="18" height="18" class="box box-api" />
                        <text x="80" y="825" class="label" style="text-anchor: start; font-size: 14px;">API Layer</text>
                        
                        <rect x="200" y="750" width="18" height="18" class="box box-cache" />
                        <text x="230" y="765" class="label" style="text-anchor: start; font-size: 14px;">Cache</text>
                        
                        <rect x="200" y="780" width="18" height="18" class="box box-github" />
                        <text x="230" y="795" class="label" style="text-anchor: start; font-size: 14px;">GitHub</text>
                        
                        <rect x="200" y="810" width="18" height="18" class="box box-data" />
                        <text x="230" y="825" class="label" style="text-anchor: start; font-size: 14px;">Data Files</text>
                        
                        <rect x="350" y="750" width="18" height="18" class="box box-scheduler" />
                        <text x="380" y="765" class="label" style="text-anchor: start; font-size: 14px;">Scheduler</text>
                        
                        <rect x="350" y="780" width="18" height="18" class="box box-excel" />
                        <text x="380" y="795" class="label" style="text-anchor: start; font-size: 14px;">Excel Export</text>
                        
                        <line x1="500" y1="760" x2="540" y2="760" class="arrow-read" />
                        <text x="550" y="765" class="label" style="text-anchor: start; font-size: 14px;">Read</text>
                        
                        <line x1="500" y1="790" x2="540" y2="790" class="arrow-write" />
                        <text x="550" y="795" class="label" style="text-anchor: start; font-size: 14px;">Write</text>
                        
                        <line x1="500" y1="820" x2="540" y2="820" class="arrow-bidi" />
                        <text x="550" y="825" class="label" style="text-anchor: start; font-size: 14px;">Bidirectional</text>
                    </svg>
                </div>
            </div>
        </div>
        
        <div class="grid">
            <div class="card full-width">
                <h2>Backend Logs (Last 500)</h2>
                <div class="log-container" id="backend-logs">Loading...</div>
            </div>
        </div>
    </div>
    
    <script>
        function formatTimestamp(isoString) {
            const date = new Date(isoString);
            return date.toLocaleString('en-CA', {
                timeZone: 'America/Toronto',
                hour12: false,
                hour: '2-digit',
                minute: '2-digit',
                second: '2-digit',
                year: 'numeric',
                month: '2-digit',
                day: '2-digit'
            }).replace(',', '');
        }
        
        function loadDashboard() {
            fetch('/api/health')
                .then(r => r.json())
                .then(data => {
                    document.getElementById('backend-start-time').textContent = formatTimestamp(data.backend_start_time);
                    document.getElementById('backend-errors').textContent = data.backend_has_errors ? 'Yes' : 'No';
                    
                    document.getElementById('github-api-status').textContent = data.github_health.api_status;
                    document.getElementById('github-branch').textContent = data.github_health.branch;
                    document.getElementById('github-rate-limit').textContent = data.github_health.rate_limit;
                    document.getElementById('github-last-commit').textContent = data.github_health.last_commit !== 'Unknown' && data.github_health.last_commit !== 'Error fetching' ? formatTimestamp(data.github_health.last_commit) : data.github_health.last_commit;
                    document.getElementById('file-employees').textContent = data.github_health.file_updates['employees.json'] !== 'N/A' && data.github_health.file_updates['employees.json'] !== 'No commits' && data.github_health.file_updates['employees.json'] !== 'Error' ? formatTimestamp(data.github_health.file_updates['employees.json']) : data.github_health.file_updates['employees.json'];
                    document.getElementById('file-schedules').textContent = data.github_health.file_updates['schedules.json'] !== 'N/A' && data.github_health.file_updates['schedules.json'] !== 'No commits' && data.github_health.file_updates['schedules.json'] !== 'Error' ? formatTimestamp(data.github_health.file_updates['schedules.json']) : data.github_health.file_updates['schedules.json'];
                    document.getElementById('file-requests').textContent = data.github_health.file_updates['availability_requests.json'] !== 'N/A' && data.github_health.file_updates['availability_requests.json'] !== 'No commits' && data.github_health.file_updates['availability_requests.json'] !== 'Error' ? formatTimestamp(data.github_health.file_updates['availability_requests.json']) : data.github_health.file_updates['availability_requests.json'];
                    document.getElementById('cache-hit-rate').textContent = data.metrics.cache_hit_rate;
                    document.getElementById('cache-hits').textContent = data.metrics.cache_hits;
                    document.getElementById('cache-misses').textContent = data.metrics.cache_misses;
                    document.getElementById('active-sessions').textContent = data.metrics.active_sessions;
                    document.getElementById('error-rate').textContent = data.metrics.error_rate;
                    
                    // Update per-card status indicators
                    const backendFrontendStatus = document.getElementById('backend-frontend-status');
                    if (data.backend_has_errors) {
                        backendFrontendStatus.className = 'status-indicator red';
                    } else {
                        backendFrontendStatus.className = 'status-indicator green';
                    }
                    
                    const githubStatus = document.getElementById('github-status');
                    if (data.github_health.api_status !== 'OK') {
                        githubStatus.className = 'status-indicator red';
                    } else {
                        githubStatus.className = 'status-indicator green';
                    }
                    
                    const systemStatus = document.getElementById('system-status');
                    if (data.metrics.error_rate !== "0%") {
                        systemStatus.className = 'status-indicator red';
                    } else {
                        systemStatus.className = 'status-indicator green';
                    }
                    
                    document.getElementById('last-updated').textContent = formatTimestamp(new Date().toISOString());
                });
            
            fetch('/api/logs/backend')
                .then(r => r.json())
                .then(data => {
                    const container = document.getElementById('backend-logs');
                    if (data.logs.length === 0) {
                        container.innerHTML = '<div class="empty-logs">No backend logs yet. Logs will appear as the server processes requests.</div>';
                    } else {
                        // Reverse to show newest logs first
                        const reversedLogs = [...data.logs].reverse();
                        container.innerHTML = reversedLogs.map(log => 
                            `<div class="log-entry ${log.level.toLowerCase()}">
                                <span class="timestamp">${formatTimestamp(log.timestamp)}</span>
                                <span class="level">[${log.level}]</span>
                                <span>${log.message}</span>
                            </div>`
                        ).join('');
                    }
                });
        }
        
        loadDashboard();
        // Auto-refresh removed to reduce GitHub API rate limit usage
        // Refresh manually by reloading the page
    </script>
</body>
</html>
    """
    return HTMLResponse(content=html_content)

@app.get("/api/health")
async def get_health_status():
    """Get overall system health status."""
    import cache_manager
    import auth
    
    logs = log_storage.get_log_storage()
    cache_stats = cache_manager.get_cache_stats()
    
    # GitHub health check
    github_health = {
        "api_status": "OK",
        "rate_limit": f"{cache_stats.get('rate_limit_remaining', 'N/A')}/{cache_stats.get('rate_limit_total', 'N/A')}",
        "last_commit": "Unknown",
        "branch": os.getenv("GITHUB_DATA_BRANCH", "data"),
        "file_updates": {}
    }
    
    try:
        import json_store
        import requests as req
        if json_store.GITHUB_AVAILABLE:
            github_health["api_status"] = "OK"
            # Get last commit time for the data branch
            try:
                commit_url = f"https://api.github.com/repos/{os.getenv('GITHUB_REPO', 'LeviSantosAraujo/IBU-Operations-Schedule')}/commits?sha={os.getenv('GITHUB_DATA_BRANCH', 'data')}&per_page=1"
                cr = req.get(commit_url, headers=json_store._headers(), timeout=5)
                if cr.status_code == 200:
                    commits = cr.json()
                    if commits:
                        github_health["last_commit"] = commits[0].get("commit", {}).get("committer", {}).get("date", "Unknown")
            except:
                github_health["last_commit"] = "Error fetching"
            # Get last update times for key files
            files = ["employees.json", "schedules.json", "availability_requests.json"]
            for filename in files:
                try:
                    url = f"https://api.github.com/repos/{os.getenv('GITHUB_REPO', 'LeviSantosAraujo/IBU-Operations-Schedule')}/commits?sha={os.getenv('GITHUB_DATA_BRANCH', 'data')}&path={filename}&per_page=1"
                    resp = req.get(url, headers=json_store._headers(), timeout=5)
                    if resp.status_code == 200:
                        commits = resp.json()
                        if commits:
                            github_health["file_updates"][filename] = commits[0].get("commit", {}).get("committer", {}).get("date", "Unknown")
                        else:
                            github_health["file_updates"][filename] = "No commits"
                except:
                    github_health["file_updates"][filename] = "Error"
        else:
            github_health["api_status"] = "Not Configured"
    except Exception as e:
        github_health["api_status"] = f"Error: {str(e)}"
    
    return {
        "backend_start_time": logs.get_server_start_time(),
        "backend_has_errors": logs.has_errors("backend"),
        "frontend_has_errors": False,  # Removed - frontend errors are client-side only
        "frontend_last_error": None,
        "frontend_start_time": None,  # Removed - not tracking frontend errors
        "github_health": github_health,
        "metrics": {
            "cache_hit_rate": f"{cache_stats.get('hit_rate_percent', 0):.1f}%",
            "cache_hits": cache_stats.get('total_hits', 0),
            "cache_misses": cache_stats.get('total_misses', 0),
            "rate_limit_remaining": cache_stats.get('rate_limit_remaining', 'N/A'),
            "rate_limit_total": cache_stats.get('rate_limit_total', 'N/A'),
            "active_sessions": auth.AuthManager.get_active_session_count(),
            "error_rate": ">0%" if logs.has_errors("backend") else "0%"
        }
    }

@app.get("/api/logs/backend")
async def get_backend_logs():
    """Get last 200 backend logs."""
    logs = log_storage.get_log_storage().get_backend_logs()
    return {"logs": logs}

@app.get("/api/logs/frontend")
async def get_frontend_logs():
    """Get last 200 frontend logs."""
    logs = log_storage.get_log_storage().get_frontend_logs()
    return {"logs": logs}

if __name__ == "__main__":
    import uvicorn
    # Initialize blob storage for cloud deployment
    if os.getenv("BLOB_READ_WRITE_TOKEN"):
        set_blob_key("ibu_schedule.xlsx")
    
    uvicorn.run(app, host="0.0.0.0", port=8000)
