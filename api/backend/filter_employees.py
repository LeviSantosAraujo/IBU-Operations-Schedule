import json
from openpyxl import load_workbook

# Load current employees
with open('data/employees.json', 'r') as f:
    employees = json.load(f)
current_names = {e['name'].lower() for e in employees}

# Load Excel to find all employee names
wb = load_workbook('/Users/levisantosaraujo/Downloads/Python/Schedule Sheet IBU/Last OPS Schedule June-Dec 2026.xlsx')
excel_names = set()

for sheet in wb.worksheets:
    for row in sheet.iter_rows(min_row=3, values_only=True):
        name = row[0]
        if name and str(name).strip() not in ['EVENTS', 'Total PT DAILY HOURS', 'IBU OPS']:
            excel_names.add(str(name).strip())

# Manual filter of actual employees
skip_keywords = ['training', 'break', 'intern', 'supervisor', 'manager', 'ops support', 
                'availabilities', 'blank', 'shifts', 'new intern', 'pending', 'last day',
                'fast track', 'wk', 'week', 'after ', 'before ', 'until ', 'anytime', 'all day',
                ' - ', ' 2', ' 3', ' 4', ' 5', ' 6', ' 7', ' c', ' b', ' tr',
                '12-3p', '1$', '2$', '3$', '4$', '5$', '6$', '7$']

actual_employees = []
for name in excel_names:
    name_lower = name.lower()
    if name_lower in current_names:
        continue
    
    skip = False
    for keyword in skip_keywords:
        if keyword in name_lower:
            skip = True
            break
    
    if not skip and len(name) > 2:
        actual_employees.append(name)

print(f'Actual employees to add: {len(actual_employees)}')
print()
for name in sorted(actual_employees):
    print(f'  - {name}')

# Add new employees to employees.json
with open('data/employees.json', 'r') as f:
    employees = json.load(f)

# Generate IDs for new employees
import uuid
for name in actual_employees:
    emp_id = f"emp_{uuid.uuid4().hex[:8]}"
    employees.append({
        "id": emp_id,
        "name": name,
        "employee_type": "employee",
        "email": "",
        "phone": "",
        "hire_date": None,
        "status": "active",
        "department": "Operations",
        "is_manager": False
    })
    print(f"Added: {name} -> {emp_id}")

with open('data/employees.json', 'w') as f:
    json.dump(employees, f, indent=2)

print(f"\nTotal employees in system: {len(employees)}")
