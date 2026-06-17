"""
Excel-based data storage system for IBU Schedule
- Uses Excel as the primary database
- Creates tabs: Config, PWDs, Employees, Availability, and weekly schedule tabs
- Real-time read/write to Excel
- Supports both local file system and Vercel Blob storage
"""

import os
import json
import re
from datetime import date, datetime, timedelta
from typing import List, Optional, Dict, Any
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from models import Employee, Availability, WeeklySchedule, Shift, EmployeeType, AvailabilityType, JobType, Floor, HourlyCoverageRequirement, Event
import io

# Import storage module
from storage import get_workbook, save_workbook, excel_file_exists as storage_file_exists

# Global path to the current Excel file (local storage)
EXCEL_FILE_PATH: Optional[str] = None
# Blob storage key (for cloud storage)
BLOB_KEY: Optional[str] = "ibu_schedule.xlsx"
# Workbook cache to reduce file loads
_workbook_cache: Optional[Workbook] = None

# Initialize with local Excel file if available
def _initialize_local_file():
    """Set Excel file path to local file if it exists"""
    global EXCEL_FILE_PATH
    try:
        local_path = Path(__file__).parent / "uploads" / "ibu_schedule.xlsx"
        if local_path.exists():
            EXCEL_FILE_PATH = str(local_path)
    except Exception as e:
        print(f"Failed to initialize local Excel file: {e}")

# NOTE: Module-level initialization removed to prevent loading Excel on import
# All data should come from data_store.py (JSON/blob storage)

def set_excel_file(path: str):
    """Set the active Excel file path (local storage)"""
    global EXCEL_FILE_PATH
    EXCEL_FILE_PATH = path
    # Ensure file exists with proper structure
    ensure_excel_structure(path)

def get_excel_file() -> Optional[str]:
    """Get the current Excel file path (local storage)"""
    return EXCEL_FILE_PATH

def set_blob_key(key: str):
    """Set the blob storage key (cloud storage)"""
    global BLOB_KEY
    BLOB_KEY = key

def get_blob_key() -> Optional[str]:
    """Get the current blob storage key"""
    return BLOB_KEY


def excel_file_exists() -> bool:
    """Check if Excel file exists in any storage"""
    # Check local file first
    if EXCEL_FILE_PATH and os.path.exists(EXCEL_FILE_PATH):
        return True
    
    # Check storage module (blob, local, memory)
    return storage_file_exists()

def _clear_workbook_cache():
    """Clear the workbook cache (call after upload/reset)"""
    global _workbook_cache
    _workbook_cache = None

def _get_workbook() -> Optional[Workbook]:
    """Get workbook from any available storage"""
    global _workbook_cache

    # Return cached workbook if available
    if _workbook_cache:
        return _workbook_cache

    # When GitHub storage is configured (production), it holds the freshest copy.
    # Prefer it over the bundled local file so live edits are not shadowed.
    try:
        from github_storage import GITHUB_AVAILABLE
    except Exception:
        GITHUB_AVAILABLE = False

    if GITHUB_AVAILABLE:
        wb = get_workbook()  # storage module reads GitHub first
        if wb:
            _workbook_cache = wb
            print("Loaded workbook from GitHub storage")
            return wb

    # Local file (used in local dev, or as a fallback seed in production)
    local_path = Path(__file__).parent / "uploads" / "ibu_schedule.xlsx"
    if local_path.exists():
        try:
            _workbook_cache = load_workbook(local_path)
            print(f"Loaded workbook from local file: {local_path}")
            return _workbook_cache
        except Exception as e:
            print(f"Error loading local file: {e}")

    # Try storage module (blob, local, memory)
    wb = get_workbook()
    if wb:
        _workbook_cache = wb
        print("Loaded workbook from storage module")
    return wb

def _save_workbook(wb: Workbook) -> bool:
    """Save workbook to storage"""
    global _workbook_cache
    print(f"[_SAVE_WORKBOOK] Starting save, calling storage.save_workbook")
    saved = save_workbook(wb)
    print(f"[_SAVE_WORKBOOK] storage.save_workbook returned: {saved}")
    try:
        local_path = Path(__file__).parent / "uploads" / "ibu_schedule.xlsx"
        local_path.parent.mkdir(exist_ok=True)
        wb.save(local_path)
        print(f"[_SAVE_WORKBOOK] Also saved to local: {local_path}")
    except OSError as e:
        print(f"[_SAVE_WORKBOOK] Local save failed (expected on Vercel): {e}")
    _workbook_cache = None
    return saved

def _invalidate_cache():
    """Clear the workbook cache"""
    global _workbook_cache
    _workbook_cache = None

def ensure_excel_structure(filepath: str):
    """Create Excel file with all required tabs if it doesn't exist"""
    if os.path.exists(filepath):
        # File exists, check if it has all required tabs
        wb = load_workbook(filepath)
        required_tabs = ['Config', 'PWDs', 'Employees', 'Availability']
        created_any = False
        for tab in required_tabs:
            if tab not in wb.sheetnames:
                wb.create_sheet(tab)
                created_any = True
                if tab == 'Config':
                    _init_config_sheet(wb[tab])
                elif tab == 'PWDs':
                    _init_pwds_sheet(wb[tab])
                elif tab == 'Employees':
                    _init_employees_sheet(wb[tab])
                elif tab == 'Availability':
                    _init_availability_sheet(wb[tab])
        if created_any:
            wb.save(filepath)
        wb.close()
        return
    
    # Create new workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Config tab
    config_sheet = wb.create_sheet('Config', 0)
    _init_config_sheet(config_sheet)
    
    # Create PWDs tab (passwords)
    pwds_sheet = wb.create_sheet('PWDs', 1)
    _init_pwds_sheet(pwds_sheet)
    
    # Create Employees tab
    emp_sheet = wb.create_sheet('Employees', 2)
    _init_employees_sheet(emp_sheet)
    
    # Create Availability tab
    avail_sheet = wb.create_sheet('Availability', 3)
    _init_availability_sheet(avail_sheet)
    
    wb.save(filepath)
    wb.close()

def _init_config_sheet(sheet):
    """Initialize Config sheet headers"""
    headers = ['Key', 'Value', 'Description']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # Default config values
    defaults = [
        ['organization', 'IBU', 'Organization name'],
        ['created_date', datetime.now().isoformat(), 'When this file was created'],
        ['last_modified', datetime.now().isoformat(), 'Last modification timestamp'],
        ['version', '1.0', 'System version'],
    ]
    for row_idx, row_data in enumerate(defaults, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 40

def _init_pwds_sheet(sheet):
    """Initialize PWDs sheet for storing manager passwords"""
    headers = ['Employee_ID', 'Employee_Name', 'Password_Hash', 'Role', 'Last_Login']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        sheet.column_dimensions[col_letter].width = 25

def _init_employees_sheet(sheet):
    """Initialize Employees sheet"""
    headers = ['ID', 'Name', 'Email', 'Type', 'Max_Hours', 'Preferences', 'Active', 'Created_At']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='9BBB59', end_color='9BBB59', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        sheet.column_dimensions[col_letter].width = 20

def _init_availability_sheet(sheet):
    """Initialize Availability sheet"""
    headers = ['ID', 'Employee_ID', 'Employee_Name', 'Week_Start', 'Monday', 'Tuesday', 
               'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday', 'Submitted_At', 'Notes',
               'Approved', 'Approved_By', 'Approved_At']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='8064A2', end_color='8064A2', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        sheet.column_dimensions[col_letter].width = 15

# ============ Password Management ============

def hash_password(password: str) -> str:
    """Simple hash for demo - use bcrypt in production"""
    import hashlib
    return hashlib.sha256(password.encode()).hexdigest()[:16]

def set_manager_password(employee_id: str, employee_name: str, password: str):
    """Set password for a manager in the PWDs tab"""
    wb = _get_workbook()
    if not wb:
        # Initialize a new workbook in memory if none exists
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        # Create PWDs sheet immediately so workbook has at least one sheet
        sheet = wb.create_sheet('PWDs')
        sheet.cell(row=1, column=1, value='Employee_ID')
        sheet.cell(row=1, column=2, value='Employee_Name')
        sheet.cell(row=1, column=3, value='Password_Hash')
        sheet.cell(row=1, column=4, value='Role')
        sheet.cell(row=1, column=5, value='Updated_At')
        _save_workbook(wb)
        wb = _get_workbook()

    if 'PWDs' not in wb.sheetnames:
        sheet = wb.create_sheet('PWDs')
        sheet.cell(row=1, column=1, value='Employee_ID')
        sheet.cell(row=1, column=2, value='Employee_Name')
        sheet.cell(row=1, column=3, value='Password_Hash')
        sheet.cell(row=1, column=4, value='Role')
        sheet.cell(row=1, column=5, value='Updated_At')
    else:
        sheet = wb['PWDs']
    
    # Check if employee already has password
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == employee_id:
            # Update existing
            sheet.cell(row=row, column=3, value=hash_password(password))
            sheet.cell(row=row, column=5, value=datetime.now().isoformat())
            _save_workbook(wb)
            wb.close()
            return
    
    # Add new entry
    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=1, value=employee_id)
    sheet.cell(row=new_row, column=2, value=employee_name)
    sheet.cell(row=new_row, column=3, value=hash_password(password))
    sheet.cell(row=new_row, column=4, value='manager')
    sheet.cell(row=new_row, column=5, value=datetime.now().isoformat())
    
    _save_workbook(wb)
    wb.close()

def verify_manager_password(employee_id: str, password: str) -> bool:
    """Verify manager password"""
    wb = _get_workbook()
    if not wb:
        return False
    
    try:
        sheet = wb['PWDs']
        
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == employee_id:
                stored_hash = sheet.cell(row=row, column=3).value
                wb.close()
                return stored_hash == hash_password(password)
        
        wb.close()
    except Exception:
        pass
    
    return False

def manager_has_password(employee_id: str) -> bool:
    """Check if manager has set a password"""
    wb = _get_workbook()
    if not wb:
        return False
    
    try:
        sheet = wb['PWDs']
        
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == employee_id:
                wb.close()
                return True
        
        wb.close()
    except Exception:
        pass
    
    return False

# ============ Employee Operations ============

def get_all_employees() -> List[Employee]:
    """Get all employees from JSON data store"""
    from data_store import get_all_employees as get_employees_from_json
    return get_employees_from_json()

def get_employee_by_id(employee_id: str) -> Optional[Employee]:
    """Get employee by ID"""
    employees = get_all_employees()
    return next((e for e in employees if e.id == employee_id), None)

def save_employee(employee: Employee) -> Employee:
    """Save or update employee in JSON data store"""
    from data_store import save_employee as save_employee_to_json
    return save_employee_to_json(employee)

def delete_employee(employee_id: str) -> bool:
    """Delete employee from JSON data store and Excel"""
    from data_store import delete_employee as delete_employee_from_json
    
    # Delete from JSON first
    json_result = delete_employee_from_json(employee_id)
    
    # Also delete from Excel Employees sheet
    try:
        wb = _get_workbook()
        if not wb:
            return json_result
        
        if 'Employees' not in wb.sheetnames:
            wb.close()
            return json_result
        
        sheet = wb['Employees']
        
        # Find and delete the employee row
        for row in range(2, sheet.max_row + 1):
            emp_id = sheet.cell(row=row, column=1).value
            if emp_id and str(emp_id) == employee_id:
                sheet.delete_rows(row)
                _save_workbook(wb)
                wb.close()
                # Clear workbook cache to force reload from GitHub
                _clear_workbook_cache()
                print(f"[EXCEL] Deleted employee {employee_id} from Excel Employees sheet and cleared cache")
                return True
        
        wb.close()
    except Exception as e:
        print(f"[EXCEL] Error deleting employee from Excel: {e}")
    
    return json_result

# ============ Hourly Coverage Requirements Operations ============

def get_coverage_requirements(week_start_date: date) -> List[HourlyCoverageRequirement]:
    """Get hourly coverage requirements for a week from Excel"""
    wb = _get_workbook()
    if not wb:
        return []
    
    try:
        if 'CoverageRequirements' not in wb.sheetnames:
            wb.close()
            return []
        
        sheet = wb['CoverageRequirements']
        requirements = []
        
        for row in range(2, sheet.max_row + 1):
            req_week = sheet.cell(row=row, column=1).value
            # Convert datetime to date for comparison
            if req_week and hasattr(req_week, 'date'):
                req_week = req_week.date()
            if req_week and isinstance(req_week, date) and req_week == week_start_date:
                req = HourlyCoverageRequirement(
                    id=str(sheet.cell(row=row, column=2).value or ''),
                    week_start_date=req_week,
                    day_of_week=str(sheet.cell(row=row, column=3).value or ''),
                    hour=int(sheet.cell(row=row, column=4).value or 0),
                    location=str(sheet.cell(row=row, column=5).value or ''),
                    required_employees=int(sheet.cell(row=row, column=6).value or 1),
                    is_call_center=sheet.cell(row=row, column=7).value == True or sheet.cell(row=row, column=7).value == 'True',
                    created_by=str(sheet.cell(row=row, column=8).value or ''),
                    created_at=datetime.fromisoformat(str(sheet.cell(row=row, column=9).value)) if sheet.cell(row=row, column=9).value else datetime.now(),
                    notes=sheet.cell(row=row, column=10).value
                )
                requirements.append(req)
        
        wb.close()
        return requirements
    except Exception as e:
        print(f"Error reading coverage requirements: {e}")
        return []

def save_coverage_requirement(requirement: HourlyCoverageRequirement) -> HourlyCoverageRequirement:
    """Save a coverage requirement to Excel"""
    wb = _get_workbook()
    if not wb:
        raise ValueError("Excel database not configured")
    
    # Create CoverageRequirements sheet if it doesn't exist
    if 'CoverageRequirements' not in wb.sheetnames:
        sheet = wb.create_sheet('CoverageRequirements')
        # Add headers
        sheet.cell(row=1, column=1, value='Week Start Date')
        sheet.cell(row=1, column=2, value='ID')
        sheet.cell(row=1, column=3, value='Day of Week')
        sheet.cell(row=1, column=4, value='Hour')
        sheet.cell(row=1, column=5, value='Location')
        sheet.cell(row=1, column=6, value='Required Employees')
        sheet.cell(row=1, column=7, value='Is Call Center')
        sheet.cell(row=1, column=8, value='Created By')
        sheet.cell(row=1, column=9, value='Created At')
        sheet.cell(row=1, column=10, value='Notes')
    else:
        sheet = wb['CoverageRequirements']
    
    # Check if requirement exists
    found_row = None
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=2).value) == requirement.id:
            found_row = row
            break
    
    if not found_row:
        found_row = sheet.max_row + 1
        if not requirement.id:
            requirement.id = f"req_{found_row}"
    
    # Write data
    sheet.cell(row=found_row, column=1, value=requirement.week_start_date)
    sheet.cell(row=found_row, column=2, value=requirement.id)
    sheet.cell(row=found_row, column=3, value=requirement.day_of_week)
    sheet.cell(row=found_row, column=4, value=requirement.hour)
    sheet.cell(row=found_row, column=5, value=requirement.location)
    sheet.cell(row=found_row, column=6, value=requirement.required_employees)
    sheet.cell(row=found_row, column=7, value=requirement.is_call_center)
    sheet.cell(row=found_row, column=8, value=requirement.created_by)
    sheet.cell(row=found_row, column=9, value=requirement.created_at.isoformat() if isinstance(requirement.created_at, datetime) else str(requirement.created_at))
    sheet.cell(row=found_row, column=10, value=requirement.notes)
    
    _save_workbook(wb)
    wb.close()
    return requirement

# ============ Availability Operations ============

def get_availabilities(week_start_date: Optional[date] = None, employee_id: Optional[str] = None) -> List[Availability]:
    """Get availabilities from Excel"""
    wb = _get_workbook()
    if not wb:
        return []
    
    try:
        if 'Availability' not in wb.sheetnames:
            wb.close()
            return []
        
        sheet = wb['Availability']
        availabilities = []
        
        for row in range(2, sheet.max_row + 1):
            avail_id = sheet.cell(row=row, column=1).value
            if not avail_id:
                continue
            
            emp_id = str(sheet.cell(row=row, column=2).value or '')
            week_str = str(sheet.cell(row=row, column=4).value or '')
            
            # Filter by employee if specified
            if employee_id and emp_id != employee_id:
                continue
            
            # Filter by week if specified
            if week_start_date:
                try:
                    row_date = datetime.strptime(week_str, '%Y-%m-%d').date()
                    if row_date != week_start_date:
                        continue
                except:
                    continue
            
            avail = Availability(
                id=str(avail_id),
                employee_id=emp_id,
                week_start_date=datetime.strptime(week_str, '%Y-%m-%d').date() if week_str else date.today(),
                monday=AvailabilityType(str(sheet.cell(row=row, column=5).value or 'blank')),
                tuesday=AvailabilityType(str(sheet.cell(row=row, column=6).value or 'blank')),
                wednesday=AvailabilityType(str(sheet.cell(row=row, column=7).value or 'blank')),
                thursday=AvailabilityType(str(sheet.cell(row=row, column=8).value or 'blank')),
                friday=AvailabilityType(str(sheet.cell(row=row, column=9).value or 'blank')),
                saturday=AvailabilityType(str(sheet.cell(row=row, column=10).value or 'blank')),
                sunday=AvailabilityType(str(sheet.cell(row=row, column=11).value or 'off')),
                submitted_at=datetime.fromisoformat(str(sheet.cell(row=row, column=12).value)) if sheet.cell(row=row, column=12).value else datetime.now(),
                notes=sheet.cell(row=row, column=13).value,
                approved=bool(sheet.cell(row=row, column=14).value),
                approved_by=sheet.cell(row=row, column=15).value,
                approved_at=datetime.fromisoformat(str(sheet.cell(row=row, column=16).value)) if sheet.cell(row=row, column=16).value else None
            )
            availabilities.append(avail)
        
        wb.close()
        return availabilities
    except Exception as e:
        print(f"Error reading availabilities: {e}")
        return []

def get_availability_for_week(employee_id: str, week_start_date: date) -> Optional[Availability]:
    """Get availability for specific employee and week"""
    availabilities = get_availabilities(week_start_date, employee_id)
    return availabilities[0] if availabilities else None

def save_availability(availability: Availability) -> Availability:
    """Save availability to Excel"""
    wb = _get_workbook()
    if not wb:
        raise ValueError("Excel database not configured. Please ask your manager to set up the Excel file.")
    
    # Create Availability sheet if it doesn't exist
    if 'Availability' not in wb.sheetnames:
        sheet = wb.create_sheet('Availability')
        # Add headers
        sheet.cell(row=1, column=1, value='ID')
        sheet.cell(row=1, column=2, value='Employee ID')
        sheet.cell(row=1, column=3, value='Employee Name')
        sheet.cell(row=1, column=4, value='Week Start Date')
        sheet.cell(row=1, column=5, value='Monday')
        sheet.cell(row=1, column=6, value='Tuesday')
        sheet.cell(row=1, column=7, value='Wednesday')
        sheet.cell(row=1, column=8, value='Thursday')
        sheet.cell(row=1, column=9, value='Friday')
        sheet.cell(row=1, column=10, value='Saturday')
        sheet.cell(row=1, column=11, value='Sunday')
        sheet.cell(row=1, column=12, value='Submitted At')
        sheet.cell(row=1, column=13, value='Notes')
        sheet.cell(row=1, column=14, value='Approved')
        sheet.cell(row=1, column=15, value='Approved By')
        sheet.cell(row=1, column=16, value='Approved At')
    else:
        sheet = wb['Availability']
    
    # Check if availability exists
    found_row = None
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value) == availability.id:
            found_row = row
            break
        # Also check by employee + week
        if (str(sheet.cell(row=row, column=2).value) == availability.employee_id and
            str(sheet.cell(row=row, column=4).value) == str(availability.week_start_date)):
            found_row = row
            break
    
    if not found_row:
        found_row = sheet.max_row + 1
    
    # Get employee name
    emp = get_employee_by_id(availability.employee_id)
    emp_name = emp.name if emp else ''
    
    # Write data
    sheet.cell(row=found_row, column=1, value=availability.id)
    sheet.cell(row=found_row, column=2, value=availability.employee_id)
    sheet.cell(row=found_row, column=3, value=emp_name)
    sheet.cell(row=found_row, column=4, value=str(availability.week_start_date))
    sheet.cell(row=found_row, column=5, value=availability.monday.value)
    sheet.cell(row=found_row, column=6, value=availability.tuesday.value)
    sheet.cell(row=found_row, column=7, value=availability.wednesday.value)
    sheet.cell(row=found_row, column=8, value=availability.thursday.value)
    sheet.cell(row=found_row, column=9, value=availability.friday.value)
    sheet.cell(row=found_row, column=10, value=availability.saturday.value)
    sheet.cell(row=found_row, column=11, value=availability.sunday.value)
    sheet.cell(row=found_row, column=12, value=availability.submitted_at.isoformat() if isinstance(availability.submitted_at, datetime) else str(availability.submitted_at))
    sheet.cell(row=found_row, column=13, value=availability.notes)
    sheet.cell(row=found_row, column=14, value=availability.approved)
    sheet.cell(row=found_row, column=15, value=availability.approved_by)
    sheet.cell(row=found_row, column=16, value=availability.approved_at.isoformat() if availability.approved_at else None)
    
    _save_workbook(wb)
    wb.close()
    return availability

# ============ Weekly Schedule Operations ============

def get_schedule_sheet_name(week_start_date: date) -> str:
    """Generate tab name for a week (legacy format)"""
    return f"Schedule_{week_start_date.strftime('%Y_%m_%d')}"

def get_schedule_sheet_name_native(week_start_date: date) -> str:
    """Generate tab name for a week in native format (e.g., 'June 1-7')"""
    import calendar
    week_end = week_start_date + timedelta(days=6)
    month_name = calendar.month_name[week_start_date.month]
    return f"{month_name} {week_start_date.day}-{week_end.day}"

def parse_schedule_sheet_dates(sheet_name: str, reference_year: int = None) -> Optional[tuple]:
    """Parse a schedule sheet name into start and end dates.
    Handles formats like:
      'June 1-7', 'May 25-31', 'April 27- May 3', '(TENTATIVE) June 8-14',
      'MARCH 16 - 22', 'FEB 23-March 1', 'Jan 26- feb 1', 'Sept 29 - OCT 5',
      'FEB 9 TO 15', 'Dec 29-jan 4'
    """
    import calendar
    import re
    
    if reference_year is None:
        reference_year = date.today().year
    
    month_names = {}
    for i, name in enumerate(calendar.month_name):
        if name:
            month_names[name.lower()] = i
            month_names[name[:3].lower()] = i
    # Add common abbreviations
    month_names['sept'] = 9
    
    # Remove prefixes like (TENTATIVE)
    clean_name = re.sub(r'\([^)]*\)\s*', '', sheet_name).strip()
    
    # Normalize separators: replace " TO " with "-"
    clean_name = re.sub(r'\s+TO\s+', '-', clean_name, flags=re.IGNORECASE)
    # Normalize spaces around dash
    clean_name = re.sub(r'\s*-\s*', '-', clean_name)
    
    # Try to parse: "Month day-day" or "Month day-Month day"
    # Split by dash
    parts = clean_name.split('-')
    if len(parts) != 2:
        return None
    
    left = parts[0].strip()
    right = parts[1].strip()
    
    # Parse left side (always has month + day)
    left_tokens = left.split()
    if len(left_tokens) < 2:
        return None
    
    start_month_str = left_tokens[0].lower()
    if start_month_str not in month_names:
        return None
    start_month = month_names[start_month_str]
    
    try:
        start_day = int(left_tokens[1])
    except (ValueError, IndexError):
        return None
    
    # Parse right side (could be just a day number, or "Month day")
    right_tokens = right.split()
    if len(right_tokens) == 1:
        # Just a number (same month) or just a month name followed by nothing
        try:
            end_day = int(right_tokens[0])
            end_month = start_month
        except ValueError:
            # Maybe it's a month name without a day? unlikely
            return None
    elif len(right_tokens) >= 2:
        # "Month day"
        end_month_str = right_tokens[0].lower()
        if end_month_str in month_names:
            end_month = month_names[end_month_str]
            try:
                end_day = int(right_tokens[1])
            except ValueError:
                return None
        else:
            # Try as just a number
            try:
                end_day = int(right_tokens[0])
                end_month = start_month
            except ValueError:
                return None
    else:
        return None
    
    try:
        start_date = date(reference_year, start_month, start_day)
        end_date = date(reference_year, end_month, end_day)
        
        # Handle year wraparound (e.g., Dec 29-Jan 4)
        if end_date < start_date:
            # End is in the next year or start is in the previous year
            if start_month > end_month:
                end_date = date(reference_year + 1, end_month, end_day)
        
        return (start_date, end_date)
    except ValueError:
        return None


def find_schedule_sheet_by_date(wb: Workbook, target_date: date) -> Optional[str]:
    """Find schedule sheet that contains the target date based on sheet names"""
    reference_year = target_date.year
    
    for sheet_name in wb.sheetnames:
        result = parse_schedule_sheet_dates(sheet_name, reference_year)
        if result:
            start_date, end_date = result
            if start_date <= target_date <= end_date:
                return sheet_name
    
    return None

def ensure_schedule_sheet(wb: Workbook, week_start_date: date):
    """Create schedule sheet if it doesn't exist"""
    # Use legacy format to avoid conflicts with native Excel format
    sheet_name = get_schedule_sheet_name(week_start_date)
    
    if sheet_name not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_name)
        headers = ['Shift_ID', 'Employee_ID', 'Employee_Name', 'Day', 'Start_Time', 
                   'End_Time', 'Job_Type', 'Floor', 'Hours', 'Is_Event', 'Event_Name',
                   'Locked', 'Locked_Avail_Type', 'Location']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            sheet.column_dimensions[col_letter].width = 15
    
    return wb[sheet_name]

def parse_shift_time_and_location(shift_text: str) -> tuple:
    """Parse shift text like '8a-3p GF' or '8a-4p f6 CC B@1' into start_time, end_time, and location display
    Also handles multiple timeframes in one cell: '8a-10a 2f 1p-3p CC' or '8a-10a 2f / 1p-3p CC'
    Handles 3+ timeframes: '9a-12p w fran (12-12:30 GR)/ B@12:30/ 1-3 w fran'
    Returns list of (start_time, end_time, location_display, comment) tuples
    """
    import re
    
    shift_text = shift_text.strip()
    
    # Split by '/' to get multiple shift segments
    # Handle patterns like: "9a-12p w fran (12-12:30 GR)/ B@12:30/ 1-3 w fran"
    segments = re.split(r'\s*/\s*', shift_text)
    
    if len(segments) > 1:
        # Multiple shifts
        shifts = []
        for segment in segments:
            segment = segment.strip()
            if segment:
                shifts.append(parse_single_shift(segment))
        return shifts
    else:
        # Check for multiple timeframes separated by space (without '/')
        # Pattern: "8a-10a 2f 1p-3p CC"
        multi_pattern = r'(\d{1,2}(?::\d{2})?[ap]?\s*-\s*\d{1,2}(?::\d{2})?[ap]?\s*[^\s/]*)\s+(?=\d{1,2}(?::\d{2})?[ap]?\s*-\s*\d{1,2}(?::\d{2})?[ap]?)'
        matches = re.finditer(multi_pattern, shift_text, re.IGNORECASE)
        
        shift_texts = [m.group(1).strip() for m in matches]
        
        if len(shift_texts) > 1:
            shifts = []
            for shift_text in shift_texts:
                shifts.append(parse_single_shift(shift_text))
            return shifts
        else:
            # Single shift
            return [parse_single_shift(shift_text)]

def parse_single_shift(shift_text: str) -> tuple:
    """Parse a single shift text into start_time, end_time, location display, and comment"""
    import re
    
    shift_text = shift_text.strip()
    
    # Parse time pattern: e.g., "8a-3p", "9a-4p", "12:30p-3:30p"
    time_pattern = r'(\d{1,2}(?::\d{2})?[ap]?)\s*-\s*(\d{1,2}(?::\d{2})?[ap]?)'
    time_match = re.search(time_pattern, shift_text, re.IGNORECASE)
    
    start_time = ''
    end_time = ''
    
    if time_match:
        start_time = time_match.group(1)
        end_time = time_match.group(2)
        
        # Normalize time format (e.g., "8a" -> "8:00 AM", "9a-4p" -> "9:00 AM-4:00 PM")
        start_time = normalize_time(start_time)
        end_time = normalize_time(end_time)
    
    # Parse and translate location codes
    location_display = translate_location_code(shift_text)
    
    # Extract any unmapped comment (text that's not time or location)
    comment = extract_comment(shift_text, start_time, end_time, location_display)
    
    return start_time, end_time, location_display, comment

def extract_comment(shift_text: str, start_time: str, end_time: str, location: Optional[str]) -> Optional[str]:
    """Extract unmapped comment text from shift cell
    Removes time patterns, location codes, and other translated information
    Only returns text that wasn't translated into location, floor, etc.
    """
    import re
    
    if not shift_text:
        return None
    
    # Remove time patterns
    comment = re.sub(r'\d{1,2}(?::\d{2})?[ap]?\s*-\s*\d{1,2}(?::\d{2})?[ap]?', '', shift_text, flags=re.IGNORECASE)
    
    # Remove location codes (both original and translated forms)
    # Original codes: 2f, f2, 6f, f6, cc, gf, gr, wfh
    # Translated forms: 2nd floor, 6th floor, call center, ground floor, working from home, 80 bloor
    comment = re.sub(r'\b(2f|f2|6f|f6|cc|gf|gr|wfh)\b', '', comment, flags=re.IGNORECASE)
    comment = re.sub(r'\b(2nd floor|6th floor|call center|ground floor|working from home|80 bloor|bloor)\b', '', comment, flags=re.IGNORECASE)
    
    # Remove common markers and separators
    comment = re.sub(r'[@#/]', '', comment)
    
    # Remove words in parentheses that might be location codes (e.g., "(12-12:30 GR)")
    comment = re.sub(r'\([^)]*\)', '', comment)
    
    # Clean up whitespace
    comment = re.sub(r'\s+', ' ', comment).strip()
    
    return comment if comment else None

def translate_location_code(text: str) -> str:
    """Extract and translate location codes to human-readable format
    2F/F2 → 2nd Floor
    6F/F6 → 6th Floor
    CC → Call Center
    Bloor/bloor → 80 Bloor
    GF → Ground Floor
    GR → Ground Floor
    WFH → Working from Home
    Other codes → returned as-is
    """
    text_lower = text.lower()
    
    # Check for multiple locations and return them all
    locations = []
    
    if '2f' in text_lower or 'f2' in text_lower:
        locations.append('2nd Floor')
    if '6f' in text_lower or 'f6' in text_lower:
        locations.append('6th Floor')
    if 'cc' in text_lower:
        locations.append('Call Center')
    if 'bloor' in text_lower:
        locations.append('80 Bloor')
    if 'gf' in text_lower or 'gr' in text_lower or 'ground' in text_lower:
        locations.append('Ground Floor')
    if 'wfh' in text_lower:
        locations.append('Working from Home')
    
    if locations:
        return ', '.join(locations)
    
    # Try to extract any location-like text (words at end)
    import re
    location_match = re.search(r'\b([A-Z][A-Za-z\s]+)$', text)
    if location_match:
        return location_match.group(1).strip()
    return None

def get_location_color(location: Optional[str]) -> Optional[str]:
    """Get color code for a location
    2nd Floor → green (#90EE90)
    Ground Floor → grey (#D3D3D3)
    6th Floor → blue (#87CEEB)
    Working from Home → red (#FF6B6B)
    80 Bloor → purple (#DDA0DD)
    Call Center → orange (#FFB347)
    Other → None (use Excel cell color)
    """
    if not location:
        return None
    
    location_lower = location.lower()
    
    if '2nd floor' in location_lower or 'f2' in location_lower:
        return '#90EE90'  # Light green
    elif 'ground floor' in location_lower or 'ground' in location_lower or 'gf' in location_lower or 'gr' in location_lower:
        return '#D3D3D3'  # Light grey
    elif '6th floor' in location_lower or 'f6' in location_lower:
        return '#87CEEB'  # Sky blue
    elif 'working from home' in location_lower or 'wfh' in location_lower:
        return '#FF6B6B'  # Light red
    elif '80 bloor' in location_lower or 'bloor' in location_lower:
        return '#DDA0DD'  # Plum purple
    elif 'call center' in location_lower or 'cc' in location_lower:
        return '#FFB347'  # Light orange
    
    return None

def normalize_time(time_str: str) -> str:
    """Normalize time like '8a' to '8:00 AM', '9:30p' to '9:30 PM', '4' to '4:00 PM' (assumes PM for single digits)"""
    import re
    time_str = time_str.strip().lower()
    
    # Pattern: 8a, 9p, 12:30p, 12p, 3:30p, etc.
    match = re.match(r'(\d{1,2})(?::(\d{2}))?([ap])', time_str)
    if match:
        hour = match.group(1)
        minute = match.group(2) if match.group(2) else '00'
        ampm = match.group(3).upper() + 'M'
        return f"{hour}:{minute} {ampm}"
    
    # Pattern: single hour like '4', '3' - assume PM for these
    match = re.match(r'^(\d{1,2})$', time_str)
    if match:
        hour = match.group(1)
        return f"{hour}:00 PM"
    
    return time_str

def calculate_hours_from_time_range(start_time: str, end_time: str) -> float:
    """Calculate hours from time range strings like '8:00 AM' to '3:00 PM'"""
    try:
        import re
        # Parse start time
        start_match = re.match(r'(\d{1,2}):(\d{2})\s*(AM|PM)', start_time.upper())
        end_match = re.match(r'(\d{1,2}):(\d{2})\s*(AM|PM)', end_time.upper())
        
        if not start_match or not end_match:
            return 0.0
        
        start_hour = int(start_match.group(1))
        start_min = int(start_match.group(2))
        start_ampm = start_match.group(3)
        
        end_hour = int(end_match.group(1))
        end_min = int(end_match.group(2))
        end_ampm = end_match.group(3)
        
        # Convert to 24-hour format
        if start_ampm == 'PM' and start_hour != 12:
            start_hour += 12
        if start_ampm == 'AM' and start_hour == 12:
            start_hour = 0
        
        if end_ampm == 'PM' and end_hour != 12:
            end_hour += 12
        if end_ampm == 'AM' and end_hour == 12:
            end_hour = 0
        
        # Calculate difference in minutes
        start_minutes = start_hour * 60 + start_min
        end_minutes = end_hour * 60 + end_min
        
        # Handle overnight shifts (end < start)
        if end_minutes < start_minutes:
            end_minutes += 24 * 60
        
        diff_minutes = end_minutes - start_minutes
        return diff_minutes / 60.0
    except Exception:
        return 0.0

def get_cell_color(sheet, row: int, col: int) -> Optional[str]:
    """Get the background color of a cell as hex string"""
    from openpyxl.styles import Color
    
    try:
        cell = sheet.cell(row=row, column=col)
        if cell.fill and cell.fill.fgColor:
            color = cell.fill.fgColor
            if color.type == 'rgb' and color.rgb:
                # RGB format: AARRGGBB or RRGGBB
                rgb = color.rgb
                if len(rgb) == 8:
                    return f"#{rgb[2:]}"  # Skip alpha, add #
                elif len(rgb) == 6:
                    return f"#{rgb}"
            elif color.type == 'theme':
                # Theme colors - return a default based on theme
                # This is a simplification; theme colors would need theme lookup
                theme_colors = {
                    0: '#FFFFFF',  # Light 1
                    1: '#000000',  # Dark 1
                    2: '#EEECE1',  # Light 2
                    3: '#1F497D',  # Dark 2
                    4: '#4F81BD',  # Accent 1
                    5: '#C0504D',  # Accent 2
                    6: '#9BBB59',  # Accent 3
                    7: '#8064A2',  # Accent 4
                    8: '#4BACC6',  # Accent 5
                    9: '#F79646',  # Accent 6
                }
                return theme_colors.get(color.theme, None)
            elif color.type == 'indexed':
                # Indexed colors from palette
                indexed_colors = [
                    '#000000', '#FFFFFF', '#FF0000', '#00FF00', '#0000FF',
                    '#FFFF00', '#FF00FF', '#00FFFF', '#800000', '#008000',
                    '#000080', '#808000', '#800080', '#008080', '#C0C0C0',
                    '#808080', '#9999FF', '#993366', '#FFFFCC', '#CCFFFF',
                    '#660066', '#FF8080', '#0066CC', '#CCCCFF', '#000080',
                    '#FF00FF', '#FFFF00', '#00FFFF', '#800080', '#800000',
                    '#008080', '#0000FF', '#00CCFF', '#CCFFFF', '#CCFFCC',
                    '#FFFF99', '#99CCFF', '#FF99CC', '#CC99FF', '#FFCC99',
                    '#3366FF', '#33CCCC', '#99CC00', '#FFCC00', '#FF9900',
                    '#FF6600', '#CCCC99', '#996699', '#669999', '#006699',
                ]
                idx = color.indexed
                if idx is not None and 0 <= idx < len(indexed_colors):
                    return indexed_colors[idx]
    except Exception:
        pass
    
    return None

def get_schedule_by_week(week_start_date: date) -> Optional[WeeklySchedule]:
    """Get schedule for a specific week from its tab"""
    wb = None
    
    # Try local file first - load without data_only to preserve colors
    if EXCEL_FILE_PATH and os.path.exists(EXCEL_FILE_PATH):
        wb = load_workbook(EXCEL_FILE_PATH, data_only=False)
    else:
        # Try storage module
        wb = _get_workbook()
    
    if not wb:
        return None
    
    try:
        
        # First try the new date-based lookup for sheets like "June 1-7"
        sheet_name = find_schedule_sheet_by_date(wb, week_start_date)
        
        # Fall back to legacy format if not found
        if not sheet_name:
            sheet_name = get_schedule_sheet_name(week_start_date)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        
        sheet = wb[sheet_name]
        shifts = []
        total_hours: Dict[str, float] = {}
        
        # Detect sheet format by checking row 1
        row1_col1 = sheet.cell(row=1, column=1).value
        row1_col2 = str(sheet.cell(row=1, column=2).value or '').lower()
        
        # Check if this is the native format (column layout: Name, Mon, HRS, Tue, HRS, ...)
        is_native_format = row1_col2 in ['monday', 'mon', 'mon '] or 'monday' in row1_col2 or 'mon' in row1_col2
        
        if is_native_format:
            # Native Excel schedule format
            # Row 1: headers (blank, Monday, HRS, Tuesday, HRS, ...)
            # Row 2: dates
            # Row 3: EVENTS
            # Row 4+: Employee name, shift_text, hours, shift_text, hours, ...
            # Columns: A=Name, B=Mon shift, C=Mon hrs, D=Tue shift, E=Tue hrs, ...
            #          F=Wed shift, G=Wed hrs, H=Thu shift, I=Thu hrs, J=Fri shift, K=Fri hrs
            #          L=Sat shift, M=Sat hrs, N=Sun shift, O=Sun hrs, P=Total
            
            day_columns = {
                'monday': (2, 3),     # shift col, hours col
                'tuesday': (4, 5),
                'wednesday': (6, 7),
                'thursday': (8, 9),
                'friday': (10, 11),
                'saturday': (12, 13),
                'sunday': (14, 15),
            }
            
            # Get events from row 3
            events_row = None
            for r in range(1, min(5, sheet.max_row + 1)):
                cell_val = str(sheet.cell(row=r, column=1).value or '').strip().upper()
                if 'EVENT' in cell_val:
                    events_row = r
                    break
            
            # Find where employee data starts (after EVENTS row or row 4)
            start_row = (events_row + 1) if events_row else 4
            
            # Stop phrases
            stop_phrases = ['total pt daily hours', 'total', 'shifts', 'availabilities', 'blank', 'availability']
            
            employees_from_excel = get_all_employees()
            emp_name_to_id = {}
            for emp in employees_from_excel:
                # Skip Availabilities entry
                if emp.name.lower().strip() in ['availabilities', 'availability']:
                    continue
                
                name_lower = emp.name.lower().strip()
                emp_name_to_id[name_lower] = emp.id
                # Also add variations with spaces removed
                emp_name_to_id[name_lower.replace(' ', '')] = emp.id
                # Add variations with different spacing
                emp_name_to_id[name_lower.replace('-', ' ')] = emp.id
                emp_name_to_id[name_lower.replace('-', '')] = emp.id
                # Add version with trailing number removed (e.g., "intern- levi 6" -> "intern- levi")
                name_no_number = re.sub(r'\s+\d+$', '', name_lower).strip()
                if name_no_number != name_lower:
                    emp_name_to_id[name_no_number] = emp.id
                    emp_name_to_id[name_no_number.replace(' ', '')] = emp.id
                    emp_name_to_id[name_no_number.replace('-', ' ')] = emp.id
                    emp_name_to_id[name_no_number.replace('-', '')] = emp.id
            
            for row in range(start_row, sheet.max_row + 1):
                emp_name = sheet.cell(row=row, column=1).value
                if not emp_name or not isinstance(emp_name, str):
                    continue
                
                emp_name = emp_name.strip()
                if not emp_name or emp_name.lower() in stop_phrases:
                    continue
                if any(phrase in emp_name.lower() for phrase in stop_phrases):
                    break
                
                # Find employee ID by name (normalize to handle trailing numbers like "Intern- Levi 6")
                sheet_name_normalized = emp_name.lower().strip()
                sheet_name_no_number = re.sub(r'\s+\d+$', '', sheet_name_normalized)
                
                emp_id = emp_name_to_id.get(sheet_name_normalized)
                if not emp_id:
                    emp_id = emp_name_to_id.get(sheet_name_no_number)
                if not emp_id:
                    emp_id = emp_name_to_id.get(sheet_name_normalized.replace(' ', ''))
                if not emp_id:
                    emp_id = emp_name_to_id.get(sheet_name_no_number.replace(' ', ''))
                if not emp_id:
                    emp_id = emp_name_to_id.get(sheet_name_normalized.replace('-', ''))
                if not emp_id:
                    emp_id = emp_name_to_id.get(sheet_name_no_number.replace('-', ''))
                if not emp_id:
                    emp_id = sheet_name_no_number.replace(' ', '_')
                
                emp_total_hours = 0.0
                
                for day_name, (shift_col, hours_col) in day_columns.items():
                    shift_text = sheet.cell(row=row, column=shift_col).value
                    hours_val = sheet.cell(row=row, column=hours_col).value
                    
                    if not shift_text or str(shift_text).strip() == '':
                        continue
                    
                    shift_str = str(shift_text).strip()
                    
                    # Skip OFF/RO entries
                    if shift_str.upper() in ['OFF', 'RO', '']:
                        continue
                    
                    # Parse hours
                    try:
                        hrs = float(hours_val) if hours_val else 0.0
                    except (ValueError, TypeError):
                        hrs = 0.0
                    
                    emp_total_hours += hrs
                    
                    # Parse time and location from shift text (may return multiple shifts)
                    parsed_shifts = parse_shift_time_and_location(shift_str)
                    
                    # Get cell color from Excel
                    cell_color = get_cell_color(sheet, row, shift_col)
                    
                    # Create shift(s) from parsed data
                    for idx, (start_time, end_time, location_display, comment) in enumerate(parsed_shifts):
                        # Prioritize Excel cell color over location-based color
                        final_color = cell_color if cell_color else get_location_color(location_display)
                        
                        # Calculate hours for this specific shift if multiple
                        if len(parsed_shifts) > 1:
                            # Approximate hours from time range
                            shift_hrs = calculate_hours_from_time_range(start_time, end_time)
                        else:
                            shift_hrs = hrs
                        
                        # Check if break is required (>5 hours)
                        requires_break = shift_hrs > 5.0
                        
                        shift = Shift(
                            id=f"{emp_id}_{day_name}_{row}_{idx}",
                            employee_id=emp_id,
                            day_of_week=day_name,
                            start_time=start_time,
                            end_time=end_time,
                            job_type=JobType.DESK,
                            floor=None,  # Using location field instead
                            location=location_display,
                            hours=shift_hrs,
                            is_event=False,
                            event_name=None,
                            color=final_color,
                            comment=comment,
                            requires_break=requires_break,
                            break_provided=False  # Will be set by manager
                        )
                        shifts.append(shift)
                
                if emp_total_hours > 0:
                    total_hours[emp_id] = emp_total_hours
        else:
            # Legacy system-generated format (Shift_ID, Employee_ID, ...)
            for row in range(2, sheet.max_row + 1):
                shift_id = sheet.cell(row=row, column=1).value
                if not shift_id:
                    continue
                
                emp_id = str(sheet.cell(row=row, column=2).value or '')
                hours = float(sheet.cell(row=row, column=9).value or 0)
                
                shift = Shift(
                    id=str(shift_id),
                    employee_id=emp_id,
                    day_of_week=str(sheet.cell(row=row, column=4).value or 'monday'),
                    start_time=str(sheet.cell(row=row, column=5).value or '09:00'),
                    end_time=str(sheet.cell(row=row, column=6).value or '17:00'),
                    job_type=JobType(str(sheet.cell(row=row, column=7).value or 'desk')),
                    floor=Floor(str(sheet.cell(row=row, column=8).value)) if sheet.cell(row=row, column=8).value else None,
                    hours=hours,
                    is_event=sheet.cell(row=row, column=10).value == True or sheet.cell(row=row, column=10).value == 'True',
                    event_name=sheet.cell(row=row, column=11).value,
                    locked=sheet.cell(row=row, column=12).value == True or sheet.cell(row=row, column=12).value == 'True',
                    locked_availability_type=str(sheet.cell(row=row, column=13).value) if sheet.cell(row=row, column=13).value else None,
                    location=str(sheet.cell(row=row, column=14).value) if sheet.cell(row=row, column=14).value else None
                )
                shifts.append(shift)
                
                # Skip day off shifts when calculating total hours
                if shift.location == 'day off' or shift.locked_availability_type == 'Day Off':
                    continue
                
                if emp_id in total_hours:
                    total_hours[emp_id] += hours
                else:
                    total_hours[emp_id] = hours
        
        wb.close()
        
        return WeeklySchedule(
            id=f"sched_{week_start_date.isoformat()}",
            week_start_date=week_start_date,
            shifts=shifts,
            total_hours=total_hours,
            created_at=datetime.now(),
            updated_at=datetime.now(),
            status='published' if shifts else 'draft'
        )
    except Exception as e:
        print(f"Error reading schedule: {e}")
        import traceback
        traceback.print_exc()
        return None

def save_schedule(schedule: WeeklySchedule) -> WeeklySchedule:
    """Save schedule to Excel in its own tab"""
    wb = _get_workbook()
    if not wb:
        raise ValueError("No Excel file available")
    
    sheet = ensure_schedule_sheet(wb, schedule.week_start_date)
    
    # Clear existing data (keep header)
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    
    # Write all shifts
    for shift in schedule.shifts:
        emp = get_employee_by_id(shift.employee_id)
        emp_name = emp.name if emp else ''
        
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=shift.id)
        sheet.cell(row=row, column=2, value=shift.employee_id)
        sheet.cell(row=row, column=3, value=emp_name)
        sheet.cell(row=row, column=4, value=shift.day_of_week)
        sheet.cell(row=row, column=5, value=shift.start_time)
        sheet.cell(row=row, column=6, value=shift.end_time)
        sheet.cell(row=row, column=7, value=shift.job_type.value)
        sheet.cell(row=row, column=8, value=shift.floor.value if shift.floor else None)
        sheet.cell(row=row, column=9, value=shift.hours)
        sheet.cell(row=row, column=10, value=shift.is_event)
        sheet.cell(row=row, column=11, value=shift.event_name)
        sheet.cell(row=row, column=12, value=getattr(shift, 'locked', False))
        sheet.cell(row=row, column=13, value=getattr(shift, 'locked_availability_type', None))
        sheet.cell(row=row, column=14, value=shift.location)
    
    _save_workbook(wb)
    wb.close()
    return schedule

def get_all_schedules() -> List[WeeklySchedule]:
    """Get all schedules from week tabs"""
    wb = _get_workbook()
    if not wb:
        return []
    
    try:
        schedules = []
        
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('Schedule_'):
                try:
                    # Parse date from sheet name
                    date_str = sheet_name.replace('Schedule_', '').replace('_', '-')
                    week_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    
                    # Read the schedule
                    sheet = wb[sheet_name]
                    shifts = []
                    total_hours: Dict[str, float] = {}
                    
                    for row in range(2, sheet.max_row + 1):
                        shift_id = sheet.cell(row=row, column=1).value
                        if not shift_id:
                            continue
                        
                        emp_id = str(sheet.cell(row=row, column=2).value or '')
                        hours = float(sheet.cell(row=row, column=9).value or 0)
                        
                        shift = Shift(
                            id=str(shift_id),
                            employee_id=emp_id,
                            day_of_week=str(sheet.cell(row=row, column=4).value or 'monday'),
                            start_time=str(sheet.cell(row=row, column=5).value or '09:00'),
                            end_time=str(sheet.cell(row=row, column=6).value or '17:00'),
                            job_type=JobType(str(sheet.cell(row=row, column=7).value or 'desk')),
                            floor=Floor(str(sheet.cell(row=row, column=8).value)) if sheet.cell(row=row, column=8).value else None,
                            hours=hours,
                            is_event=sheet.cell(row=row, column=10).value == True,
                            event_name=sheet.cell(row=row, column=11).value,
                            locked=sheet.cell(row=row, column=12).value == True or sheet.cell(row=row, column=12).value == 'True',
                            locked_availability_type=str(sheet.cell(row=row, column=13).value) if sheet.cell(row=row, column=13).value else None,
                            location=str(sheet.cell(row=row, column=14).value) if sheet.cell(row=row, column=14).value else None
                        )
                        shifts.append(shift)
                        
                        # Skip day off shifts when calculating total hours
                        if shift.location == 'day off' or shift.locked_availability_type == 'Day Off':
                            continue
                        
                        if emp_id in total_hours:
                            total_hours[emp_id] += hours
                        else:
                            total_hours[emp_id] = hours
                    
                    schedule = WeeklySchedule(
                        id=f"sched_{week_date.isoformat()}",
                        week_start_date=week_date,
                        shifts=shifts,
                        total_hours=total_hours,
                        created_at=datetime.now(),
                        updated_at=datetime.now(),
                        status='published' if shifts else 'draft'
                    )
                    schedules.append(schedule)
                except Exception as e:
                    print(f"Error reading schedule tab {sheet_name}: {e}")
        
        wb.close()
        return sorted(schedules, key=lambda x: x.week_start_date, reverse=True)
    except Exception as e:
        print(f"Error reading all schedules: {e}")
        return []

def delete_schedule(week_start_date: date) -> bool:
    """Delete a schedule by removing its tab"""
    wb = _get_workbook()
    if not wb:
        return False
    
    # Use legacy format to match the saver
    sheet_name = get_schedule_sheet_name(week_start_date)
    
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        _save_workbook(wb)
        wb.close()
        return True
    
    wb.close()
    return False

def migrate_native_to_system_format(week_start_date: date) -> bool:
    """Migrate a native Excel format schedule to the system format"""
    wb = _get_workbook()
    if not wb:
        return False
    
    try:
        # Find the native format tab
        native_sheet_name = find_schedule_sheet_by_date(wb, week_start_date)
        
        if not native_sheet_name:
            wb.close()
            return False
        
        # Read the schedule using the existing get_schedule_by_week logic
        # This will parse the native format and return a WeeklySchedule object
        schedule = get_schedule_by_week(week_start_date)
        
        if not schedule:
            wb.close()
            return False
        wb.close()
        return True
        
    except Exception as e:
        print(f"Error migrating schedule: {e}")
        wb.close()
        return False

# ============ Floor Coverage ============

def get_floor_coverage(floor: str, day_of_week: str, time_slot: str, week_start_date: date) -> Dict:
    """Get employees working on a floor at a specific time"""
    schedule = get_schedule_by_week(week_start_date)
    if not schedule:
        return {"floor": floor, "day_of_week": day_of_week, "time_slot": time_slot, "employee_count": 0, "employees": []}
    
    # Filter shifts by floor and day
    matching_shifts = [
        s for s in schedule.shifts 
        if s.day_of_week.lower() == day_of_week.lower() 
        and s.floor 
        and s.floor.value.lower() == floor.lower()
    ]
    
    employees = []
    for shift in matching_shifts:
        emp = get_employee_by_id(shift.employee_id)
        if emp:
            employees.append({
                "id": emp.id,
                "name": emp.name,
                "job_type": shift.job_type.value,
                "hours": shift.hours,
                "start_time": shift.start_time,
                "end_time": shift.end_time
            })
    
    return {
        "floor": floor,
        "day_of_week": day_of_week,
        "time_slot": time_slot,
        "employee_count": len(employees),
        "employees": employees
    }

# ============ System Config ============

def get_system_config() -> Dict[str, Any]:
    """Get system config from Excel"""
    import json
    import ast
    wb = _get_workbook()
    if not wb:
        return {}
    
    try:
        if 'Config' not in wb.sheetnames:
            wb.close()
            return {}
        
        sheet = wb['Config']
        config = {}
        
        for row in range(2, sheet.max_row + 1):
            key = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=2).value
            if key:
                # Try to parse JSON for complex values (like staffing_targets dict)
                if isinstance(value, str):
                    # Try JSON first (new format)
                    try:
                        config[key] = json.loads(value)
                    except (json.JSONDecodeError, ValueError):
                        # Fall back to Python literal eval (for old format with single quotes)
                        try:
                            config[key] = ast.literal_eval(value)
                        except (ValueError, SyntaxError):
                            config[key] = value
                else:
                    config[key] = value
        
        wb.close()
        return config
    except Exception as e:
        print(f"Error reading config: {e}")
        return {}

def save_system_config(config: Dict[str, Any]):
    """Save system config to Excel"""
    import json
    wb = _get_workbook()
    if not wb:
        raise ValueError("No Excel file available")
    
    if 'Config' not in wb.sheetnames:
        wb.create_sheet('Config', 0)
        _init_config_sheet(wb['Config'])
    
    sheet = wb['Config']
    
    # Clear existing data (keep header)
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    
    # Write new config
    for key, value in config.items():
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=key)
        # Serialize dict/list values as JSON for proper deserialization
        if isinstance(value, (dict, list)):
            sheet.cell(row=row, column=2, value=json.dumps(value))
        else:
            sheet.cell(row=row, column=2, value=str(value))
    
    _save_workbook(wb)
    wb.close()

# ============ Sample Data Initialization ============

def initialize_sample_employees():
    """Initialize the employee list when creating a new Excel file"""
    wb = _get_workbook()
    if not wb:
        return
    
    # Check if employees already exist
    existing = get_all_employees()
    if existing:
        return
    
    sample_employees = [
        Employee(id="emp1", name="Fran", employee_type=EmployeeType.MANAGER, max_hours_per_week=80),
        Employee(id="emp2", name="Aashima", employee_type=EmployeeType.MANAGER, max_hours_per_week=80),
        Employee(id="emp3", name="Mickaela C", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp4", name="Kavya C", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp5", name="Pablo 2", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp6", name="Viviana 3", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp7", name="Anastasia 3", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp8", name="MEG 3", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp9", name="Achal 2", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp10", name="PRIYANKA 2", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp11", name="Arta C", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp12", name="Taran C", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp13", name="Sagar C", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp14", name="Itshan 3", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp15", name="Arnob 2", employee_type=EmployeeType.EMPLOYEE, max_hours_per_week=24),
        Employee(id="emp16", name="Intern-Levi 7", employee_type=EmployeeType.INTERN, max_hours_per_week=15),
        Employee(id="emp17", name="Intern-Himanshu 7", employee_type=EmployeeType.INTERN, max_hours_per_week=15),
    ]
    
    for emp in sample_employees:
        save_employee(emp)

# ============ Initialization ============

def initialize_from_excel():
    """Initialize data from Excel file - call this on startup if file is set"""
    if not EXCEL_FILE_PATH or not os.path.exists(EXCEL_FILE_PATH):
        return
    
    # Ensure structure is correct
    ensure_excel_structure(EXCEL_FILE_PATH)

def get_all_week_schedule_dates() -> List[date]:
    """Get list of all weeks that have schedule tabs"""
    if not EXCEL_FILE_PATH or not os.path.exists(EXCEL_FILE_PATH):
        return []
    
    try:
        wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
        dates = []
        reference_year = date.today().year
        
        for sheet_name in wb.sheetnames:
            # Try legacy format first
            if sheet_name.startswith('Schedule_'):
                try:
                    date_str = sheet_name.replace('Schedule_', '').replace('_', '-')
                    week_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    dates.append(week_date)
                except:
                    pass
            else:
                # Try native format (e.g., "June 1-7")
                result = parse_schedule_sheet_dates(sheet_name, reference_year)
                if result:
                    dates.append(result[0])  # Use start date
        
        wb.close()
        return sorted(dates, reverse=True)
    except Exception as e:
        print(f"Error getting week schedule dates: {e}")
        return []

# ============ Availability Requests ============

def get_availability_requests() -> List[Dict]:
    """Get all availability requests from Excel"""
    # Clear cache to ensure we get fresh data
    _clear_workbook_cache()
    wb = _get_workbook()
    if not wb:
        return []

    try:
        if 'Availability_Requests' not in wb.sheetnames:
            wb.close()
            return []

        sheet = wb['Availability_Requests']
        requests = []

        # Check if sheet has new schema (has 'Request_Type' column at position 3)
        has_new_schema = sheet.cell(row=1, column=3).value == 'Request_Type'

        for row in range(2, sheet.max_row + 1):
            if has_new_schema:
                # New schema
                request = {
                    'id': sheet.cell(row=row, column=1).value,
                    'employee_id': sheet.cell(row=row, column=2).value,
                    'request_type': sheet.cell(row=row, column=3).value,
                    'start_date': sheet.cell(row=row, column=4).value,
                    'end_date': sheet.cell(row=row, column=5).value,
                    'days_of_week': sheet.cell(row=row, column=6).value,
                    'start_time': sheet.cell(row=row, column=7).value,
                    'end_time': sheet.cell(row=row, column=8).value,
                    'status': sheet.cell(row=row, column=9).value,
                    'manager_comment': sheet.cell(row=row, column=10).value,
                    'employee_comment': sheet.cell(row=row, column=11).value,
                    'preferences': sheet.cell(row=row, column=12).value,
                    'created_at': sheet.cell(row=row, column=13).value,
                    'updated_at': sheet.cell(row=row, column=14).value,
                    'approved_by': sheet.cell(row=row, column=15).value,
                    'approved_at': sheet.cell(row=row, column=16).value,
                    # Legacy fields for backward compatibility
                    'day_of_week': None,
                    'availability_type': None,
                    'week_start_date': None,
                    'description': None,
                }
                # Parse days_of_week JSON if present
                if request['days_of_week']:
                    try:
                        import json
                        request['days_of_week'] = json.loads(request['days_of_week'])
                    except:
                        pass
            else:
                # Old schema for backward compatibility
                request = {
                    'id': sheet.cell(row=row, column=1).value,
                    'employee_id': sheet.cell(row=row, column=2).value,
                    'day_of_week': sheet.cell(row=row, column=3).value,
                    'availability_type': sheet.cell(row=row, column=4).value,
                    'week_start_date': sheet.cell(row=row, column=5).value,
                    'status': sheet.cell(row=row, column=6).value,
                    'manager_comment': sheet.cell(row=row, column=7).value,
                    'description': sheet.cell(row=row, column=8).value,
                    'preferences': sheet.cell(row=row, column=9).value,
                    'created_at': sheet.cell(row=row, column=10).value,
                    'updated_at': sheet.cell(row=row, column=11).value,
                    # New fields for backward compatibility
                    'request_type': 'availability',
                    'start_date': None,
                    'end_date': None,
                    'days_of_week': None,
                    'start_time': None,
                    'end_time': None,
                    'employee_comment': None,
                    'approved_by': None,
                    'approved_at': None,
                }

            # Parse preferences JSON if present
            if request['preferences']:
                try:
                    request['preferences'] = json.loads(request['preferences'])
                except:
                    request['preferences'] = None
            requests.append(request)
        
        wb.close()
        return requests
    except Exception as e:
        print(f"Error reading availability requests: {e}")
        wb.close()
        return []

def save_availability_request(request: Dict) -> bool:
    """Save an availability request to Excel"""
    wb = _get_workbook()
    if not wb:
        return False

    try:
        import json

        if 'Availability_Requests' not in wb.sheetnames:
            sheet = wb.create_sheet('Availability_Requests')
            # New schema headers
            headers = ['ID', 'Employee_ID', 'Request_Type', 'Start_Date', 'End_Date', 'Days_of_Week', 'Start_Time', 'End_Time', 'Status', 'Manager_Comment', 'Employee_Comment', 'Preferences', 'Created_At', 'Updated_At', 'Approved_By', 'Approved_At']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
                sheet.cell(row=1, column=col).font = Font(bold=True)
        else:
            sheet = wb['Availability_Requests']
            # Check if sheet has new schema (has 'Request_Type' column)
            has_new_schema = sheet.cell(row=1, column=3).value == 'Request_Type'
            if not has_new_schema:
                # Migrate to new schema by recreating the sheet
                old_data = []
                for row in range(2, sheet.max_row + 1):
                    old_data.append({
                        'id': sheet.cell(row=row, column=1).value,
                        'employee_id': sheet.cell(row=row, column=2).value,
                        'day_of_week': sheet.cell(row=row, column=3).value,
                        'availability_type': sheet.cell(row=row, column=4).value,
                        'week_start_date': sheet.cell(row=row, column=5).value,
                        'status': sheet.cell(row=row, column=6).value,
                        'manager_comment': sheet.cell(row=row, column=7).value,
                        'description': sheet.cell(row=row, column=8).value,
                        'preferences': sheet.cell(row=row, column=9).value,
                        'created_at': sheet.cell(row=row, column=10).value,
                        'updated_at': sheet.cell(row=row, column=11).value,
                    })
                wb.remove(sheet)
                sheet = wb.create_sheet('Availability_Requests')
                headers = ['ID', 'Employee_ID', 'Request_Type', 'Start_Date', 'End_Date', 'Days_of_Week', 'Start_Time', 'End_Time', 'Status', 'Manager_Comment', 'Employee_Comment', 'Preferences', 'Created_At', 'Updated_At', 'Approved_By', 'Approved_At']
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col, value=header)
                    sheet.cell(row=1, column=col).font = Font(bold=True)
                # Migrate old data
                for i, old in enumerate(old_data, 2):
                    sheet.cell(row=i, column=1).value = old['id']
                    sheet.cell(row=i, column=2).value = old['employee_id']
                    # Determine request_type based on availability_type
                    avail_type = old.get('availability_type', '').lower()
                    if avail_type in ['day_off', 'off']:
                        sheet.cell(row=i, column=3).value = 'day_off'
                    else:
                        sheet.cell(row=i, column=3).value = 'availability'
                    sheet.cell(row=i, column=4).value = old.get('week_start_date')
                    sheet.cell(row=i, column=5).value = old.get('week_start_date')
                    # Only set days_of_week if it's a valid day name (not an availability type)
                    day_of_week = old.get('day_of_week', '').lower()
                    if day_of_week in ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']:
                        sheet.cell(row=i, column=6).value = json.dumps([day_of_week])
                    else:
                        sheet.cell(row=i, column=6).value = None
                    sheet.cell(row=i, column=7).value = None  # No start_time in old schema
                    sheet.cell(row=i, column=8).value = None  # No end_time in old schema
                    sheet.cell(row=i, column=9).value = old['status']
                    sheet.cell(row=i, column=10).value = old['manager_comment']
                    sheet.cell(row=i, column=11).value = old['description']
                    sheet.cell(row=i, column=12).value = old['preferences']
                    sheet.cell(row=i, column=13).value = old['created_at']
                    sheet.cell(row=i, column=14).value = old['updated_at']
                    sheet.cell(row=i, column=15).value = None
                    sheet.cell(row=i, column=16).value = None

        # Check if request already exists
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == request['id']:
                # Update existing
                sheet.cell(row=row, column=3).value = request.get('request_type')
                sheet.cell(row=row, column=4).value = request.get('start_date')
                sheet.cell(row=row, column=5).value = request.get('end_date')
                sheet.cell(row=row, column=6).value = json.dumps(request.get('days_of_week')) if request.get('days_of_week') else None
                sheet.cell(row=row, column=7).value = request.get('start_time')
                sheet.cell(row=row, column=8).value = request.get('end_time')
                sheet.cell(row=row, column=9).value = request['status']
                sheet.cell(row=row, column=10).value = request.get('manager_comment')
                sheet.cell(row=row, column=11).value = request.get('employee_comment')
                sheet.cell(row=row, column=12).value = str(request.get('preferences', {})) if request.get('preferences') else None
                sheet.cell(row=row, column=14).value = request.get('updated_at')
                sheet.cell(row=row, column=15).value = request.get('approved_by')
                sheet.cell(row=row, column=16).value = request.get('approved_at')
                _save_workbook(wb)
                wb.close()
                return True

        # Add new request
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1).value = request['id']
        sheet.cell(row=row, column=2).value = request['employee_id']
        sheet.cell(row=row, column=3).value = request.get('request_type')
        sheet.cell(row=row, column=4).value = request.get('start_date')
        sheet.cell(row=row, column=5).value = request.get('end_date')
        sheet.cell(row=row, column=6).value = json.dumps(request.get('days_of_week')) if request.get('days_of_week') else None
        sheet.cell(row=row, column=7).value = request.get('start_time')
        sheet.cell(row=row, column=8).value = request.get('end_time')
        sheet.cell(row=row, column=9).value = request['status']
        sheet.cell(row=row, column=10).value = request.get('manager_comment')
        sheet.cell(row=row, column=11).value = request.get('employee_comment')
        sheet.cell(row=row, column=12).value = str(request.get('preferences', {})) if request.get('preferences') else None
        sheet.cell(row=row, column=13).value = request.get('created_at')
        sheet.cell(row=row, column=14).value = request.get('updated_at')
        sheet.cell(row=row, column=15).value = request.get('approved_by')
        sheet.cell(row=row, column=16).value = request.get('approved_at')

        _save_workbook(wb)
        wb.close()
        return True
    except Exception as e:
        print(f"Error saving availability request: {e}")
        wb.close()
        return False

# ============ Notifications ============

def get_notifications(employee_id: str) -> List[Dict]:
    """Get all notifications for an employee"""
    wb = _get_workbook()
    if not wb:
        return []
    
    try:
        if 'Notifications' not in wb.sheetnames:
            wb.close()
            return []
        
        sheet = wb['Notifications']
        notifications = []
        
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value == employee_id:
                notification = {
                    'id': sheet.cell(row=row, column=1).value,
                    'employee_id': sheet.cell(row=row, column=2).value,
                    'type': sheet.cell(row=row, column=3).value,
                    'message': sheet.cell(row=row, column=4).value,
                    'created_at': sheet.cell(row=row, column=5).value,
                    'read': sheet.cell(row=row, column=6).value
                }
                notifications.append(notification)
        
        wb.close()
        return sorted(notifications, key=lambda x: x['created_at'], reverse=True)
    except Exception as e:
        print(f"Error reading notifications: {e}")
        wb.close()
        return []

def save_notification(notification: Dict) -> bool:
    """Save a notification to Excel"""
    wb = _get_workbook()
    if not wb:
        return False
    
    try:
        if 'Notifications' not in wb.sheetnames:
            sheet = wb.create_sheet('Notifications')
            headers = ['ID', 'Employee_ID', 'Type', 'Message', 'Created_At', 'Read']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
                sheet.cell(row=1, column=col).font = Font(bold=True)
        else:
            sheet = wb['Notifications']
        
        # Add new notification
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1).value = notification['id']
        sheet.cell(row=row, column=2).value = notification['employee_id']
        sheet.cell(row=row, column=3).value = notification['type']
        sheet.cell(row=row, column=4).value = notification['message']
        sheet.cell(row=row, column=5).value = notification['created_at']
        sheet.cell(row=row, column=6).value = notification.get('read', False)
        
        _save_workbook(wb)
        wb.close()
        return True
    except Exception as e:
        print(f"Error saving notification: {e}")
        wb.close()
        return False

def mark_notification_read(notification_id: str) -> bool:
    """Mark a notification as read"""
    wb = _get_workbook()
    if not wb:
        return False
    
    try:
        if 'Notifications' not in wb.sheetnames:
            wb.close()
            return False
        
        sheet = wb['Notifications']
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == notification_id:
                sheet.cell(row=row, column=6).value = True
                _save_workbook(wb)
                wb.close()
                return True
        
        wb.close()
        return False
    except Exception as e:
        print(f"Error marking notification as read: {e}")
        wb.close()
        return False

# ============ Events ============
def get_events(week_start_date: Optional[date] = None) -> List[Event]:
    """Get events from Excel, optionally filtered by week"""
    wb = _get_workbook()
    if not wb:
        return []
    
    try:
        if 'Events' not in wb.sheetnames:
            wb.close()
            return []
        
        sheet = wb['Events']
        events = []
        
        for row in range(2, sheet.max_row + 1):
            event_week_date = sheet.cell(row=row, column=3).value
            event_date = sheet.cell(row=row, column=4).value
            
            # Convert datetime to date for comparison
            if event_date and hasattr(event_date, 'date'):
                event_date = event_date.date()
            
            # Filter by week: check if event date falls within the week (week_start_date to week_start_date + 6 days)
            if week_start_date and event_date:
                from datetime import timedelta
                week_end = week_start_date + timedelta(days=6)
                if not (week_start_date <= event_date <= week_end):
                    continue
            
            event = Event(
                id=sheet.cell(row=row, column=1).value,
                name=sheet.cell(row=row, column=2).value,
                week_start_date=event_week_date,
                date=sheet.cell(row=row, column=4).value,
                start_time=sheet.cell(row=row, column=5).value,
                end_time=sheet.cell(row=row, column=6).value,
                location=sheet.cell(row=row, column=7).value,
                people_needed=sheet.cell(row=row, column=8).value,
                description=sheet.cell(row=row, column=9).value,
                created_by=sheet.cell(row=row, column=10).value,
                created_at=sheet.cell(row=row, column=11).value,
                updated_at=sheet.cell(row=row, column=12).value
            )
            events.append(event)
        
        wb.close()
        return events
    except Exception as e:
        print(f"Error getting events: {e}")
        wb.close()
        return []

def save_event(event: Event) -> Event:
    """Save an event to Excel"""
    wb = _get_workbook()
    if not wb:
        raise Exception("Could not get workbook")
    
    try:
        if 'Events' not in wb.sheetnames:
            sheet = wb.create_sheet('Events')
            headers = ['ID', 'Name', 'Week_Start_Date', 'Date', 'Start_Time', 'End_Time', 'Location', 'People_Needed', 'Description', 'Created_By', 'Created_At', 'Updated_At']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
                sheet.cell(row=1, column=col).font = Font(bold=True)
        else:
            sheet = wb['Events']
        
        # Check if event already exists
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == event.id:
                # Update existing
                sheet.cell(row=row, column=2).value = event.name
                sheet.cell(row=row, column=3).value = event.week_start_date
                sheet.cell(row=row, column=4).value = event.date
                sheet.cell(row=row, column=5).value = event.start_time
                sheet.cell(row=row, column=6).value = event.end_time
                sheet.cell(row=row, column=7).value = event.location
                sheet.cell(row=row, column=8).value = event.people_needed
                sheet.cell(row=row, column=9).value = event.description
                sheet.cell(row=row, column=12).value = event.updated_at
                _save_workbook(wb)
                wb.close()
                return event
        
        # Add new event
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1).value = event.id
        sheet.cell(row=row, column=2).value = event.name
        sheet.cell(row=row, column=3).value = event.week_start_date
        sheet.cell(row=row, column=4).value = event.date
        sheet.cell(row=row, column=5).value = event.start_time
        sheet.cell(row=row, column=6).value = event.end_time
        sheet.cell(row=row, column=7).value = event.location
        sheet.cell(row=row, column=8).value = event.people_needed
        sheet.cell(row=row, column=9).value = event.description
        sheet.cell(row=row, column=10).value = event.created_by
        sheet.cell(row=row, column=11).value = event.created_at
        sheet.cell(row=row, column=12).value = event.updated_at
        
        _save_workbook(wb)
        wb.close()
        return event
    except Exception as e:
        print(f"Error saving event: {e}")
        wb.close()
        raise

def delete_event(event_id: str) -> bool:
    """Delete an event from Excel"""
    wb = _get_workbook()
    if not wb:
        return False
    
    try:
        if 'Events' not in wb.sheetnames:
            wb.close()
            return False
        
        sheet = wb['Events']
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == event_id:
                sheet.delete_rows(row)
                _save_workbook(wb)
                wb.close()
                return True
        
        wb.close()
        return False
    except Exception as e:
        print(f"Error deleting event: {e}")
        wb.close()
        return False

def initialize_sample_data():
    """Initialize Excel with sample data if empty"""
    if not get_all_employees():
        sample_employees = [
            Employee(id="emp1", name="Fran", employee_type="manager", max_hours_per_week=80),
            Employee(id="emp2", name="Aashima", employee_type="manager", max_hours_per_week=80),
            Employee(id="emp3", name="Mickaela C", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp4", name="Kavya C", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp5", name="Pablo", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp6", name="Viviana", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp7", name="Anastasia", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp8", name="Meg", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp9", name="Achal", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp10", name="Priyanka", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp11", name="Arfa C", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp12", name="Taran C", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp13", name="Sagar C", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp14", name="Itohan", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp15", name="Arnob", employee_type="employee", max_hours_per_week=24),
            Employee(id="emp16", name="Nahim", employee_type="employee", max_hours_per_week=30),
        ]
        for emp in sample_employees:
            save_employee(emp)
        print("[INIT] Initialized Excel with sample employees")
