"""Create Employees sheet in Excel file on GitHub with current employees"""
import os
import io
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from github_storage import github_get_file, github_put_file, GITHUB_AVAILABLE

# Helper function to normalize employee names
def normalize_employee_name(name):
    """Remove single letter/number suffixes to get base name.
    - 'Sagar C' -> 'Sagar'
    - 'Pablo 2' -> 'Pablo'
    - 'Sagar Ca' -> 'Sagar Ca' (keep multi-letter suffixes)
    """
    parts = name.rsplit(' ', 1)
    if len(parts) == 2:
        base, suffix = parts
        # If suffix is a single character (letter or number), remove it
        if len(suffix) == 1:
            return base
    return name

# Definitive employee list with correct IDs matching schedule data
# Using base names (without single-letter/number suffixes)
current_employees = [
    {'id': 'emp_001', 'name': 'Fran',        'type': 'manager',  'max_hours': 40},
    {'id': 'emp_002', 'name': 'Aashima',      'type': 'manager',  'max_hours': 40},
    {'id': 'emp_003', 'name': 'NAHIM',        'type': 'employee', 'max_hours': 24},
    {'id': 'emp_004', 'name': 'Pablo',        'type': 'employee', 'max_hours': 24},
    {'id': 'emp_005', 'name': 'Kavya',        'type': 'employee', 'max_hours': 24},
    {'id': 'emp_006', 'name': 'Mickaela',     'type': 'employee', 'max_hours': 24},
    {'id': 'emp_007', 'name': 'Viviana',      'type': 'employee', 'max_hours': 24},
    {'id': 'emp_008', 'name': 'Anastasia',    'type': 'employee', 'max_hours': 24},
    {'id': 'emp_009', 'name': 'MEG',          'type': 'employee', 'max_hours': 24},
    {'id': 'emp_010', 'name': 'Achal',        'type': 'employee', 'max_hours': 24},
    {'id': 'emp_011', 'name': 'PRIYANKA',     'type': 'employee', 'max_hours': 24},
    {'id': 'emp_012', 'name': 'Arfa',         'type': 'employee', 'max_hours': 24},
    {'id': 'emp_013', 'name': 'Taran',        'type': 'employee', 'max_hours': 24},
    {'id': 'emp_014', 'name': 'Sagar',        'type': 'employee', 'max_hours': 24},
    {'id': 'emp_015', 'name': 'Itohan',       'type': 'employee', 'max_hours': 24},
    {'id': 'emp_016', 'name': 'Arnob',        'type': 'employee', 'max_hours': 24},
    {'id': 'emp_017', 'name': 'Intern Chidera','type': 'intern',   'max_hours': 15},
    {'id': 'emp_018', 'name': 'Intern Heena', 'type': 'intern',   'max_hours': 15},
]

# Load preferences from local JSON if available
with open('data/employees.json', 'r') as f:
    employees_data = json.load(f)

prefs_by_name = {}
for emp in employees_data:
    # Match by normalized name
    normalized_name = normalize_employee_name(emp['name'])
    prefs_by_name[normalized_name] = emp.get('preferences', {})

for emp in current_employees:
    emp['email'] = ''
    emp['preferences'] = prefs_by_name.get(emp['name'], {})
    emp['active'] = True
    emp['created_at'] = '2026-01-01'

print(f"Found {len(current_employees)} employees to keep")

if not GITHUB_AVAILABLE:
    print("GitHub storage not available")
    exit(1)

print("Downloading Excel file from GitHub...")
data = github_get_file()
if not data:
    print("Failed to download")
    exit(1)

print(f"Downloaded {len(data)} bytes")

wb = load_workbook(io.BytesIO(data))

# Create Employees sheet if it doesn't exist
if 'Employees' in wb.sheetnames:
    print("Employees sheet already exists, removing it...")
    wb.remove(wb['Employees'])

# Create Employees sheet at position 2 (after PWDs)
emp_sheet = wb.create_sheet('Employees', 2)

# Add headers
headers = ['ID', 'Name', 'Email', 'Type', 'Max_Hours', 'Preferences', 'Active', 'Created_At']
for col, header in enumerate(headers, 1):
    cell = emp_sheet.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='9BBB59', end_color='9BBB59', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF')

# Add employee data
for row_idx, emp in enumerate(current_employees, 2):
    emp_sheet.cell(row=row_idx, column=1, value=emp['id'])
    emp_sheet.cell(row=row_idx, column=2, value=emp['name'])
    emp_sheet.cell(row=row_idx, column=3, value=emp['email'])
    emp_sheet.cell(row=row_idx, column=4, value=emp['type'])
    emp_sheet.cell(row=row_idx, column=5, value=emp['max_hours'])
    # Convert preferences dict to string for Excel
    if isinstance(emp['preferences'], dict):
        emp_sheet.cell(row=row_idx, column=6, value=json.dumps(emp['preferences']))
    else:
        emp_sheet.cell(row=row_idx, column=6, value=emp['preferences'])
    emp_sheet.cell(row=row_idx, column=7, value=emp['active'])
    emp_sheet.cell(row=row_idx, column=8, value=emp['created_at'])

# Set column widths
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    emp_sheet.column_dimensions[col_letter].width = 20

print(f"Created Employees sheet with {len(current_employees)} employees")

# Save to buffer
buffer = io.BytesIO()
wb.save(buffer)
buffer.seek(0)
updated_data = buffer.read()
wb.close()

print(f"Uploading updated Excel file to GitHub ({len(updated_data)} bytes)...")
if github_put_file(updated_data, message="Create Employees sheet with current employees"):
    print("Successfully updated Excel file on GitHub!")
    print("\nThe Employees sheet now has only the 6 current employees.")
    print("Employee deletions should now persist correctly.")
else:
    print("Failed to upload to GitHub")
