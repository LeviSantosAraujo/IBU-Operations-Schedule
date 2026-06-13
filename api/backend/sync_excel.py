"""
Sync Excel file from Blob storage data
Run this daily to backup JSON data to Excel format
"""
import json
import io
from openpyxl import Workbook, load_workbook
from storage import get_excel_data, store_excel_data
from data_store import get_all_employees, get_all_schedules, get_availabilities, get_system_config
from excel_store import ensure_excel_structure

def sync_to_excel():
    """Sync current JSON data to Excel file format"""
    print("Starting Excel sync...")
    
    # Load existing Excel or create new structure
    existing_data = get_excel_data("ibu_schedule.xlsx")
    if existing_data:
        wb = load_workbook(io.BytesIO(existing_data))
    else:
        wb = Workbook()
        ensure_excel_structure(wb)
    
    # Update Employees sheet
    employees = get_all_employees()
    emp_sheet = wb["Employees"]
    # Clear existing data (keep header)
    for row in list(emp_sheet.iter_rows(min_row=2)):
        for cell in row:
            cell.value = None
    # Write current employees
    for i, emp in enumerate(employees, start=2):
        emp_sheet.cell(row=i, column=1, value=emp.id)
        emp_sheet.cell(row=i, column=2, value=emp.name)
        emp_sheet.cell(row=i, column=3, value=emp.email or "")
        emp_sheet.cell(row=i, column=4, value=emp.employee_type)
        emp_sheet.cell(row=i, column=5, value=emp.max_hours_per_week)
        emp_sheet.cell(row=i, column=6, value=json.dumps(emp.preferences))
        emp_sheet.cell(row=i, column=7, value=json.dumps(emp.manager_preferences))
        emp_sheet.cell(row=i, column=8, value=emp.active)
    
    # Update Config sheet
    config = get_system_config()
    config_sheet = wb["Config"]
    config_sheet.cell(row=2, column=1, value=json.dumps(config.model_dump()))
    
    # Save to storage
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    store_excel_data(buffer.read(), "ibu_schedule.xlsx")
    
    print(f"Excel sync complete: {len(employees)} employees saved")

if __name__ == "__main__":
    sync_to_excel()
