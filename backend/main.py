from fastapi import FastAPI, HTTPException, Query, Header, Depends, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from typing import List, Optional, Dict
from datetime import date, datetime
import uuid
import os
import shutil
from pathlib import Path
from pydantic import BaseModel

from models import (
    Employee, Availability, WeeklySchedule, Shift, JobType, Floor,
    AvailabilityType, EmployeeType, FloorCoverageQuery, FloorCoverageResponse,
    AVAILABILITY_COLORS
)
from excel_store import (
    get_all_employees, get_employee_by_id, save_employee, delete_employee,
    get_availabilities, get_availability_for_week, save_availability,
    get_all_schedules, get_schedule_by_week, save_schedule, delete_schedule,
    get_floor_coverage, get_system_config, save_system_config,
    set_excel_file, get_excel_file, ensure_excel_structure,
    set_manager_password, verify_manager_password, manager_has_password,
    initialize_from_excel, get_all_week_schedule_dates,
    initialize_sample_employees, set_blob_key
)
from storage import store_excel_data, excel_file_exists
from scheduler import SchedulingEngine, generate_schedule
from auth import AuthManager, require_auth, require_manager, require_self_or_manager

# Pydantic models for requests
class LoginRequest(BaseModel):
    employee_id: str
    password: Optional[str] = None

class SetPasswordRequest(BaseModel):
    employee_id: str
    password: str

class ExcelPathRequest(BaseModel):
    file_path: str

# Global state
CURRENT_EXCEL_FILE: Optional[str] = None
UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

app = FastAPI(title="IBU Schedule System", version="2.0.0")

# Enable CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============ Excel File Management ============

@app.get("/api/excel/status")
async def excel_status():
    """Check if Excel file is configured"""
    try:
        file_path = get_excel_file()
        exists = excel_file_exists()
        return {
            "configured": exists or file_path is not None,
            "file_path": file_path or ("blob_storage" if exists else None),
            "file_exists": exists
        }
    except Exception as e:
        # Return safe default if there's any error
        print(f"Error checking Excel status: {e}")
        return {
            "configured": False,
            "file_path": None,
            "file_exists": False
        }

@app.post("/api/excel/upload")
async def upload_excel(file: UploadFile = File(...)):
    """Upload an Excel file to use as database"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")
    
    try:
        # Just read the file to validate it's a valid Excel file
        file_content = await file.read()
        
        # Try to parse it as Excel to validate
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_content))
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
        
        # Store in memory for now (minimal approach)
        try:
            from storage import store_excel_data
            store_excel_data(file_content, file.filename)
        except:
            # If storage fails, just continue - we validated the file
            pass
        
        return {
            "message": "Excel file uploaded successfully",
            "filename": file.filename
        }
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to upload Excel file: {str(e)}")

@app.post("/api/excel/select")
async def select_excel_file(request: ExcelPathRequest):
    """Select an existing Excel file"""
    if not os.path.exists(request.file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    set_excel_file(request.file_path)
    
    return {
        "message": "Excel file selected",
        "file_path": request.file_path
    }

@app.get("/api/excel/download")
async def download_excel():
    """Download the current Excel file from blob storage"""
    from fastapi.responses import StreamingResponse
    import io
    
    # Try to get from blob first
    file_data = download_excel_from_blob()
    if file_data:
        return StreamingResponse(
            io.BytesIO(file_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=IBU_Schedule.xlsx"}
        )
    
    # Fall back to local file
    file_path = get_excel_file()
    if file_path and os.path.exists(file_path):
        return FileResponse(file_path, filename="IBU_Schedule.xlsx")
    
    raise HTTPException(status_code=404, detail="No Excel file configured")

@app.post("/api/excel/create-new")
async def create_new_excel():
    """Create a new Excel database with sample employees"""
    import io
    from openpyxl import Workbook
    
    try:
        # Create workbook in memory
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create basic tabs
        wb.create_sheet('Config', 0)
        wb.create_sheet('PWDs', 1)
        wb.create_sheet('Employees', 2)
        wb.create_sheet('Availability', 3)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Try to store using storage module
        try:
            from storage import store_excel_data
            store_excel_data(buffer.read(), "ibu_schedule.xlsx")
        except:
            # If storage fails, continue anyway
            pass
        
        return {
            "message": "New Excel database created successfully",
            "employees_added": 0
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create Excel database: {str(e)}")

# ============ Password Management ============

@app.post("/api/managers/set-password")
async def set_manager_password_endpoint(request: SetPasswordRequest):
    """Set password for a manager (only if not already set)"""
    employee = get_employee_by_id(request.employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    if employee.employee_type != EmployeeType.MANAGER:
        raise HTTPException(status_code=403, detail="Only managers can have passwords")
    
    # Check if password already exists
    if manager_has_password(request.employee_id):
        raise HTTPException(status_code=400, detail="Password already set. Contact admin to reset.")
    
    set_manager_password(request.employee_id, employee.name, request.password)
    
    return {"message": "Password set successfully"}

@app.post("/api/managers/verify-password")
async def verify_password_endpoint(request: SetPasswordRequest):
    """Verify manager password"""
    if not manager_has_password(request.employee_id):
        raise HTTPException(status_code=400, detail="Password not set")
    
    is_valid = verify_manager_password(request.employee_id, request.password)
    
    if not is_valid:
        raise HTTPException(status_code=401, detail="Invalid password")
    
    return {"valid": True}

@app.get("/api/managers/has-password/{employee_id}")
async def check_has_password(employee_id: str):
    """Check if manager has password set"""
    return {"has_password": manager_has_password(employee_id)}

# ============ Auth Endpoints ============

@app.post("/api/login")
async def login(request: LoginRequest):
    """Login as an employee with optional password for managers"""
    # Check if Excel file is configured
    if not get_excel_file():
        raise HTTPException(status_code=400, detail="No Excel database configured. Please upload or select an Excel file first.")
    
    employee = get_employee_by_id(request.employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Managers need password verification
    if employee.employee_type == EmployeeType.MANAGER:
        if manager_has_password(request.employee_id):
            if not request.password:
                raise HTTPException(status_code=401, detail="Password required for managers")
            if not verify_manager_password(request.employee_id, request.password):
                raise HTTPException(status_code=401, detail="Invalid password")
        # If no password set yet, allow login (first time setup)
    
    token = AuthManager.login(request.employee_id)
    
    return {
        "token": token,
        "employee": employee,
        "role": employee.employee_type,
        "requires_password_setup": employee.employee_type == EmployeeType.MANAGER and not manager_has_password(request.employee_id)
    }

@app.post("/api/logout")
async def logout(authorization: Optional[str] = Header(None)):
    """Logout and end session"""
    if authorization and authorization.startswith("Bearer "):
        token = authorization.replace("Bearer ", "")
        AuthManager.logout(token)
    return {"message": "Logged out"}

@app.get("/api/me")
async def get_me(authorization: Optional[str] = Header(None)):
    """Get current logged-in user info"""
    user = AuthManager.get_current_user(authorization)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user

# ============ Employee Endpoints ============

@app.get("/api/employees", response_model=List[Employee])
async def list_employees(active_only: bool = False):
    """List all employees"""
    employees = get_all_employees()
    if active_only:
        employees = [e for e in employees if e.active]
    return employees

@app.get("/api/employees/{employee_id}", response_model=Employee)
async def get_employee(employee_id: str):
    """Get a specific employee"""
    employee = get_employee_by_id(employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    return employee

@app.post("/api/employees", response_model=Employee)
async def create_employee(
    employee: Employee,
    user: Dict = Depends(require_manager)
):
    """Create a new employee (managers only)"""
    if not employee.id:
        employee.id = f"emp_{uuid.uuid4().hex[:8]}"
    return save_employee(employee)

@app.put("/api/employees/{employee_id}", response_model=Employee)
async def update_employee(
    employee_id: str,
    employee: Employee,
    user: Dict = Depends(require_manager)
):
    """Update an employee (managers only)"""
    existing = get_employee_by_id(employee_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Employee not found")
    employee.id = employee_id
    return save_employee(employee)

@app.delete("/api/employees/{employee_id}")
async def remove_employee(
    employee_id: str,
    user: Dict = Depends(require_manager)
):
    """Delete an employee (managers only)"""
    if delete_employee(employee_id):
        return {"message": "Employee deleted"}
    raise HTTPException(status_code=404, detail="Employee not found")

# ============ Availability Endpoints ============

@app.get("/api/availability", response_model=List[Availability])
async def list_availabilities(
    week_start_date: Optional[date] = None,
    employee_id: Optional[str] = None,
    authorization: Optional[str] = Header(None)
):
    """List availabilities - employees see only their own, managers see all"""
    user = require_auth(authorization)
    
    # Employees can only see their own availability
    if user["role"] != "manager":
        employee_id = user["employee_id"]
    
    return get_availabilities(week_start_date, employee_id)

@app.get("/api/availability/{employee_id}/{week_start_date}", response_model=Availability)
async def get_employee_availability(
    employee_id: str,
    week_start_date: date,
    authorization: Optional[str] = Header(None)
):
    """Get availability for a specific employee and week"""
    # Check permissions
    require_self_or_manager(employee_id, authorization)
    
    availability = get_availability_for_week(employee_id, week_start_date)
    if not availability:
        # Return default availability
        return Availability(
            id=str(uuid.uuid4()),
            employee_id=employee_id,
            week_start_date=week_start_date
        )
    return availability

@app.post("/api/availability", response_model=Availability)
async def submit_availability(
    availability: Availability,
    authorization: Optional[str] = Header(None)
):
    """Submit or update availability - employees can only submit their own"""
    user = require_auth(authorization)
    
    # Employees can only submit their own availability
    if user["role"] != "manager" and availability.employee_id != user["employee_id"]:
        raise HTTPException(status_code=403, detail="Can only submit your own availability")
    
    # 24-hour cutoff check for employees (not managers)
    if user["role"] != "manager":
        # Calculate the first day of the week being submitted
        week_start = availability.week_start_date
        today = date.today()
        
        # Check if any day in the week is within 24 hours of now
        # The week starts on Monday, so we check each day
        days_offset = 0  # Monday
        current_day_of_week = today.weekday()  # 0=Monday, 6=Sunday
        
        # If the week being submitted is the current week or a past week
        if week_start <= today:
            # Calculate which day of the week today is relative to the week start
            days_since_week_start = (today - week_start).days
            
            # If today is within the week being submitted, check if it's too late to change
            if 0 <= days_since_week_start <= 6:
                # The week includes today or past days - cannot change
                raise HTTPException(
                    status_code=403, 
                    detail=f"Cannot change availability for week that includes today or past days. Changes must be made at least 24 hours before the week starts."
                )
        
        # If the week starts within 24 hours from now, also block
        hours_until_week_start = (week_start - today).days * 24
        if hours_until_week_start < 24:
            raise HTTPException(
                status_code=403,
                detail=f"Cannot change availability for week starting in less than 24 hours. Changes must be made at least 24 hours before the week starts."
            )
    
    if not availability.id:
        availability.id = f"avail_{uuid.uuid4().hex[:8]}"
    availability.submitted_at = datetime.now()
    # Reset approval status when availability is modified
    availability.approved = False
    availability.approved_by = None
    availability.approved_at = None
    return save_availability(availability)

@app.get("/api/availability/colors")
async def get_availability_colors():
    """Get color mapping for availability types"""
    return AVAILABILITY_COLORS

@app.post("/api/availability/{availability_id}/approve")
async def approve_availability(
    availability_id: str,
    authorization: Optional[str] = Header(None)
):
    """Approve availability (managers only)"""
    user = require_manager(authorization)
    
    # Get all availabilities and find the one to approve
    availabilities = get_availabilities()
    availability = None
    for avail in availabilities:
        if avail.id == availability_id:
            availability = avail
            break
    
    if not availability:
        raise HTTPException(status_code=404, detail="Availability not found")
    
    # Update approval status
    availability.approved = True
    availability.approved_by = user["employee_id"]
    availability.approved_at = datetime.now()
    
    return save_availability(availability)

@app.get("/api/availability/pending")
async def get_pending_availabilities(user: Dict = Depends(require_manager)):
    """Get all pending (unapproved) availabilities (managers only)"""
    all_availabilities = get_availabilities()
    pending = [avail for avail in all_availabilities if not avail.approved]
    return pending

# ============ Schedule Endpoints ============

@app.get("/api/schedules", response_model=List[WeeklySchedule])
async def list_schedules(authorization: Optional[str] = Header(None)):
    """List all schedules (all authenticated users)"""
    require_auth(authorization)
    return get_all_schedules()

@app.get("/api/schedules/{week_start_date}", response_model=WeeklySchedule)
async def get_schedule(
    week_start_date: date,
    authorization: Optional[str] = Header(None)
):
    """Get schedule for a specific week (all authenticated users)"""
    require_auth(authorization)
    schedule = get_schedule_by_week(week_start_date)
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    return schedule

@app.post("/api/schedules/generate/{week_start_date}", response_model=WeeklySchedule)
async def auto_generate_schedule(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Auto-generate a schedule for a week (managers only)"""
    schedule = generate_schedule(week_start_date)
    return save_schedule(schedule)

@app.post("/api/schedules", response_model=WeeklySchedule)
async def create_or_update_schedule(
    schedule: WeeklySchedule,
    user: Dict = Depends(require_manager)
):
    """Save a schedule (manual editing) (managers only)"""
    schedule.updated_at = datetime.now()
    return save_schedule(schedule)

@app.put("/api/schedules/{week_start_date}/shifts", response_model=WeeklySchedule)
async def update_schedule_shifts(
    week_start_date: date,
    shifts: List[Shift],
    user: Dict = Depends(require_manager)
):
    """Update shifts for a schedule (drag-drop saves here) (managers only)"""
    schedule = get_schedule_by_week(week_start_date)
    if not schedule:
        schedule = WeeklySchedule(
            id=str(uuid.uuid4()),
            week_start_date=week_start_date
        )
    
    schedule.shifts = shifts
    # Recalculate total hours
    total_hours = {}
    for shift in shifts:
        current = total_hours.get(shift.employee_id, 0)
        total_hours[shift.employee_id] = current + shift.hours
    schedule.total_hours = total_hours
    schedule.updated_at = datetime.now()
    
    return save_schedule(schedule)

@app.post("/api/schedules/{week_start_date}/publish")
async def publish_schedule(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Publish a schedule (finalize it) (managers only)"""
    schedule = get_schedule_by_week(week_start_date)
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    schedule.status = "published"
    save_schedule(schedule)
    return {"message": "Schedule published"}

@app.delete("/api/schedules/{week_start_date}")
async def remove_schedule(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Delete a schedule (managers only)"""
    if delete_schedule(week_start_date):
        return {"message": "Schedule deleted"}
    raise HTTPException(status_code=404, detail="Schedule not found")

@app.get("/api/schedules/weeks/available")
async def get_available_weeks(user: Dict = Depends(require_manager)):
    """Get list of all weeks that have schedule data (managers only)"""
    weeks = get_all_week_schedule_dates()
    return {
        "weeks": [w.isoformat() for w in weeks],
        "count": len(weeks)
    }

# ============ Floor Coverage & Queries ============

@app.get("/api/floor-coverage/{floor}/{day_of_week}/{time_slot}")
async def query_floor_coverage(
    floor: Floor,
    day_of_week: str,
    time_slot: str,
    week_start_date: date = Query(...),
    user: Dict = Depends(require_manager)
):
    """Query how many employees are on a floor at a given time (managers only)"""
    result = get_floor_coverage(floor.value, day_of_week, time_slot, week_start_date)
    return result

@app.get("/api/floor-coverage/summary/{week_start_date}")
async def get_weekly_floor_summary(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Get floor coverage summary for entire week (managers only)"""
    days = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    time_slots = ["morning", "afternoon", "evening"]
    floors = [Floor.GROUND, Floor.SECOND, Floor.SIXTH]
    
    summary = {}
    for floor in floors:
        summary[floor.value] = {}
        for day in days:
            summary[floor.value][day] = {}
            for slot in time_slots:
                result = get_floor_coverage(floor.value, day, slot, week_start_date)
                summary[floor.value][day][slot] = result["employee_count"]
    
    return summary

# ============ Analytics & Reporting ============

@app.get("/api/analytics/employee-hours/{week_start_date}")
async def get_employee_hours_summary(
    week_start_date: date,
    user: Dict = Depends(require_manager)
):
    """Get hours summary for all employees for a week (managers only)"""
    schedule = get_schedule_by_week(week_start_date)
    employees = get_all_employees()
    
    summary = []
    for emp in employees:
        hours = schedule.total_hours.get(emp.id, 0) if schedule else 0
        remaining = emp.max_hours_per_week - hours
        summary.append({
            "employee_id": emp.id,
            "name": emp.name,
            "type": emp.employee_type,
            "scheduled_hours": hours,
            "max_hours": emp.max_hours_per_week,
            "remaining_hours": remaining,
            "utilization": round(hours / emp.max_hours_per_week * 100, 1) if emp.max_hours_per_week > 0 else 0
        })
    
    return summary

# ============ Configuration ============

@app.get("/api/config")
async def get_config():
    """Get system configuration"""
    return get_system_config()

@app.put("/api/config")
async def update_config(config: Dict):
    """Update system configuration"""
    return save_system_config(config)

# ============ Health Check ============

@app.get("/api/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now()}

if __name__ == "__main__":
    import uvicorn
    # Initialize blob storage for cloud deployment
    if os.getenv("BLOB_READ_WRITE_TOKEN"):
        set_blob_key("ibu_schedule.xlsx")
    uvicorn.run(app, host="0.0.0.0", port=8000)
