"""
Excel-based data storage system for IBU Schedule
- Uses Excel as the primary database
- Creates tabs: Config, PWDs, Employees, Availability, and weekly schedule tabs
- Real-time read/write to Excel
- Supports both local file system and Vercel Blob storage
"""

import os
import json
from datetime import date, datetime
from typing import List, Optional, Dict, Any
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from models import Employee, Availability, WeeklySchedule, Shift, EmployeeType, AvailabilityType, JobType, Floor
import io

# Import storage module
from storage import get_workbook, save_workbook, excel_file_exists as storage_file_exists

# Global path to the current Excel file (local storage)
EXCEL_FILE_PATH: Optional[str] = None
# Blob storage key (for cloud storage)
BLOB_KEY: Optional[str] = "ibu_schedule.xlsx"

def set_excel_file(path: str):
    """Set the active Excel file path (local storage)"""
    global EXCEL_FILE_PATH
    EXCEL_FILE_PATH = path
    # Ensure file exists with proper structure
    ensure_excel_structure(path)

def get_excel_file() -> Optional[str]:
    """Get the current Excel file path (local storage)"""
    return EXCEL_FILE_PATH

def set_blob_key(key: str):
    """Set the blob storage key (cloud storage)"""
    global BLOB_KEY
    BLOB_KEY = key

def get_blob_key() -> Optional[str]:
    """Get the current blob storage key"""
    return BLOB_KEY


def excel_file_exists() -> bool:
    """Check if Excel file exists in any storage"""
    # Check local file first
    if EXCEL_FILE_PATH and os.path.exists(EXCEL_FILE_PATH):
        return True
    
    # Check storage module (blob, local, memory)
    return storage_file_exists()

def _get_workbook() -> Optional[Workbook]:
    """Get workbook from any available storage"""
    print(f"_get_workbook called. EXCEL_FILE_PATH: {EXCEL_FILE_PATH}")
    
    # Try local file first
    if EXCEL_FILE_PATH and os.path.exists(EXCEL_FILE_PATH):
        print(f"Loading from local file: {EXCEL_FILE_PATH}")
        return load_workbook(EXCEL_FILE_PATH)
    
    # Try storage module (blob, local, memory)
    print("Trying storage module")
    wb = get_workbook()
    print(f"Storage module returned: {wb}")
    return wb

def _save_workbook(wb: Workbook) -> bool:
    """Save workbook to storage"""
    # Save to local if path is set
    if EXCEL_FILE_PATH:
        wb.save(EXCEL_FILE_PATH)
        return True
    
    # Use storage module
    return save_workbook(wb)

def ensure_excel_structure(filepath: str):
    """Create Excel file with all required tabs if it doesn't exist"""
    if os.path.exists(filepath):
        # File exists, check if it has all required tabs
        wb = load_workbook(filepath)
        required_tabs = ['Config', 'PWDs', 'Employees', 'Availability']
        created_any = False
        for tab in required_tabs:
            if tab not in wb.sheetnames:
                wb.create_sheet(tab)
                created_any = True
                if tab == 'Config':
                    _init_config_sheet(wb[tab])
                elif tab == 'PWDs':
                    _init_pwds_sheet(wb[tab])
                elif tab == 'Employees':
                    _init_employees_sheet(wb[tab])
                elif tab == 'Availability':
                    _init_availability_sheet(wb[tab])
        if created_any:
            wb.save(filepath)
        wb.close()
        return
    
    # Create new workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Config tab
    config_sheet = wb.create_sheet('Config', 0)
    _init_config_sheet(config_sheet)
    
    # Create PWDs tab (passwords)
    pwds_sheet = wb.create_sheet('PWDs', 1)
    _init_pwds_sheet(pwds_sheet)
    
    # Create Employees tab
    emp_sheet = wb.create_sheet('Employees', 2)
    _init_employees_sheet(emp_sheet)
    
    # Create Availability tab
    avail_sheet = wb.create_sheet('Availability', 3)
    _init_availability_sheet(avail_sheet)
    
    wb.save(filepath)
    wb.close()

def _init_config_sheet(sheet):
    """Initialize Config sheet headers"""
    headers = ['Key', 'Value', 'Description']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # Default config values
    defaults = [
        ['organization', 'IBU', 'Organization name'],
        ['created_date', datetime.now().isoformat(), 'When this file was created'],
        ['last_modified', datetime.now().isoformat(), 'Last modification timestamp'],
        ['version', '1.0', 'System version'],
    ]
    for row_idx, row_data in enumerate(defaults, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 40

def _init_pwds_sheet(sheet):
    """Initialize PWDs sheet for storing manager passwords"""
    headers = ['Employee_ID', 'Employee_Name', 'Password_Hash', 'Role', 'Last_Login']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        sheet.column_dimensions[col_letter].width = 25

def _init_employees_sheet(sheet):
    """Initialize Employees sheet"""
    headers = ['ID', 'Name', 'Email', 'Type', 'Max_Hours', 'Preferences', 'Active', 'Created_At']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='9BBB59', end_color='9BBB59', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        sheet.column_dimensions[col_letter].width = 20

def _init_availability_sheet(sheet):
    """Initialize Availability sheet"""
    headers = ['ID', 'Employee_ID', 'Employee_Name', 'Week_Start', 'Monday', 'Tuesday', 
               'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday', 'Submitted_At', 'Notes',
               'Approved', 'Approved_By', 'Approved_At']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='8064A2', end_color='8064A2', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        sheet.column_dimensions[col_letter].width = 15

# ============ Password Management ============

def hash_password(password: str) -> str:
    """Simple hash for demo - use bcrypt in production"""
    import hashlib
    return hashlib.sha256(password.encode()).hexdigest()[:16]

def set_manager_password(employee_id: str, employee_name: str, password: str):
    """Set password for a manager in the PWDs tab"""
    wb = _get_workbook()
    if not wb:
        raise ValueError("No Excel file available")
    
    sheet = wb['PWDs']
    
    # Check if employee already has password
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == employee_id:
            # Update existing
            sheet.cell(row=row, column=3, value=hash_password(password))
            sheet.cell(row=row, column=5, value=datetime.now().isoformat())
            _save_workbook(wb)
            wb.close()
            return
    
    # Add new entry
    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=1, value=employee_id)
    sheet.cell(row=new_row, column=2, value=employee_name)
    sheet.cell(row=new_row, column=3, value=hash_password(password))
    sheet.cell(row=new_row, column=4, value='manager')
    sheet.cell(row=new_row, column=5, value=datetime.now().isoformat())
    
    _save_workbook(wb)
    wb.close()

def verify_manager_password(employee_id: str, password: str) -> bool:
    """Verify manager password"""
    wb = _get_workbook()
    if not wb:
        return False
    
    try:
        sheet = wb['PWDs']
        
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == employee_id:
                stored_hash = sheet.cell(row=row, column=3).value
                wb.close()
                return stored_hash == hash_password(password)
        
        wb.close()
    except Exception:
        pass
    
    return False

def manager_has_password(employee_id: str) -> bool:
    """Check if manager has set a password"""
    wb = _get_workbook()
    if not wb:
        return False
    
    try:
        sheet = wb['PWDs']
        
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == employee_id:
                wb.close()
                return True
        
        wb.close()
    except Exception:
        pass
    
    return False

# ============ Employee Operations ============

def get_all_employees() -> List[Employee]:
    """Get all employees from Excel"""
    wb = _get_workbook()
    if not wb:
        return []
    
    try:
        if 'Employees' not in wb.sheetnames:
            wb.close()
            return []
        
        sheet = wb['Employees']
        employees = []
        
        for row in range(2, sheet.max_row + 1):
            emp_id = sheet.cell(row=row, column=1).value
            if not emp_id:
                continue
            
            try:
                preferences_str = sheet.cell(row=row, column=6).value or '{}'
                if isinstance(preferences_str, str):
                    preferences = json.loads(preferences_str)
                else:
                    preferences = {}
            except:
                preferences = {}
            
            emp = Employee(
                id=str(emp_id),
                name=str(sheet.cell(row=row, column=2).value or ''),
                email=sheet.cell(row=row, column=3).value,
                employee_type=EmployeeType(sheet.cell(row=row, column=4).value or 'student_worker'),
                max_hours_per_week=int(sheet.cell(row=row, column=5).value or 24),
                preferences=preferences,
                active=sheet.cell(row=row, column=7).value != False and sheet.cell(row=row, column=7).value != 'False',
                created_at=datetime.fromisoformat(str(sheet.cell(row=row, column=8).value)) if sheet.cell(row=row, column=8).value else datetime.now()
            )
            employees.append(emp)
        
        wb.close()
        return employees
    except Exception as e:
        print(f"Error reading employees: {e}")
        return []

def get_employee_by_id(employee_id: str) -> Optional[Employee]:
    """Get employee by ID"""
    employees = get_all_employees()
    return next((e for e in employees if e.id == employee_id), None)

def save_employee(employee: Employee) -> Employee:
    """Save or update employee in Excel"""
    wb = _get_workbook()
    if not wb:
        raise ValueError("No Excel file available")
    
    sheet = wb['Employees']
    
    # Check if employee exists
    found_row = None
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value) == employee.id:
            found_row = row
            break
    
    if not found_row:
        found_row = sheet.max_row + 1
    
    # Write data
    sheet.cell(row=found_row, column=1, value=employee.id)
    sheet.cell(row=found_row, column=2, value=employee.name)
    sheet.cell(row=found_row, column=3, value=employee.email)
    sheet.cell(row=found_row, column=4, value=employee.employee_type.value)
    sheet.cell(row=found_row, column=5, value=employee.max_hours_per_week)
    sheet.cell(row=found_row, column=6, value=json.dumps(employee.preferences))
    sheet.cell(row=found_row, column=7, value=employee.active)
    sheet.cell(row=found_row, column=8, value=employee.created_at.isoformat() if isinstance(employee.created_at, datetime) else str(employee.created_at))
    
    _save_workbook(wb)
    wb.close()
    return employee

def delete_employee(employee_id: str) -> bool:
    """Delete employee from Excel"""
    wb = _get_workbook()
    if not wb:
        return False
    
    sheet = wb['Employees']
    
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value) == employee_id:
            sheet.delete_rows(row)
            _save_workbook(wb)
            wb.close()
            return True
    
    wb.close()
    return False

# ============ Availability Operations ============

def get_availabilities(week_start_date: Optional[date] = None, employee_id: Optional[str] = None) -> List[Availability]:
    """Get availabilities from Excel"""
    wb = _get_workbook()
    if not wb:
        return []
    
    try:
        if 'Availability' not in wb.sheetnames:
            wb.close()
            return []
        
        sheet = wb['Availability']
        availabilities = []
        
        for row in range(2, sheet.max_row + 1):
            avail_id = sheet.cell(row=row, column=1).value
            if not avail_id:
                continue
            
            emp_id = str(sheet.cell(row=row, column=2).value or '')
            week_str = str(sheet.cell(row=row, column=4).value or '')
            
            # Filter by employee if specified
            if employee_id and emp_id != employee_id:
                continue
            
            # Filter by week if specified
            if week_start_date:
                try:
                    row_date = datetime.strptime(week_str, '%Y-%m-%d').date()
                    if row_date != week_start_date:
                        continue
                except:
                    continue
            
            avail = Availability(
                id=str(avail_id),
                employee_id=emp_id,
                week_start_date=datetime.strptime(week_str, '%Y-%m-%d').date() if week_str else date.today(),
                monday=AvailabilityType(str(sheet.cell(row=row, column=5).value or 'blank')),
                tuesday=AvailabilityType(str(sheet.cell(row=row, column=6).value or 'blank')),
                wednesday=AvailabilityType(str(sheet.cell(row=row, column=7).value or 'blank')),
                thursday=AvailabilityType(str(sheet.cell(row=row, column=8).value or 'blank')),
                friday=AvailabilityType(str(sheet.cell(row=row, column=9).value or 'blank')),
                saturday=AvailabilityType(str(sheet.cell(row=row, column=10).value or 'blank')),
                sunday=AvailabilityType(str(sheet.cell(row=row, column=11).value or 'off')),
                submitted_at=datetime.fromisoformat(str(sheet.cell(row=row, column=12).value)) if sheet.cell(row=row, column=12).value else datetime.now(),
                notes=sheet.cell(row=row, column=13).value,
                approved=bool(sheet.cell(row=row, column=14).value),
                approved_by=sheet.cell(row=row, column=15).value,
                approved_at=datetime.fromisoformat(str(sheet.cell(row=row, column=16).value)) if sheet.cell(row=row, column=16).value else None
            )
            availabilities.append(avail)
        
        wb.close()
        return availabilities
    except Exception as e:
        print(f"Error reading availabilities: {e}")
        return []

def get_availability_for_week(employee_id: str, week_start_date: date) -> Optional[Availability]:
    """Get availability for specific employee and week"""
    availabilities = get_availabilities(week_start_date, employee_id)
    return availabilities[0] if availabilities else None

def save_availability(availability: Availability) -> Availability:
    """Save availability to Excel"""
    wb = _get_workbook()
    if not wb:
        raise ValueError("Excel database not configured. Please ask your manager to set up the Excel file.")
    
    sheet = wb['Availability']
    
    # Check if availability exists
    found_row = None
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value) == availability.id:
            found_row = row
            break
        # Also check by employee + week
        if (str(sheet.cell(row=row, column=2).value) == availability.employee_id and
            str(sheet.cell(row=row, column=4).value) == str(availability.week_start_date)):
            found_row = row
            break
    
    if not found_row:
        found_row = sheet.max_row + 1
    
    # Get employee name
    emp = get_employee_by_id(availability.employee_id)
    emp_name = emp.name if emp else ''
    
    # Write data
    sheet.cell(row=found_row, column=1, value=availability.id)
    sheet.cell(row=found_row, column=2, value=availability.employee_id)
    sheet.cell(row=found_row, column=3, value=emp_name)
    sheet.cell(row=found_row, column=4, value=str(availability.week_start_date))
    sheet.cell(row=found_row, column=5, value=availability.monday.value)
    sheet.cell(row=found_row, column=6, value=availability.tuesday.value)
    sheet.cell(row=found_row, column=7, value=availability.wednesday.value)
    sheet.cell(row=found_row, column=8, value=availability.thursday.value)
    sheet.cell(row=found_row, column=9, value=availability.friday.value)
    sheet.cell(row=found_row, column=10, value=availability.saturday.value)
    sheet.cell(row=found_row, column=11, value=availability.sunday.value)
    sheet.cell(row=found_row, column=12, value=availability.submitted_at.isoformat() if isinstance(availability.submitted_at, datetime) else str(availability.submitted_at))
    sheet.cell(row=found_row, column=13, value=availability.notes)
    sheet.cell(row=found_row, column=14, value=availability.approved)
    sheet.cell(row=found_row, column=15, value=availability.approved_by)
    sheet.cell(row=found_row, column=16, value=availability.approved_at.isoformat() if availability.approved_at else None)
    
    _save_workbook(wb)
    wb.close()
    return availability

# ============ Weekly Schedule Operations ============

def get_schedule_sheet_name(week_start_date: date) -> str:
    """Generate tab name for a week"""
    return f"Schedule_{week_start_date.strftime('%Y_%m_%d')}"

def ensure_schedule_sheet(wb: Workbook, week_start_date: date):
    """Create schedule sheet if it doesn't exist"""
    sheet_name = get_schedule_sheet_name(week_start_date)
    
    if sheet_name not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_name)
        headers = ['Shift_ID', 'Employee_ID', 'Employee_Name', 'Day', 'Start_Time', 
                   'End_Time', 'Job_Type', 'Floor', 'Hours', 'Is_Event', 'Event_Name']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            sheet.column_dimensions[col_letter].width = 15
    
    return wb[sheet_name]

def get_schedule_by_week(week_start_date: date) -> Optional[WeeklySchedule]:
    """Get schedule for a specific week from its tab"""
    if not EXCEL_FILE_PATH or not os.path.exists(EXCEL_FILE_PATH):
        return None
    
    try:
        wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
        sheet_name = get_schedule_sheet_name(week_start_date)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        
        sheet = wb[sheet_name]
        shifts = []
        total_hours: Dict[str, float] = {}
        
        for row in range(2, sheet.max_row + 1):
            shift_id = sheet.cell(row=row, column=1).value
            if not shift_id:
                continue
            
            emp_id = str(sheet.cell(row=row, column=2).value or '')
            hours = float(sheet.cell(row=row, column=9).value or 0)
            
            shift = Shift(
                id=str(shift_id),
                employee_id=emp_id,
                day_of_week=str(sheet.cell(row=row, column=4).value or 'monday'),
                start_time=str(sheet.cell(row=row, column=5).value or '09:00'),
                end_time=str(sheet.cell(row=row, column=6).value or '17:00'),
                job_type=JobType(str(sheet.cell(row=row, column=7).value or 'desk')),
                floor=Floor(str(sheet.cell(row=row, column=8).value)) if sheet.cell(row=row, column=8).value else None,
                hours=hours,
                is_event=sheet.cell(row=row, column=10).value == True or sheet.cell(row=row, column=10).value == 'True',
                event_name=sheet.cell(row=row, column=11).value
            )
            shifts.append(shift)
            
            # Accumulate hours
            if emp_id in total_hours:
                total_hours[emp_id] += hours
            else:
                total_hours[emp_id] = hours
        
        wb.close()
        
        return WeeklySchedule(
            id=f"sched_{week_start_date.isoformat()}",
            week_start_date=week_start_date,
            shifts=shifts,
            total_hours=total_hours,
            created_at=datetime.now(),
            updated_at=datetime.now(),
            status='published' if shifts else 'draft'
        )
    except Exception as e:
        print(f"Error reading schedule: {e}")
        return None

def save_schedule(schedule: WeeklySchedule) -> WeeklySchedule:
    """Save schedule to Excel in its own tab"""
    wb = _get_workbook()
    if not wb:
        raise ValueError("No Excel file available")
    
    sheet = ensure_schedule_sheet(wb, schedule.week_start_date)
    
    # Clear existing data (keep header)
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    
    # Write all shifts
    for shift in schedule.shifts:
        emp = get_employee_by_id(shift.employee_id)
        emp_name = emp.name if emp else ''
        
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=shift.id)
        sheet.cell(row=row, column=2, value=shift.employee_id)
        sheet.cell(row=row, column=3, value=emp_name)
        sheet.cell(row=row, column=4, value=shift.day_of_week)
        sheet.cell(row=row, column=5, value=shift.start_time)
        sheet.cell(row=row, column=6, value=shift.end_time)
        sheet.cell(row=row, column=7, value=shift.job_type.value)
        sheet.cell(row=row, column=8, value=shift.floor.value if shift.floor else None)
        sheet.cell(row=row, column=9, value=shift.hours)
        sheet.cell(row=row, column=10, value=shift.is_event)
        sheet.cell(row=row, column=11, value=shift.event_name)
    
    _save_workbook(wb)
    wb.close()
    return schedule

def get_all_schedules() -> List[WeeklySchedule]:
    """Get all schedules from week tabs"""
    wb = _get_workbook()
    if not wb:
        return []
    
    try:
        schedules = []
        
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('Schedule_'):
                try:
                    # Parse date from sheet name
                    date_str = sheet_name.replace('Schedule_', '').replace('_', '-')
                    week_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    
                    # Read the schedule
                    sheet = wb[sheet_name]
                    shifts = []
                    total_hours: Dict[str, float] = {}
                    
                    for row in range(2, sheet.max_row + 1):
                        shift_id = sheet.cell(row=row, column=1).value
                        if not shift_id:
                            continue
                        
                        emp_id = str(sheet.cell(row=row, column=2).value or '')
                        hours = float(sheet.cell(row=row, column=9).value or 0)
                        
                        shift = Shift(
                            id=str(shift_id),
                            employee_id=emp_id,
                            day_of_week=str(sheet.cell(row=row, column=4).value or 'monday'),
                            start_time=str(sheet.cell(row=row, column=5).value or '09:00'),
                            end_time=str(sheet.cell(row=row, column=6).value or '17:00'),
                            job_type=JobType(str(sheet.cell(row=row, column=7).value or 'desk')),
                            floor=Floor(str(sheet.cell(row=row, column=8).value)) if sheet.cell(row=row, column=8).value else None,
                            hours=hours,
                            is_event=sheet.cell(row=row, column=10).value == True,
                            event_name=sheet.cell(row=row, column=11).value
                        )
                        shifts.append(shift)
                        
                        if emp_id in total_hours:
                            total_hours[emp_id] += hours
                        else:
                            total_hours[emp_id] = hours
                    
                    schedule = WeeklySchedule(
                        id=f"sched_{week_date.isoformat()}",
                        week_start_date=week_date,
                        shifts=shifts,
                        total_hours=total_hours,
                        created_at=datetime.now(),
                        updated_at=datetime.now(),
                        status='published' if shifts else 'draft'
                    )
                    schedules.append(schedule)
                except Exception as e:
                    print(f"Error reading schedule tab {sheet_name}: {e}")
        
        wb.close()
        return sorted(schedules, key=lambda x: x.week_start_date, reverse=True)
    except Exception as e:
        print(f"Error reading all schedules: {e}")
        return []

def delete_schedule(week_start_date: date) -> bool:
    """Delete a schedule by removing its tab"""
    wb = _get_workbook()
    if not wb:
        return False
    
    sheet_name = get_schedule_sheet_name(week_start_date)
    
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        _save_workbook(wb)
        wb.close()
        return True
    
    wb.close()
    return False

# ============ Floor Coverage ============

def get_floor_coverage(floor: str, day_of_week: str, time_slot: str, week_start_date: date) -> Dict:
    """Get employees working on a floor at a specific time"""
    schedule = get_schedule_by_week(week_start_date)
    if not schedule:
        return {"floor": floor, "day_of_week": day_of_week, "time_slot": time_slot, "employee_count": 0, "employees": []}
    
    # Filter shifts by floor and day
    matching_shifts = [
        s for s in schedule.shifts 
        if s.day_of_week.lower() == day_of_week.lower() 
        and s.floor 
        and s.floor.value.lower() == floor.lower()
    ]
    
    employees = []
    for shift in matching_shifts:
        emp = get_employee_by_id(shift.employee_id)
        if emp:
            employees.append({
                "id": emp.id,
                "name": emp.name,
                "job_type": shift.job_type.value,
                "hours": shift.hours,
                "start_time": shift.start_time,
                "end_time": shift.end_time
            })
    
    return {
        "floor": floor,
        "day_of_week": day_of_week,
        "time_slot": time_slot,
        "employee_count": len(employees),
        "employees": employees
    }

# ============ System Config ============

def get_system_config() -> Dict[str, Any]:
    """Get system config from Excel"""
    wb = _get_workbook()
    if not wb:
        return {}
    
    try:
        if 'Config' not in wb.sheetnames:
            wb.close()
            return {}
        
        sheet = wb['Config']
        config = {}
        
        for row in range(2, sheet.max_row + 1):
            key = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=2).value
            if key:
                config[key] = value
        
        wb.close()
        return config
    except Exception as e:
        print(f"Error reading config: {e}")
        return {}

def save_system_config(config: Dict[str, Any]):
    """Save system config to Excel"""
    wb = _get_workbook()
    if not wb:
        raise ValueError("No Excel file available")
    
    if 'Config' not in wb.sheetnames:
        wb.create_sheet('Config', 0)
        _init_config_sheet(wb['Config'])
    
    sheet = wb['Config']
    
    # Clear existing data (keep header)
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    
    # Write new config
    for key, value in config.items():
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=str(value))
    
    _save_workbook(wb)
    wb.close()

# ============ Sample Data Initialization ============

def initialize_sample_employees():
    """Initialize the employee list when creating a new Excel file"""
    wb = _get_workbook()
    if not wb:
        return
    
    # Check if employees already exist
    existing = get_all_employees()
    if existing:
        return
    
    sample_employees = [
        Employee(id="emp1", name="Fran", employee_type=EmployeeType.MANAGER, max_hours_per_week=80),
        Employee(id="emp2", name="Aashima", employee_type=EmployeeType.MANAGER, max_hours_per_week=80),
        Employee(id="emp3", name="Mickaela C", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp4", name="Kavya C", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp5", name="Pablo 2", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp6", name="Viviana 3", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp7", name="Anastasia 3", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp8", name="MEG 3", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp9", name="Achal 2", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp10", name="PRIYANKA 2", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp11", name="Arta C", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp12", name="Taran C", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp13", name="Sagar C", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp14", name="Itshan 3", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp15", name="Arnob 2", employee_type=EmployeeType.STUDENT_WORKER, max_hours_per_week=24),
        Employee(id="emp16", name="Intern-Levi 7", employee_type=EmployeeType.INTERN, max_hours_per_week=15),
        Employee(id="emp17", name="Intern-Himanshu 7", employee_type=EmployeeType.INTERN, max_hours_per_week=15),
    ]
    
    for emp in sample_employees:
        save_employee(emp)

# ============ Initialization ============

def initialize_from_excel():
    """Initialize data from Excel file - call this on startup if file is set"""
    if not EXCEL_FILE_PATH or not os.path.exists(EXCEL_FILE_PATH):
        return
    
    # Ensure structure is correct
    ensure_excel_structure(EXCEL_FILE_PATH)

def get_all_week_schedule_dates() -> List[date]:
    """Get list of all weeks that have schedule tabs"""
    if not EXCEL_FILE_PATH or not os.path.exists(EXCEL_FILE_PATH):
        return []
    
    try:
        wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
        dates = []
        
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('Schedule_'):
                try:
                    date_str = sheet_name.replace('Schedule_', '').replace('_', '-')
                    week_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    dates.append(week_date)
                except:
                    pass
        
        wb.close()
        return sorted(dates, reverse=True)
    except:
        return []
