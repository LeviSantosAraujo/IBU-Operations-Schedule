#!/usr/bin/env python3
"""Sync employees from Excel to JSON"""
import sys
sys.path.insert(0, '/Users/levisantosaraujo/Downloads/Python/Schedule Sheet IBU/backend')

from openpyxl import load_workbook
from datetime import datetime
import json

EXCEL_PATH = '/Users/levisantosaraujo/Downloads/Python/Schedule Sheet IBU/backend/uploads/ibu_schedule.xlsx'
JSON_PATH = '/Users/levisantosaraujo/Downloads/Python/Schedule Sheet IBU/backend/data/employees.json'

def sync_employees():
    wb = load_workbook(EXCEL_PATH)
    
    if 'Employees' not in wb.sheetnames:
        print("No Employees sheet found")
        return
    
    sheet = wb['Employees']
    employees = []
    
    for row in range(2, sheet.max_row + 1):
        emp_id = sheet.cell(row=row, column=1).value
        if not emp_id:
            continue
        
        emp_name = str(sheet.cell(row=row, column=2).value or '')
        emp_type = str(sheet.cell(row=row, column=4).value or 'employee')
        
        # Normalize legacy role names
        if emp_type == 'student_worker':
            emp_type = 'employee'
        
        # Detect Intern prefix and set role accordingly
        if emp_name.lower().startswith('intern'):
            emp_type = 'intern'
        
        emp = {
            'id': str(emp_id),
            'name': emp_name,
            'email': sheet.cell(row=row, column=3).value,
            'employee_type': emp_type,
            'max_hours_per_week': int(sheet.cell(row=row, column=5).value or 24),
            'preferences': {},
            'active': sheet.cell(row=row, column=7).value != False and sheet.cell(row=row, column=7).value != 'False',
            'created_at': str(sheet.cell(row=row, column=8).value) if sheet.cell(row=row, column=8).value else str(datetime.now())
        }
        employees.append(emp)
    
    wb.close()
    
    # Save to JSON
    with open(JSON_PATH, 'w') as f:
        json.dump(employees, f, indent=2)
    
    print(f"Synced {len(employees)} employees from Excel to JSON")

if __name__ == '__main__':
    sync_employees()
